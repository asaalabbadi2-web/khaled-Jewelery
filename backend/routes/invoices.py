"""Invoice domain routes — invoices_bp registered under /api in app.py."""
from __future__ import annotations

import io
import json
import logging
import os
import zipfile
from collections import defaultdict
from datetime import date, datetime, time, timedelta

from flask import Blueprint, current_app, g, jsonify, request, send_file
from sqlalchemy import Integer, String, and_, case, cast, func, not_, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload

from models import (
    db,
    Account,
    AccountingMapping,
    AuditLog,
    Category,
    Customer,
    Employee,
    InventoryCostingConfig,
    Invoice,
    InvoiceItem,
    InvoiceKaratLine,
    InvoicePayment,
    Item,
    JournalEntry,
    JournalEntryLine,
    Office,
    OfficeReservation,
    PaymentMethod,
    SafeBox,
    SafeBoxTransaction,
    SettlementLine,
    Settings,
    Supplier,
    SupplierGoldTransaction,
    User,
    Voucher,
    VoucherAccountLine,
)

try:
    from backend.config import REQUIRE_AUTH_FOR_INVOICE_CREATE, WEIGHT_SUPPORT_ACCOUNTS
except ImportError:
    from config import REQUIRE_AUTH_FOR_INVOICE_CREATE, WEIGHT_SUPPORT_ACCOUNTS

from allocation_service import AllocationService
from core.database import _db_has_column
from core.dates import _parse_iso_date
from core.number_helpers import _coerce_float
from core.responses import _wrap_api_exceptions
from auth_decorators import get_current_user, require_admin, require_auth, require_permission
from category_weight_tracking import (
    get_category_weight_balances,
    record_category_weight_movements_for_invoice_payload,
)
from code_generator import (
    generate_barcode_from_item_code,
    generate_item_code,
    validate_item_code,
)
from dual_system_helpers import (
    create_dual_journal_entry,
    get_account_balances,
    link_memo_accounts_helper,
    verify_dual_balance,
)
from gold_costing_service import GoldCostingService, ScrapCostingService
from office_supplier_service import ensure_office_supplier
from party_account_service import ensure_customer_accounts, ensure_supplier_accounts
from services.journals import create_wage_weight_release_journal
from services.weight_execution import list_weight_profiles, resolve_weight_profile
from settlement_state_service import get_settled_amounts, is_locked
from utils import normalize_number

from pricing.karat_service import convert_from_main_karat, convert_to_main_karat, get_main_karat
from accounting.voucher_engine import (
    _generate_journal_entry_number,
    generate_voucher_number,
    create_journal_entry_from_voucher,
    _append_safe_transactions_for_voucher,
)
from accounting.mappings import DEFAULT_MAPPING_OPERATION_TYPE, get_account_id_by_number, get_account_id_for_mapping
from accounting.weight_closing import _load_weight_closing_settings
from accounting.safe_boxes import _ensure_safe_box_transactions_for_invoice_je
from accounting.wages import _ensure_manufacturing_wage_expense_account, _ensure_gold24k_commission_revenue_account
from accounting.inventory import get_inventory_average_cost
from accounting.balances import _recalculate_account_balances_for_accounts
from routes import (
    get_current_gold_price,
    _next_invoice_type_id,
    create_item_from_invoice_payload,
    InlineItemCreationError,
    _resolve_inventory_account_id_for_invoice,
    _get_inventory_account_by_karat,
    _get_manufacturing_wage_mode,
    _get_manufacturing_wage_inventory_account_id,
    validate_bridge_account_balance,
    _try_process_due_auto_clearing_settlements,
    _ensure_weight_tracking_account,
    _upsert_weight_closing_order,
    _weight_kwargs_for_karat,
    _weight_kwargs_from_map,
)

invoices_bp = Blueprint('invoices', __name__)

@invoices_bp.route('/invoices/pending-post', methods=['GET'])
def pending_post_invoices():
    """قائمة مختصرة بالفواتير غير المرحّلة — للـ Dialog في الرئيسية"""
    try:
        limit = request.args.get('limit', 10, type=int)
        limit = max(1, min(limit, 50))  # clamp to safe range

        base_q = Invoice.query.filter(
            Invoice.is_posted.is_(False),
            Invoice.status != 'rejected',
        )

        total = base_q.count()

        invoices = (
            base_q
            .order_by(Invoice.date.desc(), Invoice.id.desc())
            .limit(limit)
            .all()
        )

        result = []
        for inv in invoices:
            # party name
            party_name = ''
            if inv.customer:
                party_name = inv.customer.name or ''
            elif inv.supplier:
                party_name = inv.supplier.name or ''

            # creator: use posted_by if available, else fallback to employee name
            creator = ''
            if getattr(inv, 'posted_by', None):
                creator = inv.posted_by
            elif getattr(inv, 'employee', None) and inv.employee.name:
                creator = inv.employee.name

            inv_type = inv.invoice_type or ''
            _type_labels = {
                'شراء':           'فاتورة شراء كسر من مكتب التسكير',
                'purchase':       'فاتورة شراء كسر من مكتب التسكير',
                'بيع':            'فاتورة بيع',
                'sale':           'فاتورة بيع',
                'مرتجع بيع':     'مرتجع بيع',
                'sales_return':   'مرتجع بيع',
                'شراء من عميل':  'شراء كسر من عميل',
                'scrap_purchase': 'شراء كسر من عميل',
                'مرتجع شراء':    'مرتجع شراء كسر',
                'purchase_return':'مرتجع شراء كسر',
                'scrap_sale':     'بيع كسر',
            }
            type_label = _type_labels.get(inv_type, 'فاتورة')
            stored_reason = getattr(inv, 'pending_approval_reason', None) or ''
            if stored_reason:
                approval_reason = f'{type_label} — {stored_reason}'
            else:
                approval_reason = f'{type_label} — بانتظار اعتماد الترحيل'

            result.append({
                'id': inv.id,
                'invoice_number': inv.invoice_number,
                'invoice_type': inv_type,
                'total_amount': float(inv.total or 0),
                'party_name': party_name,
                'created_by_name': creator,
                'created_at': inv.date.isoformat() if inv.date else None,
                'approval_reason_message': approval_reason,
            })

        return jsonify({
            'invoices': result,
            'total': total,
            'showing': len(result),
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'invoices': [],
            'total': 0,
            'showing': 0,
            'error': str(e),
        }), 500

@invoices_bp.route('/pending-actions', methods=['GET'])
def pending_actions():
    """نقطة موحّدة للمعلّقات: حجوزات بانتظار التسوية + فواتير بانتظار الاعتماد"""
    try:
        # ── حجوزات بانتظار التسوية ──────────────────────────────────────────
        res_query = (
            OfficeReservation.query
            .options(joinedload(OfficeReservation.office))
            .filter(
                OfficeReservation.purchase_invoice_id.is_(None),
                OfficeReservation.status.in_(['approved', 'pending']),
            )
            .order_by(OfficeReservation.reservation_date.desc())
            .limit(50)
        )
        reservations = res_query.all()
        pending_reservations = []
        for r in reservations:
            office_name = r.office.name if r.office else ''
            pending_reservations.append({
                'id': r.id,
                'reservation_code': r.reservation_code,
                'office_id': r.office_id,
                'office_name': office_name,
                'karat': r.karat,
                'weight_main_karat': float(r.weight_main_karat or 0),
                'weight_remaining_main_karat': float(r.weight_remaining_main_karat or 0),
                'price_per_gram': float(r.price_per_gram or 0),
                'total_amount': float(r.total_amount or 0),
                'paid_amount': float(r.paid_amount or 0),
                'payment_status': r.payment_status,
                'status': r.status,
                'contact_person': r.contact_person or '',
                'contact_phone': r.contact_phone or '',
                'notes': r.notes or '',
                'reservation_date': r.reservation_date.isoformat() if r.reservation_date else None,
            })

        # ── فواتير بانتظار الاعتماد ──────────────────────────────────────────
        inv_query = Invoice.query.filter(
            Invoice.is_posted.is_(False),
            Invoice.status != 'rejected',
        )
        total_invoices = inv_query.count()
        invoices = inv_query.order_by(Invoice.date.desc(), Invoice.id.desc()).limit(50).all()
        pending_invoices = []
        for inv in invoices:
            party_name = ''
            if inv.customer:
                party_name = inv.customer.name or ''
            elif inv.supplier:
                party_name = inv.supplier.name or ''

            creator = ''
            if getattr(inv, 'posted_by', None):
                creator = inv.posted_by
            elif getattr(inv, 'employee', None) and inv.employee.name:
                creator = inv.employee.name

            inv_type = inv.invoice_type or ''
            _type_labels = {
                'شراء':           'فاتورة شراء كسر من مكتب التسكير',
                'purchase':       'فاتورة شراء كسر من مكتب التسكير',
                'بيع':            'فاتورة بيع',
                'sale':           'فاتورة بيع',
                'مرتجع بيع':     'مرتجع بيع',
                'sales_return':   'مرتجع بيع',
                'شراء من عميل':  'شراء كسر من عميل',
                'scrap_purchase': 'شراء كسر من عميل',
                'مرتجع شراء':    'مرتجع شراء كسر',
                'purchase_return':'مرتجع شراء كسر',
                'scrap_sale':     'بيع كسر',
            }
            type_label = _type_labels.get(inv_type, 'فاتورة')
            stored_reason = getattr(inv, 'pending_approval_reason', None) or ''
            if stored_reason:
                approval_reason = f'{type_label} — {stored_reason}'
            else:
                approval_reason = f'{type_label} — بانتظار اعتماد الترحيل'

            pending_invoices.append({
                'id': inv.id,
                'invoice_number': inv.invoice_number,
                'invoice_type': inv_type,
                'total_amount': float(inv.total or 0),
                'party_name': party_name,
                'created_by_name': creator,
                'created_at': inv.date.isoformat() if inv.date else None,
                'approval_reason_message': approval_reason,
            })

        return jsonify({
            'pending_reservations': pending_reservations,
            'pending_invoices': pending_invoices,
            'total_pending_reservations': len(pending_reservations),
            'total_pending_invoices': total_invoices,
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'pending_reservations': [],
            'pending_invoices': [],
            'total_pending_reservations': 0,
            'total_pending_invoices': 0,
            'error': str(e),
        }), 500

@invoices_bp.route('/invoices', methods=['GET'])
def get_invoices():
    # Pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    # Sorting parameters
    sort_by = request.args.get('sort_by', 'date')
    sort_order = request.args.get('sort_order', 'desc')

    # Filtering parameters
    search = request.args.get('search')
    search_type = (request.args.get('search_type') or 'all').strip().lower()
    status = request.args.get('status')
    invoice_type = request.args.get('invoice_type')
    invoice_types = request.args.get('invoice_types')
    invoice_group = request.args.get('invoice_group')
    party = request.args.get('party')
    creator = request.args.get('creator')
    employee_id = request.args.get('employee_id', type=int)
    karat = request.args.get('karat')
    gold_type = request.args.get('gold_type')
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')

    # Base query
    query = Invoice.query

    customer_joined = False
    supplier_joined = False
    employee_joined = False

    def _ensure_party_outerjoins():
        nonlocal query, customer_joined, supplier_joined
        if not customer_joined:
            query = query.outerjoin(Customer, Invoice.customer_id == Customer.id)
            customer_joined = True
        if not supplier_joined:
            query = query.outerjoin(Supplier, Invoice.supplier_id == Supplier.id)
            supplier_joined = True

    def _ensure_employee_outerjoin():
        nonlocal query, employee_joined
        if not employee_joined:
            query = query.outerjoin(Employee, Invoice.employee_id == Employee.id)
            employee_joined = True

    def _is_return_clause():
        return or_(
            Invoice.invoice_type.ilike('%مرتجع%'),
            Invoice.invoice_type.ilike('%return%'),
        )

    def _is_sale_clause():
        return and_(
            not_(_is_return_clause()),
            or_(
                Invoice.invoice_type.ilike('%بيع%'),
                Invoice.invoice_type.ilike('%sale%'),
                Invoice.invoice_type == 'sell',
            ),
        )

    def _is_purchase_clause():
        return and_(
            not_(_is_return_clause()),
            or_(
                Invoice.invoice_type.ilike('%شراء%'),
                Invoice.invoice_type.ilike('%purchase%'),
                Invoice.invoice_type == 'buy',
            ),
        )

    def _is_customer_purchase_clause():
        return and_(
            _is_purchase_clause(),
            or_(
                Invoice.customer_id.isnot(None),
                Invoice.invoice_type.in_(['شراء من عميل', 'شراء خردة', 'شراء مستعمل']),
            ),
            Invoice.supplier_id.is_(None),
        )

    def _is_supplier_purchase_clause():
        return and_(
            _is_purchase_clause(),
            or_(
                Invoice.supplier_id.isnot(None),
                Invoice.invoice_type == 'شراء',
            ),
        )

    creator_name_cache = {}

    def _resolve_invoice_creator_name(invoice):
        try:
            if getattr(invoice, 'employee', None) and getattr(invoice.employee, 'name', None):
                return str(invoice.employee.name).strip()
        except Exception:
            pass

        posted_by_raw = str(getattr(invoice, 'posted_by', '') or '').strip()
        if not posted_by_raw:
            return None

        posted_by_key = posted_by_raw.lower()
        if posted_by_key in creator_name_cache:
            return creator_name_cache[posted_by_key]

        resolved_name = None
        try:
            from models import AppUser

            app_user = AppUser.query.filter(
                func.lower(func.trim(AppUser.username)) == posted_by_key
            ).first()
            if app_user and getattr(app_user, 'employee', None) and getattr(app_user.employee, 'name', None):
                resolved_name = str(app_user.employee.name).strip()
        except Exception:
            resolved_name = None

        creator_name_cache[posted_by_key] = resolved_name or posted_by_raw
        return creator_name_cache[posted_by_key]

    # Filtering
    if search:
        search = (search or '').strip()
        if search:
            like = f'%{search}%'
            clauses = []

            if search_type in ('all', 'name'):
                _ensure_party_outerjoins()
                clauses.extend([
                    Customer.name.ilike(like),
                    Supplier.name.ilike(like),
                ])

            if search_type == 'all':
                clauses.append(Invoice.invoice_type.ilike(like))

            normalized_numeric_search = search.replace(',', '').strip()
            try:
                search_number = float(normalized_numeric_search)
                if search_type in ('all', 'amount'):
                    clauses.append(
                        func.abs(func.coalesce(Invoice.total, 0) - search_number) < 0.0001
                    )
                if search_type in ('all', 'weight'):
                    clauses.append(
                        func.abs(func.coalesce(Invoice.total_weight, 0) - search_number) < 0.0001
                    )
            except Exception:
                pass

            if search_type in ('all', 'number'):
                try:
                    search_int = int(search)
                    clauses.extend([
                        Invoice.id == search_int,
                        Invoice.invoice_type_id == search_int,
                    ])
                except Exception:
                    tail = search.split('-')[-1].strip()
                    try:
                        search_tail_int = int(tail)
                        clauses.extend([
                            Invoice.id == search_tail_int,
                            Invoice.invoice_type_id == search_tail_int,
                        ])
                    except Exception:
                        pass

            if clauses:
                query = query.filter(or_(*clauses))
            else:
                query = query.filter(Invoice.id == -1)

    if status and status != 'all':
        normalized_status = (status or '').strip().lower()
        if normalized_status in ('paid_full', 'paid'):
            query = query.filter(Invoice.status == 'paid')
        elif normalized_status == 'remaining':
            query = query.filter(Invoice.status.in_(['unpaid', 'partially_paid']))
        else:
            query = query.filter(Invoice.status == status)

    if invoice_group and invoice_group not in ('all', 'الكل'):
        group_value = (invoice_group or '').strip().lower()
        if group_value == 'sales':
            query = query.filter(_is_sale_clause())
        elif group_value == 'customer_purchase':
            query = query.filter(_is_customer_purchase_clause())
        elif group_value == 'supplier_purchase':
            query = query.filter(_is_supplier_purchase_clause())
        elif group_value == 'returns':
            query = query.filter(_is_return_clause())

    def _expand_invoice_type_aliases(values):
        alias_map = {
            'بيع': {'بيع', 'sell', 'sale'},
            'sell': {'بيع', 'sell', 'sale'},
            'sale': {'بيع', 'sell', 'sale'},
            'شراء من عميل': {
                'شراء من عميل',
                'buy',
                'customer purchase',
                'purchase from customer',
                'شراء خردة',
                'شراء مستعمل',
            },
            'buy': {
                'شراء من عميل',
                'buy',
                'customer purchase',
                'purchase from customer',
                'شراء خردة',
                'شراء مستعمل',
            },
            'شراء': {'شراء', 'شراء من مورد'},
            'شراء من مورد': {'شراء', 'شراء من مورد'},
            'supplier purchase': {'شراء', 'شراء من مورد', 'supplier purchase'},
            'purchase': {'شراء', 'شراء من مورد', 'supplier purchase'},
            'مرتجع بيع': {'مرتجع بيع', 'sales return', 'sale return'},
            'sales return': {'مرتجع بيع', 'sales return', 'sale return'},
            'sale return': {'مرتجع بيع', 'sales return', 'sale return'},
            'مرتجع شراء': {
                'مرتجع شراء',
                'مرتجع شراء من عميل',
                'purchase return',
                'customer purchase return',
            },
            'مرتجع شراء من عميل': {
                'مرتجع شراء',
                'مرتجع شراء من عميل',
                'purchase return',
                'customer purchase return',
            },
            'purchase return': {
                'مرتجع شراء',
                'مرتجع شراء من عميل',
                'purchase return',
                'customer purchase return',
            },
            'مرتجع شراء (مورد)': {'مرتجع شراء (مورد)', 'مرتجع شراء من مورد'},
            'مرتجع شراء من مورد': {'مرتجع شراء (مورد)', 'مرتجع شراء من مورد'},
            'supplier purchase return': {
                'مرتجع شراء (مورد)',
                'مرتجع شراء من مورد',
                'supplier purchase return',
            },
        }
        expanded = []
        seen = set()
        for raw_value in values:
            value = (raw_value or '').strip()
            if not value:
                continue
            for alias in alias_map.get(value, {value}):
                if alias not in seen:
                    seen.add(alias)
                    expanded.append(alias)
        return expanded

    if invoice_type and invoice_type not in ('الكل', 'all'):
        expanded_invoice_types = _expand_invoice_type_aliases([invoice_type])
        if len(expanded_invoice_types) == 1:
            query = query.filter(Invoice.invoice_type == expanded_invoice_types[0])
        elif expanded_invoice_types:
            query = query.filter(Invoice.invoice_type.in_(expanded_invoice_types))

    if invoice_types:
        parsed_invoice_types = [
            value.strip()
            for value in (invoice_types or '').split(',')
            if value.strip()
        ]
        if parsed_invoice_types:
            query = query.filter(
                Invoice.invoice_type.in_(_expand_invoice_type_aliases(parsed_invoice_types))
            )

    if gold_type and gold_type not in ('all', 'الكل'):
        normalized_gold_type = (gold_type or '').strip().lower()
        query = query.filter(func.lower(func.coalesce(Invoice.gold_type, 'new')) == normalized_gold_type)

    if party:
        party = (party or '').strip()
        if party:
            _ensure_party_outerjoins()
            query = query.filter(
                or_(
                    Customer.name.ilike(f'%{party}%'),
                    Supplier.name.ilike(f'%{party}%'),
                )
            )

    if employee_id:
        query = query.filter(Invoice.employee_id == employee_id)

    creator_source_query = query
    available_creators = []
    try:
        creator_rows = creator_source_query.order_by(None).all()
        seen_creator_names = set()
        creator_filter_value = (creator or '').strip().lower()
        matching_creator_ids = []

        for invoice in creator_rows:
            creator_name = _resolve_invoice_creator_name(invoice)
            if creator_name:
                creator_name = creator_name.strip()
            if creator_name and creator_name not in seen_creator_names:
                seen_creator_names.add(creator_name)
                available_creators.append({'name': creator_name})
            if creator_filter_value and creator_name and creator_name.lower() == creator_filter_value:
                matching_creator_ids.append(int(invoice.id))

        available_creators.sort(key=lambda entry: entry['name'])

        if creator_filter_value:
            if matching_creator_ids:
                query = query.filter(Invoice.id.in_(matching_creator_ids))
            else:
                query = query.filter(Invoice.id == -1)
    except Exception:
        available_creators = []

    if karat:
        try:
            karat_value = float(karat)
            query = query.filter(
                or_(
                    Invoice.items.any(InvoiceItem.karat == karat_value),
                    Invoice.karat_lines.any(InvoiceKaratLine.karat == karat_value),
                )
            )
        except Exception:
            pass

    def _parse_iso_date(v: str):
        if not v:
            return None
        cleaned = v.strip()
        if cleaned.endswith('Z'):
            cleaned = cleaned[:-1] + '+00:00'
        return datetime.fromisoformat(cleaned)

    if date_from_str:
        date_from = _parse_iso_date(date_from_str)
        if date_from is not None:
            query = query.filter(Invoice.date >= date_from)

    if date_to_str:
        date_to = _parse_iso_date(date_to_str)
        if date_to is not None:
            # Date pickers send end-date at 00:00:00 (midnight), which would exclude
            # every invoice created that day. Extend to 23:59:59.999999 when no
            # explicit time was provided.
            if date_to.hour == 0 and date_to.minute == 0 and date_to.second == 0 and date_to.microsecond == 0:
                date_to = date_to.replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(Invoice.date <= date_to)

    # Sorting
    created_sort_expr = Invoice.date

    if sort_by == 'date':
        if sort_order == 'desc':
            query = query.order_by(created_sort_expr.desc(), Invoice.id.desc())
        else:
            query = query.order_by(created_sort_expr.asc(), Invoice.id.asc())
        order = None
    elif sort_by == 'customer':
        _ensure_party_outerjoins()
        party_name = func.coalesce(Customer.name, Supplier.name, '')
        if sort_order == 'desc':
            query = query.order_by(party_name.desc(), created_sort_expr.desc(), Invoice.id.desc())
        else:
            query = query.order_by(party_name.asc(), created_sort_expr.desc(), Invoice.id.desc())
        order = None
    elif sort_by == 'amount':
        if sort_order == 'desc':
            query = query.order_by(Invoice.total.desc(), created_sort_expr.desc(), Invoice.id.desc())
        else:
            query = query.order_by(Invoice.total.asc(), created_sort_expr.desc(), Invoice.id.desc())
        order = None
    elif sort_by == 'weight':
        if sort_order == 'desc':
            query = query.order_by(Invoice.total_weight.desc(), created_sort_expr.desc(), Invoice.id.desc())
        else:
            query = query.order_by(Invoice.total_weight.asc(), created_sort_expr.desc(), Invoice.id.asc())
        order = None
    elif sort_by == 'status':
        if sort_order == 'desc':
            query = query.order_by(Invoice.status.desc(), created_sort_expr.desc(), Invoice.id.desc())
        else:
            query = query.order_by(Invoice.status.asc(), created_sort_expr.desc(), Invoice.id.asc())
        order = None
    elif sort_by == 'type':
        if sort_order == 'desc':
            query = query.order_by(Invoice.gold_type.desc(), Invoice.invoice_type.desc(), created_sort_expr.desc(), Invoice.id.desc())
        else:
            query = query.order_by(Invoice.gold_type.asc(), Invoice.invoice_type.asc(), created_sort_expr.desc(), Invoice.id.asc())
        order = None
    elif sort_by == 'number':
        if sort_order == 'desc':
            query = query.order_by(Invoice.invoice_type.desc(), Invoice.invoice_type_id.desc(), created_sort_expr.desc(), Invoice.id.desc())
        else:
            query = query.order_by(Invoice.invoice_type.asc(), Invoice.invoice_type_id.asc(), created_sort_expr.asc(), Invoice.id.asc())
        order = None
    else:
        query = query.order_by(created_sort_expr.desc(), Invoice.id.desc())
        order = None

    if order is not None:
        query = query.order_by(order)

    # Silent audit counters (kept in response meta; no server logs)
    audit = {
        'enabled': True,
        'filters': {
            'search': bool(search),
            'status': status if status else None,
            'invoice_type': invoice_type if invoice_type else None,
            'invoice_group': invoice_group if invoice_group else None,
            'party': party if party else None,
            'creator': creator if creator else None,
            'karat': karat if karat else None,
            'gold_type': gold_type if gold_type else None,
            'date_from': date_from_str if date_from_str else None,
            'date_to': date_to_str if date_to_str else None,
        },
        'sort': {
            'sort_by': sort_by,
            'sort_order': sort_order,
        },
        'counts': {
            'filtered_total': None,
            'filtered_customer_invoices': None,
            'filtered_supplier_invoices': None,
            'filtered_unlinked_invoices': None,
            'page_customer_invoices': None,
            'page_supplier_invoices': None,
            'page_unlinked_invoices': None,
        }
    }

    try:
        audit_query = query.order_by(None)
        audit_row = audit_query.with_entities(
            func.count(Invoice.id),
            func.coalesce(func.sum(case((Invoice.customer_id.isnot(None), 1), else_=0)), 0),
            func.coalesce(func.sum(case((Invoice.supplier_id.isnot(None), 1), else_=0)), 0),
            func.coalesce(
                func.sum(
                    case(
                        (and_(Invoice.customer_id.is_(None), Invoice.supplier_id.is_(None)), 1),
                        else_=0,
                    )
                ),
                0,
            ),
        ).first()

        if audit_row:
            audit['counts']['filtered_total'] = int(audit_row[0] or 0)
            audit['counts']['filtered_customer_invoices'] = int(audit_row[1] or 0)
            audit['counts']['filtered_supplier_invoices'] = int(audit_row[2] or 0)
            audit['counts']['filtered_unlinked_invoices'] = int(audit_row[3] or 0)
    except Exception:
        # Keep audit best-effort and non-fatal
        pass

    # Global tab summary (over FULL filtered result set, before pagination)
    def _empty_bucket():
        return {
            'total_invoices': 0,
            'total_amount': 0.0,
            'paid_amount': 0.0,
            'unpaid_amount': 0.0,
            'vat_total': 0.0,
            'sold_weight_total': 0.0,
        }

    tab_summary = {
        'sales': _empty_bucket(),
        'customer_purchase': _empty_bucket(),
        'supplier_purchase': _empty_bucket(),
        'returns': _empty_bucket(),
    }
    current_summary = _empty_bucket()
    available_employees = []

    def _classify_tab(invoice_type_value, supplier_id_value):
        t = (invoice_type_value or '').strip()
        if not t:
            return None
        lower = t.lower()
        is_return = ('مرتجع' in t) or ('return' in lower)
        if is_return:
            return 'returns'
        is_sale = ('بيع' in t) or (lower == 'sell') or ('sale' in lower)
        if is_sale:
            return 'sales'
        is_purchase = ('شراء' in t) or (lower == 'buy') or ('purchase' in lower)
        if is_purchase:
            return 'supplier_purchase' if supplier_id_value else 'customer_purchase'
        return None

    try:
        _mk_val = float(get_main_karat() or 21)

        # Pre-fetch weight (main karat equiv) from karat_lines AND items
        _all_inv_ids_q = query.order_by(None).with_entities(Invoice.id).all()
        _all_inv_ids = [r[0] for r in _all_inv_ids_q]

        _kl_weight_map = {}
        if _all_inv_ids:
            # 1) من karat_lines (أولوية)
            _kl_rows = (
                db.session.query(
                    InvoiceKaratLine.invoice_id,
                    func.sum(InvoiceKaratLine.weight_grams * InvoiceKaratLine.karat / _mk_val),
                )
                .filter(InvoiceKaratLine.invoice_id.in_(_all_inv_ids))
                .group_by(InvoiceKaratLine.invoice_id)
                .all()
            )
            for row in _kl_rows:
                v = float(row[1] or 0.0)
                if v > 0:
                    _kl_weight_map[row[0]] = v

            # 2) من InvoiceItem (للفواتير التي لا تملك karat_lines)
            _missing_ids = [iid for iid in _all_inv_ids if iid not in _kl_weight_map]
            if _missing_ids:
                _ii_rows = (
                    db.session.query(
                        InvoiceItem.invoice_id,
                        func.sum(InvoiceItem.weight * func.coalesce(InvoiceItem.karat, _mk_val) / _mk_val),
                    )
                    .filter(InvoiceItem.invoice_id.in_(_missing_ids), InvoiceItem.weight > 0)
                    .group_by(InvoiceItem.invoice_id)
                    .all()
                )
                for row in _ii_rows:
                    v = float(row[1] or 0.0)
                    if v > 0:
                        _kl_weight_map[row[0]] = v

        summary_rows = query.order_by(None).with_entities(
            Invoice.invoice_type,
            Invoice.status,
            Invoice.total,
            Invoice.total_tax,
            Invoice.total_weight,
            Invoice.amount_paid,
            Invoice.barter_total,
            Invoice.supplier_id,
            Invoice.id,
        ).all()

        for row in summary_rows:
            tab_key = _classify_tab(row[0], row[7])
            if not tab_key:
                continue

            bucket = tab_summary[tab_key]
            bucket['total_invoices'] += 1
            current_summary['total_invoices'] += 1

            status_value = (row[1] or '').strip().lower()
            if status_value == 'cancelled' or status_value == 'ملغاة':
                continue

            total_value = float(row[2] or 0.0)
            tax_value = float(row[3] or 0.0)
            inv_id = row[8]
            weight_value = _kl_weight_map.get(inv_id, float(row[4] or 0.0))
            paid_cash_value = float(row[5] or 0.0)
            barter_value = float(row[6] or 0.0)
            settled_value = paid_cash_value + barter_value
            paid_clamped = max(0.0, min(total_value, settled_value))
            remaining_value = max(0.0, total_value - settled_value)

            bucket['total_amount'] += total_value
            bucket['paid_amount'] += paid_clamped
            bucket['unpaid_amount'] += remaining_value
            bucket['vat_total'] += tax_value
            bucket['sold_weight_total'] += weight_value

            current_summary['total_amount'] += total_value
            current_summary['paid_amount'] += paid_clamped
            current_summary['unpaid_amount'] += remaining_value
            current_summary['vat_total'] += tax_value
            current_summary['sold_weight_total'] += weight_value

        for _, bucket in tab_summary.items():
            bucket['total_amount'] = round(float(bucket['total_amount']), 2)
            bucket['paid_amount'] = round(float(bucket['paid_amount']), 2)
            bucket['unpaid_amount'] = round(float(bucket['unpaid_amount']), 2)
            bucket['vat_total'] = round(float(bucket['vat_total']), 2)
            bucket['sold_weight_total'] = round(float(bucket['sold_weight_total']), 4)
        current_summary['total_amount'] = round(float(current_summary['total_amount']), 2)
        current_summary['paid_amount'] = round(float(current_summary['paid_amount']), 2)
        current_summary['unpaid_amount'] = round(float(current_summary['unpaid_amount']), 2)
        current_summary['vat_total'] = round(float(current_summary['vat_total']), 2)
        current_summary['sold_weight_total'] = round(float(current_summary['sold_weight_total']), 4)
        # Keep legacy 'purchase' key for backwards compatibility
        tab_summary['purchase'] = tab_summary['customer_purchase']
    except Exception as _summary_exc:
        # Non-fatal; frontend falls back to current-page aggregation.
        import traceback as _tb; _tb.print_exc()
        pass

    try:
        _ensure_employee_outerjoin()
        employee_rows = (
            query.order_by(None)
            .with_entities(Invoice.employee_id, Employee.name)
            .filter(Invoice.employee_id.isnot(None), Employee.name.isnot(None))
            .distinct()
            .all()
        )
        available_employees = [
            {
                'id': int(employee_id),
                'name': str(employee_name).strip(),
            }
            for employee_id, employee_name in employee_rows
            if employee_id is not None and str(employee_name or '').strip()
        ]
        available_employees.sort(key=lambda entry: entry['name'])
    except Exception:
        available_employees = []

    # Pagination
    paginated_invoices = query.paginate(page=page, per_page=per_page, error_out=False)
    invoices = paginated_invoices.items

    result = []
    for inv in invoices:
        invoice_dict = inv.to_dict()  # 🆕 استخدام to_dict() لتضمين payments
        
        # إضافة أسماء العملاء والموردين
        customer_name = inv.customer.name if inv.customer else (inv.supplier.name if inv.supplier else "N/A")
        supplier_name = inv.supplier.name if inv.supplier else "N/A"
        
        invoice_dict['customer_name'] = customer_name
        invoice_dict['supplier_name'] = supplier_name
        
        result.append(invoice_dict)

    try:
        page_customer = 0
        page_supplier = 0
        page_unlinked = 0
        for inv in invoices:
            if getattr(inv, 'customer_id', None) is not None:
                page_customer += 1
            elif getattr(inv, 'supplier_id', None) is not None:
                page_supplier += 1
            else:
                page_unlinked += 1

        audit['counts']['page_customer_invoices'] = page_customer
        audit['counts']['page_supplier_invoices'] = page_supplier
        audit['counts']['page_unlinked_invoices'] = page_unlinked
    except Exception:
        pass

    return jsonify({
        'invoices': result,
        'total': paginated_invoices.total,
        'pages': paginated_invoices.pages,
        'current_page': paginated_invoices.page,
        'per_page': paginated_invoices.per_page,
        'meta': {
            'audit': audit,
            'tab_summary': tab_summary,
            'current_summary': current_summary,
            'available_creators': available_creators,
            'available_employees': available_employees,
        },
    })

@invoices_bp.route('/invoices/<int:invoice_id>', methods=['GET'])
def get_invoice_by_id(invoice_id: int):
    """Fetch full invoice details by id.

    Flutter expects this endpoint for print/share flows.
    Returns Invoice.to_dict() plus customer/supplier display names.
    """

    invoice = Invoice.query.get_or_404(invoice_id)
    invoice_dict = invoice.to_dict()

    customer_name = (
        invoice.customer.name
        if invoice.customer
        else (invoice.supplier.name if invoice.supplier else "N/A")
    )
    supplier_name = invoice.supplier.name if invoice.supplier else "N/A"

    invoice_dict['customer_name'] = customer_name
    invoice_dict['supplier_name'] = supplier_name

    return jsonify(invoice_dict)

@invoices_bp.route('/invoices/<int:invoice_id>', methods=['PUT'])
@require_permission('invoice.edit')
def update_unposted_invoice(invoice_id: int):
    """Edit an unposted (draft / pending-approval) invoice.

    Only invoices with is_posted=False can be edited.
    Replaces items, payments, and header fields, then
    **deletes + recreates** the draft via the standard add_invoice
    flow (which handles JE creation, approval checks, etc.).

    The strategy is delete-and-recreate rather than update-in-place,
    because add_invoice has complex side-effects (journal entries,
    category weight movements, approval alerts) that must be consistent.
    The new invoice keeps the same invoice_type_id to preserve the
    display number.
    """

    from models import (
        CategoryWeightMovement, SystemAlert, InvoiceWeightSettlement,
    )

    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid or missing JSON body'}), 400

    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        return jsonify({'error': 'not_found', 'message': 'الفاتورة غير موجودة'}), 404

    if invoice.is_posted:
        return jsonify({
            'error': 'invoice_already_posted',
            'message': 'لا يمكن تعديل فاتورة مرحّلة. استخدم المرتجعات بدلاً من ذلك.',
        }), 400

    # Preserve the original invoice_type_id so the display number stays the same.
    original_type_id = invoice.invoice_type_id
    original_invoice_type = invoice.invoice_type
    original_date = invoice.date

    # --- 1. Delete all related entities ---
    try:
        # Journal entries linked to this invoice
        linked_jes = JournalEntry.query.filter_by(
            reference_type='invoice', reference_id=invoice_id
        ).all()
        for je in linked_jes:
            JournalEntryLine.query.filter_by(journal_entry_id=je.id).delete()
            db.session.delete(je)

        # Vouchers linked to this invoice
        linked_vouchers = Voucher.query.filter_by(
            reference_type='invoice', reference_id=invoice_id
        ).all()
        for v in linked_vouchers:
            # Delete all JEs linked to this voucher (via journal_entry_id FK + reference_type)
            vje_ids: set = set()
            if v.journal_entry_id:
                vje_ids.add(v.journal_entry_id)
            for vje in JournalEntry.query.filter_by(
                reference_type='voucher', reference_id=v.id
            ).all():
                vje_ids.add(vje.id)
            for vje_id in vje_ids:
                JournalEntryLine.query.filter_by(journal_entry_id=vje_id).delete()
                vje_obj = JournalEntry.query.get(vje_id)
                if vje_obj:
                    db.session.delete(vje_obj)
            # Detach FK before deleting voucher to avoid constraint errors
            if v.journal_entry_id:
                v.journal_entry_id = None
                db.session.flush()
            VoucherAccountLine.query.filter_by(voucher_id=v.id).delete()
            db.session.delete(v)

        # SafeBox transactions
        SafeBoxTransaction.query.filter_by(invoice_id=invoice_id).delete()

        # Category weight movements
        CategoryWeightMovement.query.filter_by(invoice_id=invoice_id).delete()

        # System alerts
        try:
            SystemAlert.query.filter_by(
                entity_type='Invoice', entity_id=invoice_id
            ).delete()
        except Exception:
            pass

        # Invoice weight settlements
        try:
            InvoiceWeightSettlement.query.filter_by(invoice_id=invoice_id).delete()
        except Exception:
            pass

        # WeightClosingOrder has invoice_id NOT NULL — must delete before invoice
        try:
            from models import WeightClosingOrder, WeightClosingExecution
            wco = WeightClosingOrder.query.filter_by(invoice_id=invoice_id).first()
            if wco:
                WeightClosingExecution.query.filter_by(order_id=wco.id).delete()
                db.session.delete(wco)
                db.session.flush()
        except Exception:
            pass

        # Child rows (cascade would handle these on delete, but be explicit)
        InvoiceItem.query.filter_by(invoice_id=invoice_id).delete()
        InvoicePayment.query.filter_by(invoice_id=invoice_id).delete()
        InvoiceKaratLine.query.filter_by(invoice_id=invoice_id).delete()

        # Delete the invoice itself
        db.session.delete(invoice)
        db.session.flush()

    except Exception as exc:
        db.session.rollback()
        return jsonify({
            'error': 'cleanup_failed',
            'message': f'فشل حذف بيانات الفاتورة القديمة: {exc}',
        }), 500

    # --- 2. Re-create via the standard add_invoice flow ---
    # Merge caller data with preserved original fields.
    create_data = dict(data)
    create_data['invoice_type_id'] = original_type_id
    if 'invoice_type' not in create_data:
        create_data['invoice_type'] = original_invoice_type
    if 'date' not in create_data:
        create_data['date'] = original_date.isoformat() if original_date else datetime.now().isoformat()

    # Inject into Flask request context and call add_invoice.
    from app import app as _app  # type: ignore

    current_user = getattr(g, 'current_user', None)
    headers = {}
    if current_user:
        try:
            from auth_decorators import generate_token
            token = generate_token(current_user, expires_in_minutes=2)
            if token:
                headers['Authorization'] = f'Bearer {token}'
        except Exception:
            pass

    with _app.test_request_context(
        '/api/invoices', method='POST',
        json=create_data, headers=headers,
    ):
        rv = add_invoice()

    status_code = 200
    resp_obj = None
    if isinstance(rv, tuple) and len(rv) >= 2:
        resp_obj, status_code = rv[0], rv[1]
    else:
        resp_obj, status_code = rv, 201

    try:
        resp_data = resp_obj.get_json(silent=True) if hasattr(resp_obj, 'get_json') else None
    except Exception:
        resp_data = None

    if int(status_code) >= 400:
        # Re-creation failed — the old invoice is already deleted.
        # Return the error so the user can fix and retry.
        db.session.rollback()
        return jsonify({
            'error': 'recreate_failed',
            'message': 'فشل إعادة إنشاء الفاتورة بعد الحذف. يرجى إنشاء فاتورة جديدة.',
            'original_invoice_id': invoice_id,
            'inner_error': resp_data,
        }), int(status_code)

    # Commit (add_invoice already committed internally).
    try:
        db.session.commit()
    except Exception:
        pass

    # Attach old_invoice_id for the frontend.
    result = resp_data or {}
    result['old_invoice_id'] = invoice_id
    result['edit_mode'] = True

    return jsonify(result), 200

@invoices_bp.route('/invoices/<int:invoice_id>', methods=['DELETE'])
@require_permission('invoice.edit')
def delete_unposted_invoice(invoice_id: int):
    """Delete an unposted invoice and all related records."""

    from models import (
        CategoryWeightMovement, SystemAlert, InvoiceWeightSettlement,
    )

    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        return jsonify({'error': 'not_found', 'message': 'الفاتورة غير موجودة'}), 404

    if invoice.is_posted:
        return jsonify({
            'error': 'invoice_already_posted',
            'message': 'لا يمكن حذف فاتورة مرحّلة.',
        }), 400

    try:
        # Journal entries
        linked_jes = JournalEntry.query.filter_by(
            reference_type='invoice', reference_id=invoice_id
        ).all()
        for je in linked_jes:
            JournalEntryLine.query.filter_by(journal_entry_id=je.id).delete()
            db.session.delete(je)

        # Vouchers + their JEs + their SBTs — all cleaned atomically
        linked_vouchers = Voucher.query.filter_by(
            reference_type='invoice', reference_id=invoice_id
        ).all()
        for v in linked_vouchers:
            # Collect all JE IDs linked to this voucher (via FK + reference_type)
            vje_ids: set = set()
            if v.journal_entry_id:
                vje_ids.add(v.journal_entry_id)
            for vje in JournalEntry.query.filter_by(
                reference_type='voucher', reference_id=v.id
            ).all():
                vje_ids.add(vje.id)
            for vje_id in vje_ids:
                JournalEntryLine.query.filter_by(journal_entry_id=vje_id).delete()
                vje_obj = JournalEntry.query.get(vje_id)
                if vje_obj:
                    db.session.delete(vje_obj)
            # Delete SBTs for this voucher (including any stray reversal rows)
            SafeBoxTransaction.query.filter(
                SafeBoxTransaction.ref_type.in_(['voucher', 'voucher_reversal']),
                SafeBoxTransaction.ref_id == v.id,
            ).delete(synchronize_session=False)
            VoucherAccountLine.query.filter_by(voucher_id=v.id).delete()
            db.session.delete(v)

        SafeBoxTransaction.query.filter_by(invoice_id=invoice_id).delete()
        CategoryWeightMovement.query.filter_by(invoice_id=invoice_id).delete()

        try:
            SystemAlert.query.filter_by(entity_type='Invoice', entity_id=invoice_id).delete()
        except Exception:
            pass
        try:
            InvoiceWeightSettlement.query.filter_by(invoice_id=invoice_id).delete()
        except Exception:
            pass

        # WeightClosingOrder has invoice_id NOT NULL — must delete before invoice
        try:
            from models import WeightClosingOrder, WeightClosingExecution
            wco = WeightClosingOrder.query.filter_by(invoice_id=invoice_id).first()
            if wco:
                WeightClosingExecution.query.filter_by(order_id=wco.id).delete()
                db.session.delete(wco)
                db.session.flush()
        except Exception:
            pass

        InvoiceItem.query.filter_by(invoice_id=invoice_id).delete()
        InvoicePayment.query.filter_by(invoice_id=invoice_id).delete()
        InvoiceKaratLine.query.filter_by(invoice_id=invoice_id).delete()

        db.session.delete(invoice)
        db.session.commit()

        return jsonify({'success': True, 'message': 'تم حذف الفاتورة بنجاح'}), 200

    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'delete_failed', 'message': str(exc)}), 500

@invoices_bp.route('/invoices/<int:invoice_id>/status', methods=['PATCH'])
@require_permission('invoice.edit')
def update_invoice_status(invoice_id: int):
    """Update invoice payment status.

    Flutter uses this endpoint from the invoices list screen.
    Allowed statuses: unpaid, partially_paid, paid.
    """

    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        return jsonify({'error': 'not_found', 'message': 'الفاتورة غير موجودة'}), 404

    data = request.get_json(silent=True) or {}
    status = str(data.get('status') or '').strip().lower()

    allowed = {'unpaid', 'partially_paid', 'paid'}
    if status not in allowed:
        return jsonify({
            'error': 'invalid_status',
            'message': 'حالة الفاتورة غير صالحة',
            'allowed': sorted(list(allowed)),
        }), 400

    try:
        invoice.status = status
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'update_failed', 'message': str(exc)}), 500

    return jsonify(invoice.to_dict()), 200

@invoices_bp.route('/invoices/<int:invoice_id>/reassign-employee', methods=['PATCH'])
@require_permission('manager')
def reassign_invoice_employee(invoice_id: int):
    """تغيير موظف الفاتورة — للمدير فقط، وللفواتير غير المرحّلة حصراً."""
    invoice = Invoice.query.get_or_404(invoice_id)

    if invoice.is_posted:
        return jsonify({'error': 'لا يمكن تغيير منشئ فاتورة مرحّلة'}), 400

    data = request.get_json(silent=True) or {}
    new_emp_id = data.get('employee_id')
    if new_emp_id is None:
        return jsonify({'error': 'employee_id مطلوب'}), 400
    try:
        new_emp_id = int(new_emp_id)
    except (TypeError, ValueError):
        return jsonify({'error': 'employee_id غير صالح'}), 400

    new_emp = Employee.query.get(new_emp_id)
    if not new_emp:
        return jsonify({'error': 'الموظف غير موجود'}), 404

    old_name   = invoice.posted_by or str(invoice.employee_id)
    old_emp_id = invoice.employee_id
    new_name   = new_emp.name

    # ── تحديث الفاتورة ──────────────────────────────────────────
    invoice.employee_id = new_emp_id
    invoice.posted_by   = new_name
    # scrap_holder_employee_id يُحدَّث دائماً لأن _resolve_employee_for_invoice يقرأه أولاً
    invoice_type_for_scrap = (getattr(invoice, 'invoice_type', '') or '').strip()
    gold_type_for_scrap    = (str(getattr(invoice, 'gold_type', '') or '')).strip().lower()
    _is_scrap_invoice = (
        invoice_type_for_scrap in ('شراء من عميل', 'مرتجع شراء')
        and gold_type_for_scrap == 'scrap'
    )
    if _is_scrap_invoice:
        invoice.scrap_holder_employee_id = new_emp_id

    # ── نقل حركات الخزينة المؤقتة (invoice_scrap_receipt) للموظف الجديد ──
    # هذه الحركات تُنشأ عند حفظ فاتورة الشراء من عميل قبل الترحيل
    try:
        invoice_type = (getattr(invoice, 'invoice_type', '') or '').strip()
        gold_type    = (str(getattr(invoice, 'gold_type', '') or '')).strip().lower()
        is_scrap_purchase = (invoice_type in ('شراء من عميل', 'مرتجع شراء') and gold_type == 'scrap')

        if is_scrap_purchase:
            # خزينة الموظف الجديد
            new_gold_safe = None
            try:
                new_emp_fresh = Employee.query.get(new_emp_id)
                new_gold_safe_id = getattr(new_emp_fresh, 'gold_safe_box_id', None) if new_emp_fresh else None
                if new_gold_safe_id:
                    new_gold_safe = SafeBox.query.get(new_gold_safe_id)
                    if new_gold_safe and (new_gold_safe.safe_type or '').lower() != 'gold':
                        new_gold_safe = None
            except Exception:
                pass

            if new_gold_safe:
                # 1) نقل SBTs المؤقتة (قبل الترحيل)
                provisional_sbts = SafeBoxTransaction.query.filter(
                    SafeBoxTransaction.ref_id == invoice_id,
                    SafeBoxTransaction.ref_type.in_(['invoice_scrap_receipt', 'invoice_scrap_return']),
                ).all()
                for sbt in provisional_sbts:
                    sbt.safe_box_id = new_gold_safe.id

                # 2) نقل SBTs الذهب الحقيقية (invoice_gold + invoice_gold_reversal) إلى الخزينة الجديدة
                # يضمن أن تاريخ الخزينة مترابط: الذهب الذي عُكس من الخزينة القديمة
                # يُنقَل معه سجلاته الأصلية حتى يبقى الرصيد صفراً في الخزينة القديمة
                for _ref_type in ('invoice_gold', 'invoice_gold_reversal'):
                    for sbt in SafeBoxTransaction.query.filter_by(
                        ref_id=invoice_id, ref_type=_ref_type
                    ).all():
                        sbt.safe_box_id = new_gold_safe.id
    except Exception:
        pass

    # ── تحديث القيود المرتبطة مباشرةً بالفاتورة ──────────────────
    try:
        for je in JournalEntry.query.filter_by(
            reference_type='invoice', reference_id=invoice_id
        ).all():
            if getattr(je, 'created_by', None) is not None:
                je.created_by = new_name
            if getattr(je, 'posted_by', None) is not None:
                je.posted_by = new_name
    except Exception:
        pass

    # ── تحديث السندات وقيودها المرتبطة بالفاتورة ─────────────────
    try:
        for v in Voucher.query.filter_by(
            reference_type='invoice', reference_id=invoice_id
        ).all():
            if getattr(v, 'created_by', None) is not None:
                v.created_by = new_name
            if getattr(v, 'approved_by', None) is not None:
                v.approved_by = new_name
            # قيود السند
            if v.journal_entry_id:
                vje = JournalEntry.query.get(v.journal_entry_id)
                if vje:
                    if getattr(vje, 'created_by', None) is not None:
                        vje.created_by = new_name
                    if getattr(vje, 'posted_by', None) is not None:
                        vje.posted_by = new_name
            for vje2 in JournalEntry.query.filter_by(
                reference_type='voucher', reference_id=v.id
            ).all():
                if getattr(vje2, 'created_by', None) is not None:
                    vje2.created_by = new_name
                if getattr(vje2, 'posted_by', None) is not None:
                    vje2.posted_by = new_name
    except Exception:
        pass

    try:
        actor = _actor_username()
        db.session.add(AuditLog(
            action='reassign_employee',
            entity_type='invoice',
            entity_id=invoice_id,
            actor=actor,
            details=f'تغيير منشئ {invoice.invoice_number}: {old_name} ← {new_name} (شامل القيود والسندات)',
        ))
    except Exception:
        pass

    db.session.commit()
    return jsonify({'success': True, 'new_employee': {'id': new_emp.id, 'name': new_emp.name}})

_PAYMENT_SAFE_BOX_GUARD_LOGGER = logging.getLogger('payment_safe_box_guard')

def _warn_if_safe_account_mismatches_payment_method(pm, safe_account_id, context: str = '') -> None:
    """Stage-2-style warning-only check (never raises, never blocks).

    Flags the case where a non-cash, non-receivable payment method resolves
    to a safe account that differs from its OWN configured default safe
    box's account -- the exact signature of a known frontend bug (an async
    safe-box lookup race in *_loadSafeBoxesForPaymentMethod*) that silently
    routed مدى/تحويل payments into the main cash account. Logged, not
    enforced, since an admin could legitimately override the safe box via
    "الخيارات المتقدمة" for a one-off reason; this is an audit trail for
    spotting recurrences, not a hard gate.
    """
    try:
        if pm is None or safe_account_id is None:
            return
        payment_type = str(getattr(pm, 'payment_type', '') or '').strip().lower()
        if payment_type in ('cash', 'receivable'):
            return
        default_safe_box_id = getattr(pm, 'default_safe_box_id', None)
        if not default_safe_box_id:
            return
        expected_box = SafeBox.query.get(default_safe_box_id)
        expected_account_id = getattr(expected_box, 'account_id', None)
        if expected_account_id and int(expected_account_id) != int(safe_account_id):
            _PAYMENT_SAFE_BOX_GUARD_LOGGER.warning(
                "payment_method=%r (id=%s, type=%s) resolved to account_id=%s but its "
                "own default safe box (id=%s) maps to account_id=%s -- possible stale "
                "safe-box selection from the client [%s]",
                getattr(pm, 'name', None), getattr(pm, 'id', None), payment_type,
                safe_account_id, default_safe_box_id, expected_account_id, context,
            )
    except Exception:
        pass

def _compute_commission_fields(amount: float, pm) -> dict:
    """Same formula used when an InvoicePayment is first created."""
    rate = float(getattr(pm, 'commission_rate', 0.0) or 0.0)
    try:
        timing = str(getattr(pm, 'commission_timing', 'invoice') or 'invoice').strip().lower()
    except Exception:
        timing = 'invoice'
    if timing == 'settlement':
        return {'commission_rate': rate, 'commission_amount': 0.0, 'commission_vat': 0.0, 'net_amount': amount}
    commission_amount = amount * (rate / 100.0) if rate > 0 else 0.0
    commission_vat = commission_amount * 0.15
    return {
        'commission_rate': rate,
        'commission_amount': commission_amount,
        'commission_vat': commission_vat,
        'net_amount': amount - commission_amount - commission_vat,
    }

def _correct_invoice_payment_method_multi_split(invoice, ip, data):
    """N-way split: one InvoicePayment row that mixed 3+ payment methods.

    Body shape for this mode:
      { "reason": "<string>", "splits": [{"payment_method_id": <int>, "amount": <float>}, ...] }

    Same rules as the 2-way split (single new_payment_method_id +
    correction_amount): the original row shrinks to the remainder and stays
    under its OLD payment method; one new InvoicePayment row is created per
    split entry; one independent reclassification voucher moves exactly the
    split amounts between accounts (never touches the original posted
    JournalEntryLine). Sum of all split amounts may equal the payment's full
    amount (nothing genuinely left under the old method) -- in that case the
    original row and its SafeBoxTransaction are removed instead of being
    shrunk to a 0.00 leftover; the full reclassification is still recorded
    via the voucher/JE + AuditLog either way. Sum may not EXCEED the full
    amount.
    """
    payment_id = ip.id
    reason = str(data.get('reason') or '').strip()
    if not reason:
        return jsonify({'error': 'reason is required', 'message': 'سبب التصحيح إلزامي'}), 400

    splits_raw = data.get('splits')
    if not isinstance(splits_raw, list) or len(splits_raw) < 1:
        return jsonify({'error': 'splits must be a non-empty list'}), 400

    full_amount = float(ip.amount or 0.0)
    old_pm_id = ip.payment_method_id
    old_pm = PaymentMethod.query.get(old_pm_id)
    old_sb_id = getattr(old_pm, 'default_safe_box_id', None) if old_pm else None
    old_safe_account_id = None
    if old_sb_id is not None:
        old_sb_obj = SafeBox.query.get(old_sb_id)
        old_safe_account_id = getattr(old_sb_obj, 'account_id', None) if old_sb_obj else None

    parsed_splits = []
    seen_pm_ids = set()
    total_split = 0.0
    for i, entry in enumerate(splits_raw):
        if not isinstance(entry, dict):
            return jsonify({'error': f'splits[{i}] must be an object'}), 400
        try:
            pm_id = int(entry.get('payment_method_id'))
        except (TypeError, ValueError):
            return jsonify({'error': f'splits[{i}].payment_method_id is invalid'}), 400
        try:
            amt = float(entry.get('amount'))
        except (TypeError, ValueError):
            return jsonify({'error': f'splits[{i}].amount is invalid'}), 400
        if amt <= 0:
            return jsonify({'error': f'splits[{i}].amount must be positive'}), 400
        pm = PaymentMethod.query.get(pm_id)
        if pm is None:
            return jsonify({'error': 'not_found', 'message': f'وسيلة الدفع في splits[{i}] غير موجودة'}), 404
        if pm_id == old_pm_id:
            return jsonify({'error': f'splits[{i}].payment_method_id يطابق الوسيلة الحالية'}), 400
        if pm_id in seen_pm_ids:
            return jsonify({
                'error': 'duplicate_payment_method',
                'message': f'وسيلة الدفع "{pm.name}" مكرّرة في أكثر من سطر تقسيم — اجمعها في سطر واحد',
            }), 400
        seen_pm_ids.add(pm_id)
        sb_id = getattr(pm, 'default_safe_box_id', None)
        safe_account_id = None
        if sb_id is not None:
            sb_obj = SafeBox.query.get(sb_id)
            safe_account_id = getattr(sb_obj, 'account_id', None) if sb_obj else None
        parsed_splits.append({'pm': pm, 'pm_id': pm_id, 'sb_id': sb_id, 'account_id': safe_account_id, 'amount': amt})
        total_split += amt

    total_split = round(total_split, 2)
    if total_split > full_amount + 0.005:
        return jsonify({
            'error': 'splits_exceed_payment',
            'message': (
                f'إجمالي مبالغ التقسيم ({total_split:.2f}) أكبر من مبلغ الدفعة الكامل '
                f'({full_amount:.2f}).'
            ),
        }), 400

    try:
        old_commission = {
            'commission_rate': ip.commission_rate,
            'commission_amount': ip.commission_amount,
            'commission_vat': ip.commission_vat,
            'net_amount': ip.net_amount,
        }

        # total_split may legitimately equal full_amount: the whole row was
        # mixed across N other methods with nothing genuinely left under the
        # old one (e.g. 3000 تمارا really being 1000 نقد + 500 مدى + 1500
        # تابي). Forcing a positive remainder would mislabel part of the
        # payment as still belonging to the old method just to satisfy a
        # validation rule -- so when the remainder rounds to ~0, the original
        # row (and its SafeBoxTransaction) is removed instead of shrunk to a
        # confusing 0.00 leftover. Nothing here touches the posted
        # JournalEntryLine or erases history: the full reclassification is
        # still recorded via the voucher/JE created below + AuditLog.
        remainder = round(full_amount - total_split, 2)
        original_removed = remainder <= 0.005
        old_sbt = SafeBoxTransaction.query.filter_by(invoice_payment_id=payment_id).first()
        if original_removed:
            if old_sbt:
                db.session.delete(old_sbt)
            db.session.delete(ip)
        else:
            ip.amount = remainder
            for k, v in _compute_commission_fields(remainder, old_pm).items():
                setattr(ip, k, v)
            if old_sbt:
                old_sbt.amount_cash = round(float(old_sbt.amount_cash or 0.0) - total_split, 2)

        actor = g.current_user.username if hasattr(g, 'current_user') and g.current_user else 'admin'
        new_ip_ids = []
        for s in parsed_splits:
            new_ip = InvoicePayment(
                invoice_id=invoice.id,
                payment_method_id=s['pm_id'],
                amount=s['amount'],
                notes=f'تقسيم من دفعة #{payment_id} — {reason}',
                **_compute_commission_fields(s['amount'], s['pm']),
            )
            db.session.add(new_ip)
            db.session.flush()
            s['new_ip_id'] = new_ip.id
            new_ip_ids.append(new_ip.id)

            if s['sb_id'] is not None:
                db.session.add(SafeBoxTransaction(
                    safe_box_id=s['sb_id'],
                    ref_type='invoice_payment',
                    ref_id=new_ip.id,
                    invoice_id=invoice.id,
                    invoice_payment_id=new_ip.id,
                    payment_method_id=s['pm_id'],
                    direction='in',
                    amount_cash=s['amount'],
                    notes=f'تقسيم من دفعة #{payment_id} — {reason}',
                    created_by=actor,
                ))

        reclass_voucher_number = None
        reclass_je_id = None
        # Splits whose target account differs from the OLD account are the
        # only ones that actually need cash to move in the GL -- a split
        # landing on the SAME account as the old method (e.g. بطاقة/تحويل
        # sharing one safe box) only changes InvoicePayment-level
        # classification, no GL movement. total_moved (not total_split) is
        # therefore the only amount-correct figure for the voucher/balances.
        debit_targets = [s for s in parsed_splits if s['account_id'] and s['account_id'] != old_safe_account_id]
        total_moved = round(sum(s['amount'] for s in debit_targets), 2)
        if old_safe_account_id and debit_targets:
            recon_dt = datetime.now()
            voucher_number = generate_voucher_number('adjustment', voucher_date=recon_dt)
            methods_desc = '، '.join(f"{s['pm'].name} ({s['amount']:.2f})" for s in parsed_splits)
            voucher = Voucher(
                voucher_number=voucher_number,
                voucher_type='adjustment',
                date=recon_dt,
                description=(
                    f'إعادة تصنيف وسيلة دفع (تقسيم متعدد): فاتورة {invoice.invoice_number} '
                    f'دفعة #{payment_id} ({getattr(old_pm, "name", old_pm_id)} → {methods_desc}) — {reason}'
                ),
                notes=json.dumps({
                    'old_payment_method_id': old_pm_id,
                    'old_commission': old_commission,
                    'reason': reason,
                    'is_partial': True,
                    'is_multi_split': True,
                    'full_amount': full_amount,
                    'total_split': total_split,
                    'total_moved': total_moved,
                    'original_payment_removed': original_removed,
                    'splits': [
                        {'payment_method_id': s['pm_id'], 'amount': s['amount'], 'new_invoice_payment_id': s['new_ip_id']}
                        for s in parsed_splits
                    ],
                }, ensure_ascii=False),
                created_by=actor,
                status='approved',
                approved_by=actor,
                approved_at=recon_dt,
                amount_cash=total_moved,
                amount_gold=0.0,
            )
            db.session.add(voucher)
            db.session.flush()

            db.session.add(VoucherAccountLine(
                voucher_id=voucher.id,
                account_id=int(old_safe_account_id),
                line_type='credit',
                amount_type='cash',
                amount=total_moved,
                description=f'إعادة تصنيف دفعة #{payment_id} من {getattr(old_pm, "name", old_pm_id)} (تقسيم متعدد)',
            ))
            # Group debit lines by account in case two splits share one.
            by_account: dict[int, float] = {}
            for s in debit_targets:
                by_account[s['account_id']] = round(by_account.get(s['account_id'], 0.0) + s['amount'], 2)
            for acc_id, amt in by_account.items():
                db.session.add(VoucherAccountLine(
                    voucher_id=voucher.id,
                    account_id=int(acc_id),
                    line_type='debit',
                    amount_type='cash',
                    amount=amt,
                    description=f'إعادة تصنيف دفعة #{payment_id} (تقسيم متعدد)',
                ))
            db.session.flush()

            journal_entry = create_journal_entry_from_voucher(voucher)
            if journal_entry:
                journal_entry.reference_type = 'payment_method_correction'
                journal_entry.reference_id = payment_id
                voucher.journal_entry_id = journal_entry.id
                reclass_je_id = journal_entry.id

            old_acc = Account.query.get(int(old_safe_account_id))
            if old_acc is not None:
                old_acc.update_balance(cash_amount=-total_moved)
            for acc_id, amt in by_account.items():
                acc = Account.query.get(int(acc_id))
                if acc is not None:
                    acc.update_balance(cash_amount=amt)

            reclass_voucher_number = voucher.voucher_number

        try:
            AuditLog.log_action(
                user_name=actor,
                action='correct_payment_method',
                entity_type='InvoicePayment',
                entity_id=payment_id,
                entity_number=invoice.invoice_number,
                details=json.dumps({
                    'old_payment_method_id': old_pm_id,
                    'old_commission': old_commission,
                    'is_partial': True,
                    'is_multi_split': True,
                    'full_amount': full_amount,
                    'total_split': total_split,
                    'total_moved': total_moved,
                    'original_payment_removed': original_removed,
                    'splits': [
                        {'payment_method_id': s['pm_id'], 'amount': s['amount'], 'new_invoice_payment_id': s['new_ip_id']}
                        for s in parsed_splits
                    ],
                    'reclassification_voucher_number': reclass_voucher_number,
                    'reclassification_journal_entry_id': reclass_je_id,
                    'reason': reason,
                }, ensure_ascii=False),
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent'),
                success=True,
            )
        except Exception:
            pass

        db.session.commit()

        return jsonify({
            'success': True,
            'payment_id': payment_id,
            'old_payment_method_id': old_pm_id,
            'is_partial': True,
            'is_multi_split': True,
            'full_amount': full_amount,
            'total_split': total_split,
            'total_moved': total_moved,
            'original_payment_removed': original_removed,
            'remaining_amount_on_original': 0.0 if original_removed else float(remainder),
            'new_invoice_payment_ids': new_ip_ids,
            'reclassification_voucher_number': reclass_voucher_number,
            'reclassification_journal_entry_id': reclass_je_id,
        }), 200

    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 500

@invoices_bp.route('/invoices/<int:invoice_id>/payments/<int:payment_id>/correct-method', methods=['POST'])
@require_admin
def correct_invoice_payment_method(invoice_id: int, payment_id: int):
    """Correct the payment method of an invoice payment -- fully or partially.

    Only allowed when the payment has not yet been settled (no SettlementLine
    at all). Once a SettlementLine exists, payment_method_id becomes part of
    the historical record (InvoicePayment -> SettlementLine -> Settlement
    Voucher -> real bank transfer already happened under that
    classification) and must never be rewritten -- any genuine economic
    correction at that point has to be a separate reclassification/transfer
    entry that does not touch payment_method_id.

    For the allowed (pre-settlement) case, this never mutates an existing
    posted JournalEntryLine. It creates a new, independent reclassification
    voucher (credit old account, debit new account) via the same
    create_journal_entry_from_voucher path every other voucher uses, tagged
    reference_type='payment_method_correction'.

    Body: {
      "new_payment_method_id": <int>,
      "reason": "<string>",
      "correction_amount": <float, optional>   # omit/equal to full amount = full reclassification (default)
    }

    If correction_amount is given and is LESS than the payment's full
    amount, this is a SPLIT correction: the original row was a single
    InvoicePayment that actually mixed two payment methods (e.g. one row of
    1000 recorded entirely as cash, when really 600 was مدى + 400 cash).
    The original row's own amount/commission shrink to the remainder under
    the OLD payment method (its payment_method_id is NOT changed); a NEW
    InvoicePayment row is created for correction_amount under the NEW
    payment method; the reclassification voucher moves only
    correction_amount, not the full original amount.
    """
    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        return jsonify({'error': 'not_found', 'message': 'الفاتورة غير موجودة'}), 404

    ip = InvoicePayment.query.filter_by(id=payment_id, invoice_id=invoice_id).first()
    if ip is None:
        return jsonify({'error': 'not_found', 'message': 'الدفعة غير موجودة'}), 404

    # Block if already settled -- payment_method_id is now a historical fact.
    if is_locked(payment_id):
        return jsonify({
            'error': 'already_settled',
            'message': (
                'لا يمكن تصحيح وسيلة الدفع بعد دخول الدفعة في سلسلة تسوية فعلية '
                '(InvoicePayment → SettlementLine → سند تسوية → تحويل بنكي). '
                'أي تصحيح اقتصادي مطلوب يجب أن يتم بقيد تحويل/مطابقة مستقل '
                'لا يغيّر تصنيف هذه الدفعة.'
            ),
        }), 409

    data = request.get_json(silent=True) or {}

    # N-way split mode: { "reason": ..., "splits": [{"payment_method_id", "amount"}, ...] }
    if isinstance(data.get('splits'), list):
        return _correct_invoice_payment_method_multi_split(invoice, ip, data)

    new_pm_id = data.get('new_payment_method_id')
    reason = str(data.get('reason') or '').strip()
    correction_amount_raw = data.get('correction_amount')

    if not new_pm_id:
        return jsonify({'error': 'new_payment_method_id is required'}), 400
    if not reason:
        return jsonify({'error': 'reason is required', 'message': 'سبب التصحيح إلزامي'}), 400

    new_pm = PaymentMethod.query.get(new_pm_id)
    if new_pm is None:
        return jsonify({'error': 'not_found', 'message': 'وسيلة الدفع غير موجودة'}), 404

    old_pm_id = ip.payment_method_id
    old_pm = PaymentMethod.query.get(old_pm_id)
    old_sb_id = getattr(old_pm, 'default_safe_box_id', None) if old_pm else None
    new_sb_id = getattr(new_pm, 'default_safe_box_id', None)

    if old_pm_id == new_pm_id:
        return jsonify({'error': 'same_method', 'message': 'وسيلة الدفع هي نفسها الحالية'}), 400

    full_amount = float(ip.amount or 0.0)
    correction_amount = full_amount
    if correction_amount_raw is not None:
        try:
            correction_amount = float(correction_amount_raw)
        except (TypeError, ValueError):
            return jsonify({'error': 'invalid correction_amount'}), 400
        if correction_amount <= 0:
            return jsonify({'error': 'correction_amount must be positive'}), 400
        if correction_amount > full_amount + 0.005:
            return jsonify({
                'error': 'correction_amount_exceeds_payment',
                'message': f'المبلغ المطلوب تصحيحه ({correction_amount:.2f}) أكبر من مبلغ الدفعة نفسها ({full_amount:.2f})',
            }), 400
    is_partial = abs(correction_amount - full_amount) > 0.005

    old_safe_account_id = None
    new_safe_account_id = None
    if old_sb_id is not None:
        old_sb_obj = SafeBox.query.get(old_sb_id)
        old_safe_account_id = getattr(old_sb_obj, 'account_id', None) if old_sb_obj else None
    if new_sb_id is not None:
        new_sb_obj = SafeBox.query.get(new_sb_id)
        new_safe_account_id = getattr(new_sb_obj, 'account_id', None) if new_sb_obj else None

    try:
        old_commission = {
            'commission_rate': ip.commission_rate,
            'commission_amount': ip.commission_amount,
            'commission_vat': ip.commission_vat,
            'net_amount': ip.net_amount,
        }

        new_ip_id = None
        if not is_partial:
            # Full reclassification: same row, new payment method.
            ip.payment_method_id = new_pm_id
            for k, v in _compute_commission_fields(full_amount, new_pm).items():
                setattr(ip, k, v)

            if new_sb_id is not None:
                sbt = SafeBoxTransaction.query.filter_by(invoice_payment_id=payment_id).first()
                if sbt:
                    sbt.safe_box_id = new_sb_id
        else:
            # Split: original row shrinks to the remainder, stays under the
            # OLD payment method; a new row is created for correction_amount
            # under the NEW payment method.
            remainder = round(full_amount - correction_amount, 2)
            ip.amount = remainder
            for k, v in _compute_commission_fields(remainder, old_pm).items():
                setattr(ip, k, v)

            old_sbt = SafeBoxTransaction.query.filter_by(invoice_payment_id=payment_id).first()
            if old_sbt:
                old_sbt.amount_cash = round(float(old_sbt.amount_cash or 0.0) - correction_amount, 2)

            new_ip = InvoicePayment(
                invoice_id=invoice.id,
                payment_method_id=new_pm_id,
                amount=correction_amount,
                notes=f'تقسيم من دفعة #{payment_id} — {reason}',
                **_compute_commission_fields(correction_amount, new_pm),
            )
            db.session.add(new_ip)
            db.session.flush()
            new_ip_id = new_ip.id

            if new_sb_id is not None:
                db.session.add(SafeBoxTransaction(
                    safe_box_id=new_sb_id,
                    ref_type='invoice_payment',
                    ref_id=new_ip.id,
                    invoice_id=invoice.id,
                    invoice_payment_id=new_ip.id,
                    payment_method_id=new_pm_id,
                    direction='in',
                    amount_cash=correction_amount,
                    notes=f'تقسيم من دفعة #{payment_id} — {reason}',
                    created_by=(g.current_user.username if hasattr(g, 'current_user') and g.current_user else 'admin'),
                ))

        # Independent reclassification voucher -- never touches the original
        # posted JournalEntryLine. Credit old account, debit new account, for
        # correction_amount only (== full_amount when not partial).
        reclass_voucher_number = None
        reclass_je_id = None
        if old_safe_account_id and new_safe_account_id and old_safe_account_id != new_safe_account_id:
            recon_dt = datetime.now()
            voucher_number = generate_voucher_number('adjustment', voucher_date=recon_dt)
            voucher = Voucher(
                voucher_number=voucher_number,
                voucher_type='adjustment',
                date=recon_dt,
                description=(
                    f'إعادة تصنيف وسيلة دفع{" (جزئي)" if is_partial else ""}: فاتورة {invoice.invoice_number} '
                    f'دفعة #{payment_id} ({getattr(old_pm, "name", old_pm_id)} → {new_pm.name}، '
                    f'{correction_amount:.2f} من {full_amount:.2f}) — {reason}'
                ),
                notes=json.dumps({
                    'old_payment_method_id': old_pm_id,
                    'new_payment_method_id': new_pm_id,
                    'old_commission': old_commission,
                    'reason': reason,
                    'is_partial': is_partial,
                    'correction_amount': correction_amount,
                    'full_amount': full_amount,
                    'new_invoice_payment_id': new_ip_id,
                }, ensure_ascii=False),
                created_by=(g.current_user.username if hasattr(g, 'current_user') and g.current_user else 'admin'),
                status='approved',
                approved_by=(g.current_user.username if hasattr(g, 'current_user') and g.current_user else 'admin'),
                approved_at=recon_dt,
                amount_cash=correction_amount,
                amount_gold=0.0,
            )
            db.session.add(voucher)
            db.session.flush()

            db.session.add(VoucherAccountLine(
                voucher_id=voucher.id,
                account_id=int(new_safe_account_id),
                line_type='debit',
                amount_type='cash',
                amount=correction_amount,
                description=f'إعادة تصنيف دفعة #{payment_id} إلى {new_pm.name}',
            ))
            db.session.add(VoucherAccountLine(
                voucher_id=voucher.id,
                account_id=int(old_safe_account_id),
                line_type='credit',
                amount_type='cash',
                amount=correction_amount,
                description=f'إعادة تصنيف دفعة #{payment_id} من {getattr(old_pm, "name", old_pm_id)}',
            ))
            db.session.flush()

            journal_entry = create_journal_entry_from_voucher(voucher)
            if journal_entry:
                journal_entry.reference_type = 'payment_method_correction'
                journal_entry.reference_id = payment_id
                voucher.journal_entry_id = journal_entry.id
                reclass_je_id = journal_entry.id

            new_acc = Account.query.get(int(new_safe_account_id))
            old_acc = Account.query.get(int(old_safe_account_id))
            if new_acc is not None:
                new_acc.update_balance(cash_amount=correction_amount)
            if old_acc is not None:
                old_acc.update_balance(cash_amount=-correction_amount)

            reclass_voucher_number = voucher.voucher_number

        try:
            AuditLog.log_action(
                user_name=g.current_user.username if hasattr(g, 'current_user') and g.current_user else 'admin',
                action='correct_payment_method',
                entity_type='InvoicePayment',
                entity_id=payment_id,
                entity_number=invoice.invoice_number,
                details=json.dumps({
                    'old_payment_method_id': old_pm_id,
                    'new_payment_method_id': new_pm_id,
                    'old_safe_box_id': old_sb_id,
                    'new_safe_box_id': new_sb_id,
                    'old_safe_account_id': old_safe_account_id,
                    'new_safe_account_id': new_safe_account_id,
                    'old_commission': old_commission,
                    'is_partial': is_partial,
                    'full_amount': full_amount,
                    'correction_amount': correction_amount,
                    'new_invoice_payment_id': new_ip_id,
                    'reclassification_voucher_number': reclass_voucher_number,
                    'reclassification_journal_entry_id': reclass_je_id,
                    'reason': reason,
                }, ensure_ascii=False),
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent'),
                success=True,
            )
        except Exception:
            pass

        db.session.commit()

        return jsonify({
            'success': True,
            'payment_id': payment_id,
            'old_payment_method_id': old_pm_id,
            'new_payment_method_id': new_pm_id,
            'old_safe_box_id': old_sb_id,
            'new_safe_box_id': new_sb_id,
            'old_safe_account_id': old_safe_account_id,
            'new_safe_account_id': new_safe_account_id,
            'is_partial': is_partial,
            'full_amount': full_amount,
            'correction_amount': correction_amount,
            'remaining_amount_on_original': float(ip.amount or 0.0) if is_partial else None,
            'new_invoice_payment_id': new_ip_id,
            'reclassification_voucher_number': reclass_voucher_number,
            'reclassification_journal_entry_id': reclass_je_id,
        }), 200

    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 500

@invoices_bp.route('/invoices/<int:invoice_id>/approve', methods=['POST'])
@require_permission('invoice.edit')
def approve_invoice(invoice_id: int):
    """ترحيل فاتورة غير مرحّلة (تحتاج اعتماد المدير أو خيار الترحيل التلقائي معطّل).

    يُرحّل الفاتورة وجميع قيودها المحاسبية المرتبطة دفعةً واحدة.
    """
    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        return jsonify({'error': 'not_found', 'message': 'الفاتورة غير موجودة'}), 404

    if invoice.is_posted:
        return jsonify({'error': 'already_posted', 'message': 'الفاتورة مرحّلة بالفعل'}), 400

    data = request.get_json(silent=True) or {}
    approved_by = (
        (getattr(getattr(g, 'current_user', None), 'username', None))
        or data.get('approved_by')
        or 'system'
    )

    try:
        now = datetime.now()

        # 1. Cascade: post all linked invoice JEs
        linked_jes = JournalEntry.query.filter_by(
            reference_type='invoice', reference_id=invoice_id
        ).all()
        for je in linked_jes:
            if not je.is_posted:
                je.is_posted = True
                je.is_draft = False
                je.posted_at = now
                je.posted_by = approved_by

        # 1b. Cascade: post all linked voucher JEs
        try:
            _linked_vouchers = Voucher.query.filter_by(
                reference_type='invoice', reference_id=invoice_id
            ).all()
            _voucher_ids = [v.id for v in _linked_vouchers]
            if _voucher_ids:
                _voucher_jes = JournalEntry.query.filter(
                    JournalEntry.reference_type == 'voucher',
                    JournalEntry.reference_id.in_(_voucher_ids),
                ).all()
                for _vje in _voucher_jes:
                    if not _vje.is_posted:
                        _vje.is_posted = True
                        _vje.is_draft = False
                        _vje.posted_at = now
                        _vje.posted_by = approved_by
                linked_jes.extend(_voucher_jes)  # include in linked_jes for SBT sync below
        except Exception as exc:
            print(f"⚠️ Auto-post voucher JEs on approve skipped: {exc}")

        # 2. Post the invoice itself
        invoice.is_posted = True
        invoice.posted_at = now
        if not invoice.posted_by:
            invoice.posted_by = approved_by

        # 3. Sync payment status
        try:
            total_amount = float(invoice.total or 0.0)
            paid_amount = float(invoice.amount_paid or 0.0)
            barter_total = float(getattr(invoice, 'barter_total', 0.0) or 0.0)
            total_settled = paid_amount + barter_total
            eps = 0.01
            if total_amount <= eps:
                invoice.status = 'paid' if total_settled > eps else 'unpaid'
            elif total_settled <= eps:
                invoice.status = 'unpaid'
            elif total_settled >= total_amount - eps:
                invoice.status = 'paid'
            else:
                invoice.status = 'partially_paid'
        except Exception:
            pass

        db.session.flush()

        # 4. Record category-weight movements (skipped if already recorded)
        try:
            from category_weight_tracking import record_category_weight_movements_for_invoice_payload
            # Build minimal items payload from stored InvoiceItem rows
            items_payload = [
                {'item_id': ii.item_id, 'weight': float(ii.weight or 0.0)}
                for ii in InvoiceItem.query.filter_by(invoice_id=invoice_id).all()
                if ii.item_id
            ]
            record_category_weight_movements_for_invoice_payload(
                invoice_id=invoice_id,
                items_payload=items_payload or None,
            )
        except Exception as exc:
            print(f"⚠️ Category weight tracking skipped on approve: {exc}")

        # 📒 Inventory ledger posting on approval
        try:
            from services.inventory_posting_service import InventoryPostingService
            InventoryPostingService.post(invoice)
        except Exception as exc:
            print(f"⚠️ Inventory ledger posting skipped on approve: {exc}")

        # 5. Mark any SystemAlerts for this invoice as reviewed
        try:
            from models import SystemAlert
            SystemAlert.query.filter_by(
                entity_type='Invoice', entity_id=invoice_id, is_reviewed=False
            ).update({'is_reviewed': True, 'reviewed_by': approved_by, 'reviewed_at': now})
        except Exception:
            pass

        # 6. Sync safe-box transactions for invoice JE lines
        try:
            for je in linked_jes:
                _ensure_safe_box_transactions_for_invoice_je(
                    invoice_id=invoice_id,
                    journal_entry_id=je.id,
                    created_by=approved_by,
                )
        except Exception as exc:
            print(f"⚠️ safe-box SBT sync skipped on approve: {exc}")

        # Recompute stored Account.balance_* for every account touched by the
        # now-posted JEs.  Without this the cached balance never reflects the
        # newly-posted lines and the trial balance drifts.
        try:
            _approve_affected_ids = set()
            for _aje in linked_jes:
                _approve_affected_ids.update(
                    l.account_id for l in (_aje.lines or []) if l.account_id
                )
            if _approve_affected_ids:
                _recalculate_account_balances_for_accounts(_approve_affected_ids)
        except Exception as _rc_exc:
            print(f"⚠️ recalculate balances after approve skipped: {_rc_exc}")

        db.session.commit()
        return jsonify({'success': True, 'invoice': invoice.to_dict()}), 200

    except Exception as exc:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': 'approve_failed', 'message': str(exc)}), 500

@invoices_bp.route('/invoices/<int:invoice_id>/reject', methods=['POST'])
@require_permission('invoice.edit')
def reject_invoice(invoice_id: int):
    """رفض فاتورة غير مرحّلة وإعادة الحجز المرتبط بها (إن وُجد) إلى حالة pending."""
    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        return jsonify({'error': 'not_found', 'message': 'الفاتورة غير موجودة'}), 404

    if invoice.is_posted:
        return jsonify({'error': 'already_posted', 'message': 'لا يمكن رفض فاتورة مرحّلة بالفعل'}), 400

    if invoice.status == 'rejected':
        return jsonify({'success': True, 'invoice': invoice.to_dict()}), 200

    data = request.get_json(silent=True) or {}
    rejection_reason = (data.get('reason') or '').strip()

    try:
        # ── إعادة الحجز المرتبط إلى pending ────────────────────────────────
        linked_reservation = OfficeReservation.query.filter_by(
            purchase_invoice_id=invoice_id
        ).first()
        if linked_reservation:
            linked_reservation.purchase_invoice_id = None
            linked_reservation.status = 'pending'
            db.session.add(linked_reservation)

        # ── وضع علامة رفض على الفاتورة ─────────────────────────────────────
        invoice.status = 'rejected'
        if rejection_reason:
            current_notes = (getattr(invoice, 'notes', None) or '').strip()
            invoice.notes = f'[مرفوض: {rejection_reason}]\n{current_notes}'.strip() if hasattr(invoice, 'notes') else None
        db.session.add(invoice)

        db.session.commit()
        return jsonify({
            'success': True,
            'invoice': invoice.to_dict(),
            'reservation_reset': linked_reservation.id if linked_reservation else None,
        }), 200

    except Exception as exc:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': 'reject_failed', 'message': str(exc)}), 500

@invoices_bp.route('/invoices/<int:invoice_id>/unpost', methods=['POST'])
@require_permission('invoice.edit')
def unpost_invoice(invoice_id: int):
    """إلغاء ترحيل فاتورة (يتطلب تمكين خيار allow_unposting في الإعدادات).

    يُلغي ترحيل الفاتورة وجميع قيودها المحاسبية المرتبطة دفعةً واحدة.
    """
    # Check setting
    try:
        settings_row = Settings.query.first()
        allow_unposting = bool(getattr(settings_row, 'allow_unposting', False)) if settings_row else False
    except Exception:
        allow_unposting = False

    if not allow_unposting:
        return jsonify({
            'error': 'unposting_disabled',
            'message': 'إلغاء الترحيل غير مفعّل. يمكن تفعيله من إعدادات النظام.',
        }), 403

    invoice = Invoice.query.get(invoice_id)
    if not invoice:
        return jsonify({'error': 'not_found', 'message': 'الفاتورة غير موجودة'}), 404

    if not invoice.is_posted:
        return jsonify({'error': 'not_posted', 'message': 'الفاتورة غير مرحّلة أصلاً'}), 400

    data = request.get_json(silent=True) or {}
    unposted_by = (
        (getattr(getattr(g, 'current_user', None), 'username', None))
        or data.get('unposted_by')
        or 'system'
    )

    try:
        # Accounts touched by any JE we're about to unpost -- their stored
        # balance_cash/balance_*k must be recomputed afterward, otherwise the
        # cached balance keeps counting a JE that no longer affects the live
        # (is_posted=True) ledger total, drifting further out of sync on
        # every unpost. (Confirmed in production: 76 historical unposts with
        # no corresponding reversal, accounting for a large, previously
        # undiagnosed gap between cached and ledger-computed balances.)
        affected_account_ids = set()

        # 1. Cascade: unpost all linked invoice JEs
        linked_jes = JournalEntry.query.filter_by(
            reference_type='invoice', reference_id=invoice_id
        ).all()
        for je in linked_jes:
            affected_account_ids.update(l.account_id for l in je.lines if l.account_id)
            je.is_posted = False
            je.posted_at = None
            je.posted_by = None

        # 2. Cascade: reset linked payment vouchers + their JEs atomically
        linked_vouchers = Voucher.query.filter_by(
            reference_type='invoice', reference_id=invoice_id
        ).all()
        voucher_ids = [v.id for v in linked_vouchers]
        for v in linked_vouchers:
            # Unpost the voucher JE via FK
            if v.journal_entry_id:
                vje = JournalEntry.query.get(v.journal_entry_id)
                if vje and vje.is_posted:
                    affected_account_ids.update(l.account_id for l in vje.lines if l.account_id)
                    vje.is_posted = False
                    vje.posted_at = None
                    vje.posted_by = None
            # Unpost any additional voucher JEs via reference_type link
            for vje2 in JournalEntry.query.filter_by(
                reference_type='voucher', reference_id=v.id, is_posted=True
            ).all():
                affected_account_ids.update(l.account_id for l in vje2.lines if l.account_id)
                vje2.is_posted = False
                vje2.posted_at = None
                vje2.posted_by = None
            # Reset voucher to pending so it can be reposted if needed
            v.status = 'pending'

        # 3. Atomically delete ALL SafeBoxTransactions linked to this invoice.
        # Using direct DELETE rather than append-only reversal rows avoids orphan
        # voucher_reversal SBTs that have no matching GL reversal JE.
        SafeBoxTransaction.query.filter(
            SafeBoxTransaction.invoice_id == invoice_id
        ).delete(synchronize_session=False)
        if voucher_ids:
            SafeBoxTransaction.query.filter(
                SafeBoxTransaction.ref_type.in_(['voucher', 'voucher_reversal']),
                SafeBoxTransaction.ref_id.in_(voucher_ids),
            ).delete(synchronize_session=False)

        # 4. Unpost the invoice
        invoice.is_posted = False
        invoice.posted_at = None

        # 5. Remove category-weight movements (only valid for posted invoices)
        try:
            from models import CategoryWeightMovement
            CategoryWeightMovement.query.filter_by(invoice_id=invoice_id).delete()
        except Exception:
            pass

        # 6. Recompute stored balances for every account these now-unposted
        # JEs used to count toward, so balance_cash/balance_*k matches the
        # live ledger total immediately, not just after a manual rebuild.
        if affected_account_ids:
            _recalculate_account_balances_for_accounts(affected_account_ids)

        db.session.commit()
        return jsonify({'success': True, 'invoice': invoice.to_dict()}), 200

    except Exception as exc:
        db.session.rollback()
        return jsonify({'error': 'unpost_failed', 'message': str(exc)}), 500

def _add_payment_lines_to_consolidated_je(
    *,
    invoice,
    voucher,
    voucher_number: str,
    safe_account_id: int,
    party_account_id: int,
    amount: float,
    payment_id: int,
    direction: str,
    voucher_date,
    created_by: str,
):
    """Find or create a consolidated JE for this invoice's payments, then add lines.

    All payment vouchers for the same invoice share ONE JournalEntry
    (reference_type='invoice_payments'). Each payment adds 2 lines (debit/credit).

    Returns the JournalEntry.
    """
    consolidated_je = (
        JournalEntry.query
        .filter(
            JournalEntry.is_deleted == False,
            JournalEntry.reference_type == 'invoice_payments',
            JournalEntry.reference_id == int(invoice.id),
        )
        .first()
    )

    if consolidated_je is None:
        inv_num = (
            getattr(invoice, 'invoice_number', None)
            or getattr(invoice, 'invoice_type_id', None)
            or str(invoice.id)
        )
        consolidated_je = JournalEntry(
            entry_number=_generate_journal_entry_number(entry_date=voucher_date),
            date=voucher_date,
            description=f'دفعات فاتورة {inv_num}',
            reference_type='invoice_payments',
            reference_id=int(invoice.id),
            reference_number=str(inv_num),
            is_posted=True,
            posted_at=datetime.now(),
            posted_by=created_by,
            created_by=created_by,
        )
        db.session.add(consolidated_je)
        db.session.flush()

    # Add lines for this payment
    if direction == 'in':
        db.session.add(JournalEntryLine(
            journal_entry_id=consolidated_je.id,
            account_id=int(safe_account_id),
            cash_debit=float(amount),
            description=f'استلام نقد - دفعة #{payment_id} ({voucher_number})',
        ))
        db.session.add(JournalEntryLine(
            journal_entry_id=consolidated_je.id,
            account_id=int(party_account_id),
            cash_credit=float(amount),
            description=f'تسوية ذمم - دفعة #{payment_id} ({voucher_number})',
        ))
    else:
        db.session.add(JournalEntryLine(
            journal_entry_id=consolidated_je.id,
            account_id=int(party_account_id),
            cash_debit=float(amount),
            description=f'تسوية ذمم - دفعة #{payment_id} ({voucher_number})',
        ))
        db.session.add(JournalEntryLine(
            journal_entry_id=consolidated_je.id,
            account_id=int(safe_account_id),
            cash_credit=float(amount),
            description=f'صرف نقد - دفعة #{payment_id} ({voucher_number})',
        ))

    # Update description with all voucher numbers
    try:
        inv_num = (
            getattr(invoice, 'invoice_number', None)
            or getattr(invoice, 'invoice_type_id', None)
            or str(invoice.id)
        )
        linked_vouchers = (
            Voucher.query
            .filter(
                Voucher.reference_type == 'invoice',
                Voucher.reference_id == int(invoice.id),
                Voucher.status != 'cancelled',
            )
            .all()
        )
        voucher_nums = [v.voucher_number for v in linked_vouchers if v.voucher_number]
        if voucher_number not in voucher_nums:
            voucher_nums.append(voucher_number)
        consolidated_je.description = f'دفعات فاتورة {inv_num} ({", ".join(voucher_nums)})'
    except Exception:
        pass

    return consolidated_je

@invoices_bp.route('/invoices/<int:invoice_id>/payments', methods=['POST'])
def add_invoice_payment(invoice_id: int):
    """Add a payment entry to an existing invoice.

    Body JSON:
      - payment_method_id (int)
      - amount (float)
      - notes (optional)
    """

    invoice = Invoice.query.get_or_404(invoice_id)
    data = request.get_json(silent=True) or {}

    pm_id = data.get('payment_method_id')
    amount = data.get('amount')
    notes = data.get('notes')

    try:
        pm_id = int(pm_id)
    except Exception:
        return jsonify({'error': 'invalid_payment_method_id'}), 400

    def _to_float(v, default=0.0):
        try:
            if v in (None, ''):
                return default
            return float(v)
        except Exception:
            return default

    amount = _to_float(amount, 0.0)
    if amount <= 0:
        return jsonify({'error': 'invalid_amount'}), 400

    # Determine if partial payments are allowed.
    allow_partial_payments = False
    try:
        env_flag = str(os.getenv('ALLOW_PARTIAL_INVOICE_PAYMENTS', '')).strip().lower()
        if env_flag in ('1', 'true', 'yes', 'on'):
            allow_partial_payments = True
    except Exception:
        allow_partial_payments = False

    if not allow_partial_payments:
        try:
            settings_row = Settings.query.first()
            allow_partial_payments = bool(getattr(settings_row, 'allow_partial_invoice_payments', False)) if settings_row else False
        except Exception:
            allow_partial_payments = False

    # Compute already paid.
    paid_amount = 0.0
    try:
        if invoice.amount_paid is not None:
            paid_amount = float(invoice.amount_paid or 0.0)
        elif invoice.payments:
            paid_amount = float(sum(p.amount or 0.0 for p in invoice.payments))
    except Exception:
        paid_amount = 0.0

    total_amount = float(invoice.total or 0.0)
    remaining = max(total_amount - paid_amount, 0.0)
    eps = 0.01

    if remaining <= eps:
        return jsonify({'error': 'invoice_already_paid'}), 400

    if amount > remaining + eps:
        return jsonify({'error': 'amount_exceeds_remaining', 'remaining': round(remaining, 2)}), 400

    if not allow_partial_payments and abs(amount - remaining) > eps:
        return jsonify({'error': 'partial_payments_not_allowed', 'remaining': round(remaining, 2)}), 400

    pm_obj = PaymentMethod.query.get(pm_id)
    if not pm_obj:
        return jsonify({'error': 'payment_method_not_found'}), 404

    def _fallback_cash_safe_box_id() -> int | None:
        """Fallback cash SafeBox when none is supplied/configured.

        Precedence:
        - If Settings.employee_cash_safes_enabled and invoice.employee has cash_safe_box_id -> use it
        - Else Settings.main_cash_safe_box_id
        - Else default cash safe
        """
        try:
            settings_row = Settings.query.first()
        except Exception:
            settings_row = None

        if bool(getattr(settings_row, 'employee_cash_safes_enabled', False)):
            try:
                emp = getattr(invoice, 'employee', None)
                if not emp and getattr(invoice, 'employee_id', None):
                    emp = Employee.query.get(int(invoice.employee_id))
                emp_cash = getattr(emp, 'cash_safe_box_id', None) if emp else None
                if emp_cash not in (None, '', 0, '0', False):
                    return int(emp_cash)
            except Exception:
                pass

        try:
            main_cash = getattr(settings_row, 'main_cash_safe_box_id', None) if settings_row else None
            if main_cash not in (None, '', 0, '0', False):
                return int(main_cash)
        except Exception:
            pass

        try:
            sb = SafeBox.get_default_by_type('cash')
            if sb and sb.id:
                return int(sb.id)
        except Exception:
            pass

        # If no explicit default is configured, but there is exactly one active
        # cash safe box, use it as a conservative fallback.
        try:
            safes = SafeBox.query.filter_by(safe_type='cash', is_active=True).all()
            if isinstance(safes, list) and len(safes) == 1 and getattr(safes[0], 'id', None):
                return int(safes[0].id)
        except Exception:
            pass

        # If multiple active cash safes exist (common in production), pick a stable
        # fallback instead of failing: prefer default then lowest id.
        try:
            sb = (
                SafeBox.query.filter_by(safe_type='cash', is_active=True)
                .order_by(SafeBox.is_default.desc(), SafeBox.id.asc())
                .first()
            )
            if sb and getattr(sb, 'id', None):
                return int(sb.id)
        except Exception:
            pass

        # Last resort: if there is exactly one active safe box in the system,
        # use it rather than failing (helps when safe_type is misconfigured).
        try:
            safes = SafeBox.query.filter_by(is_active=True).all()
            if isinstance(safes, list) and len(safes) == 1 and getattr(safes[0], 'id', None):
                return int(safes[0].id)
        except Exception:
            pass

        # If multiple active safes exist, pick a stable fallback rather than failing.
        try:
            sb = SafeBox.query.filter_by(is_active=True).order_by(SafeBox.id.asc()).first()
            if sb and getattr(sb, 'id', None):
                return int(sb.id)
        except Exception:
            pass

        return None

    def _fallback_non_cash_safe_box_id(pm: PaymentMethod | None) -> int | None:
        """Fallback SafeBox for non-cash payment methods when none is supplied/configured.

        Precedence:
        - If auto settlement enabled: default clearing safe, then default bank safe
        - Otherwise: default bank safe, then default clearing safe

        This is intentionally conservative and only kicks in when the payment
        method has no `default_safe_box_id` and the request didn't supply one.
        """
        try:
            if pm is None:
                return None
            auto_settle = bool(getattr(pm, 'auto_settlement_enabled', False))
        except Exception:
            auto_settle = False

        def pick_default(t: str) -> int | None:
            try:
                sb = SafeBox.get_default_by_type(t)
                if sb and sb.id:
                    return int(sb.id)
            except Exception:
                return None

            # If no explicit default is configured, but there is exactly one
            # active safe box of this type, use it as a conservative fallback.
            try:
                safes = SafeBox.query.filter_by(safe_type=t, is_active=True).all()
                if isinstance(safes, list) and len(safes) == 1 and getattr(safes[0], 'id', None):
                    return int(safes[0].id)
            except Exception:
                return None

            # If multiple active safes exist and none is marked default, pick a
            # stable fallback to avoid blocking invoice creation.
            try:
                sb = (
                    SafeBox.query.filter_by(safe_type=t, is_active=True)
                    .order_by(SafeBox.is_default.desc(), SafeBox.id.asc())
                    .first()
                )
                if sb and getattr(sb, 'id', None):
                    return int(sb.id)
            except Exception:
                return None
            return None

        if auto_settle:
            return pick_default('clearing') or pick_default('bank')
        return pick_default('bank') or pick_default('clearing')

    def _is_cash_payment_method(pm: PaymentMethod | None) -> bool:
        if pm is None:
            return False
        try:
            pt = str(getattr(pm, 'payment_type', '') or '').strip().lower()
            name = str(getattr(pm, 'name', '') or '').strip()
            if pt in {'cash'}:
                return True
            return 'نقد' in name
        except Exception:
            return False

    # Resolve safe box (single source of truth: PaymentMethod -> SafeBox -> Account)
    resolved_safe_box_id = None
    try:
        raw_safe_box_id = data.get('safe_box_id')
        if raw_safe_box_id not in (None, '', False):
            resolved_safe_box_id = int(raw_safe_box_id)
    except Exception:
        resolved_safe_box_id = None

    # invoice.safe_box_id is the invoice-level default (usually cash).
    # For non-cash payment methods (bank transfer, mada, etc.) skip it so the
    # PM's own default_safe_box_id (bank/clearing) is used instead.
    if resolved_safe_box_id is None and _is_cash_payment_method(pm_obj):
        resolved_safe_box_id = invoice.safe_box_id

    if resolved_safe_box_id is None and _is_cash_payment_method(pm_obj):
        # When employee cash safes are enabled, we want cash payments to go to the
        # employee cash safe (or main cash safe) even if the payment method has a default.
        try:
            settings_row = Settings.query.first()
        except Exception:
            settings_row = None
        if bool(getattr(settings_row, 'employee_cash_safes_enabled', False)):
            resolved_safe_box_id = _fallback_cash_safe_box_id()

    if resolved_safe_box_id is None:
        resolved_safe_box_id = getattr(pm_obj, 'default_safe_box_id', None)

    if resolved_safe_box_id is None:
        if _is_cash_payment_method(pm_obj):
            resolved_safe_box_id = _fallback_cash_safe_box_id()
        else:
            resolved_safe_box_id = _fallback_non_cash_safe_box_id(pm_obj)

    # Ultimate fallback: use cash safe as last resort for any payment method.
    if resolved_safe_box_id is None:
        resolved_safe_box_id = _fallback_cash_safe_box_id()

    # Enforce employee cash safe toggle: if employee cash safes are disabled,
    # do not allow routing payments into the employee cash safe (fallback to main cash).
    try:
        settings_row = Settings.query.first()
    except Exception:
        settings_row = None
    try:
        emp_cash_safe_id = None
        emp = getattr(invoice, 'employee', None)
        if not emp and getattr(invoice, 'employee_id', None):
            emp = Employee.query.get(int(invoice.employee_id))
        raw_emp_cash = getattr(emp, 'cash_safe_box_id', None) if emp else None
        if raw_emp_cash not in (None, '', 0, '0', False):
            emp_cash_safe_id = int(raw_emp_cash)
        if (
            emp_cash_safe_id
            and resolved_safe_box_id is not None
            and int(resolved_safe_box_id) == int(emp_cash_safe_id)
            and not bool(getattr(settings_row, 'employee_cash_safes_enabled', False))
        ):
            main_cash = getattr(settings_row, 'main_cash_safe_box_id', None) if settings_row else None
            if main_cash not in (None, '', 0, '0', False):
                resolved_safe_box_id = int(main_cash)
            else:
                try:
                    sb = SafeBox.get_default_by_type('cash')
                    if sb and sb.id:
                        resolved_safe_box_id = int(sb.id)
                except Exception:
                    pass
    except Exception:
        pass

    # شراء من عميل / مرتجع شراء: safe_box_id قد يكون خزينة ذهبية (لتتبع الوزن).
    # سند الصرف النقدي يجب أن يستخدم دائماً خزينة نقدية.
    if (
        str(getattr(invoice, 'invoice_type', '') or '').strip() in ('شراء من عميل', 'مرتجع شراء')
        and resolved_safe_box_id is not None
    ):
        try:
            _pmt_sb = SafeBox.query.get(resolved_safe_box_id)
            if _pmt_sb and ((_pmt_sb.safe_type or '').lower() == 'gold'):
                _pmt_settings = Settings.query.first()
                _pmt_main = getattr(_pmt_settings, 'main_cash_safe_box_id', None) if _pmt_settings else None
                if _pmt_main not in (None, '', 0, '0', False):
                    resolved_safe_box_id = int(_pmt_main)
                else:
                    _pmt_cs = SafeBox.get_default_by_type('cash')
                    if _pmt_cs and _pmt_cs.id:
                        resolved_safe_box_id = int(_pmt_cs.id)
        except Exception:
            pass

    if resolved_safe_box_id is None:
        return jsonify({
            'error': 'missing_safe_box_for_payment_method',
            'message': 'يجب تحديد خزينة (SafeBox) لوسيلة الدفع أو ضبط خزينة افتراضية لها',
            'payment_method_id': pm_id,
            'payment_method_name': getattr(pm_obj, 'name', None),
        }), 400

    # Validate safe box exists (avoid FK failures later; still keep atomic rollback below)
    safe_box_obj = SafeBox.query.get(resolved_safe_box_id)
    if not safe_box_obj:
        return jsonify({
            'error': 'safe_box_not_found',
            'message': 'الخزينة المحددة غير موجودة',
            'safe_box_id': resolved_safe_box_id,
        }), 400

    commission_rate = _to_float(getattr(pm_obj, 'commission_rate', 0.0), 0.0)
    try:
        pm_commission_timing = str(getattr(pm_obj, 'commission_timing', 'invoice') or 'invoice').strip().lower()
    except Exception:
        pm_commission_timing = 'invoice'

    if pm_commission_timing == 'settlement':
        commission_amount = 0.0
        commission_vat = 0.0
        net_amount = amount
    else:
        commission_amount = amount * (commission_rate / 100.0) if commission_rate > 0 else 0.0
        commission_vat = commission_amount * 0.15
        net_amount = amount - commission_amount - commission_vat

    try:
        payment = InvoicePayment(
            invoice_id=invoice.id,
            payment_method_id=pm_id,
            amount=amount,
            commission_rate=commission_rate,
            commission_amount=commission_amount,
            commission_vat=commission_vat,
            net_amount=net_amount,
            notes=notes,
        )

        db.session.add(payment)

        # Also record safe box ledger transaction (رقابة)
        def _direction_for_invoice_type(t: str) -> str:
            t = (t or '').strip()
            if not t:
                return 'in'
            if t == 'بيع':
                return 'in'
            if t == 'مرتجع بيع':
                return 'out'
            if t in ('شراء من عميل', 'شراء'):
                return 'out'
            if t in ('مرتجع شراء', 'مرتجع شراء (مورد)'):
                return 'in'
            # fallback
            return 'in'

        created_by_name = None
        try:
            created_by_name = getattr(g, 'current_user', None).username if getattr(g, 'current_user', None) else None
        except Exception:
            created_by_name = None

        db.session.flush()  # ensure payment.id is available

        # Auto-create + approve a voucher for this payment (for auditing/printing and ledger impact).
        safe_account_id = getattr(safe_box_obj, 'account_id', None)
        if not safe_account_id:
            raise ValueError('safe_box_missing_account_id')
        _warn_if_safe_account_mismatches_payment_method(pm_obj, safe_account_id, context='add_invoice_payment')

        direction = _direction_for_invoice_type(getattr(invoice, 'invoice_type', None))
        voucher_type = 'receipt' if direction == 'in' else 'payment'

        party_type = None
        party_id = None
        party_account_id = None
        if getattr(invoice, 'supplier_id', None):
            party_type = 'supplier'
            party_id = int(invoice.supplier_id)
            supplier = Supplier.query.get(party_id)
            if not supplier:
                raise ValueError('supplier_not_found')
            party_account_id = int(ensure_supplier_accounts(supplier).financial.id)
        elif getattr(invoice, 'customer_id', None):
            party_type = 'customer'
            party_id = int(invoice.customer_id)
            customer = Customer.query.get(party_id)
            if not customer:
                raise ValueError('customer_not_found')
            party_account_id = int(ensure_customer_accounts(customer).financial.id)
        else:
            raise ValueError('missing_party_for_payment_voucher')

        voucher_number = generate_voucher_number(voucher_type)
        voucher_date = datetime.now()
        try:
            voucher_date = invoice.date or voucher_date
        except Exception:
            pass

        voucher_notes = None
        try:
            voucher_notes = json.dumps({
                'source': 'invoice_payment',
                'invoice_id': int(invoice.id),
                'invoice_payment_id': int(payment.id),
                'payment_method_id': int(pm_id),
            }, ensure_ascii=False)
        except Exception:
            voucher_notes = None

        voucher = Voucher(
            voucher_number=voucher_number,
            voucher_type=voucher_type,
            date=voucher_date,
            party_type=party_type,
            customer_id=party_id if party_type == 'customer' else None,
            supplier_id=party_id if party_type == 'supplier' else None,
            amount_cash=float(amount),
            amount_gold=0.0,
            description=f"دفعة فاتورة {getattr(invoice, 'invoice_type_id', '')}".strip(),
            reference_type='invoice',
            reference_id=int(invoice.id),
            reference_number=str(getattr(invoice, 'invoice_type_id', '') or '') or None,
            notes=voucher_notes,
            created_by=created_by_name or 'system',
            status='pending',
        )

        db.session.add(voucher)
        db.session.flush()

        safe_line_type = 'debit' if direction == 'in' else 'credit'
        party_line_type = 'credit' if direction == 'in' else 'debit'

        db.session.add(VoucherAccountLine(
            voucher_id=voucher.id,
            account_id=int(safe_account_id),
            line_type=safe_line_type,
            amount_type='cash',
            amount=float(amount),
            description=notes,
        ))
        db.session.add(VoucherAccountLine(
            voucher_id=voucher.id,
            account_id=int(party_account_id),
            line_type=party_line_type,
            amount_type='cash',
            amount=float(amount),
            description=notes,
        ))
        db.session.flush()

        # ── قيد مجمّع: نجمع كل دفعات الفاتورة في قيد واحد ──
        consolidated_je = _add_payment_lines_to_consolidated_je(
            invoice=invoice,
            voucher=voucher,
            voucher_number=voucher_number,
            safe_account_id=int(safe_account_id),
            party_account_id=int(party_account_id),
            amount=float(amount),
            payment_id=payment.id,
            direction=direction,
            voucher_date=voucher_date,
            created_by=created_by_name or 'system',
        )

        voucher.status = 'approved'
        voucher.approved_at = datetime.now()
        voucher.approved_by = created_by_name or 'system'
        voucher.journal_entry_id = consolidated_je.id

        _append_safe_transactions_for_voucher(voucher, created_by=voucher.approved_by)

        new_paid = paid_amount + amount
        invoice.amount_paid = round(new_paid, 2)
        if invoice.amount_paid >= total_amount - eps:
            invoice.status = 'paid'
        else:
            invoice.status = 'partially_paid'

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"❌ Payment atomicity failure (invoice_id={invoice_id}, safe_box_id={resolved_safe_box_id}): {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': 'payment_post_failed',
            'message': 'تعذر تسجيل الدفعة بسبب فشل تسجيل حركة الخزينة/المعاملة. لم يتم حفظ أي تغييرات.',
        }), 500

    invoice_dict = invoice.to_dict()
    customer_name = (
        invoice.customer.name
        if invoice.customer
        else (invoice.supplier.name if invoice.supplier else "N/A")
    )
    supplier_name = invoice.supplier.name if invoice.supplier else "N/A"
    invoice_dict['customer_name'] = customer_name
    invoice_dict['supplier_name'] = supplier_name
    return jsonify(invoice_dict), 201

@invoices_bp.route('/invoices/<int:invoice_id>/print-template', methods=['PUT'])
def set_invoice_print_template(invoice_id: int):
    """Set per-invoice print template preset key.

    Body JSON supports either:
    - {"preset_key": "a4_portrait"}
    - {"template_preset_key": "a4_portrait"}
    - {"print_template_preset_key": "a4_portrait"}
    - {"clear": true} to unset
    """
    invoice = Invoice.query.get_or_404(invoice_id)
    data = request.get_json(silent=True) or {}

    if bool(data.get('clear')) is True:
        invoice.print_template_preset_key = None
        db.session.commit()
        return jsonify(invoice.to_dict())

    preset_key = (
        data.get('preset_key')
        or data.get('template_preset_key')
        or data.get('print_template_preset_key')
    )
    preset_key = (preset_key or '').strip()
    if not preset_key:
        return jsonify({'error': 'preset_key is required'}), 400

    invoice.print_template_preset_key = preset_key
    db.session.commit()
    return jsonify(invoice.to_dict())

def _create_gold24k_settlement_entries(invoice, posted_by: str = 'system'):
    """
    Create the commission JE for a purchase invoice settled with 24k gold.

    The weight movement is already handled by the existing gold settlement
    system (SafeBoxTransactions + weight JEs from gold_settlements).
    We only add: Dr. Supplier cash account / Cr. إيرادات عمولة السداد بذهب صافي
    """
    if not invoice.gold24k_settlement or not invoice.gold24k_commission_total:
        return

    commission_total = float(invoice.gold24k_commission_total or 0)
    if commission_total <= 0:
        return

    supplier = Supplier.query.get(invoice.supplier_id) if invoice.supplier_id else None
    if not supplier:
        return

    supplier_cash_account_id = supplier.account_id
    revenue_account_id = _ensure_gold24k_commission_revenue_account()

    if not supplier_cash_account_id or not revenue_account_id:
        return

    weight = float(invoice.gold24k_weight or 0)
    je_commission = JournalEntry(
        entry_number=_generate_journal_entry_number(entry_date=invoice.date),
        date=invoice.date,
        description=f'عمولة السداد بذهب صافي — فاتورة #{invoice.invoice_number} ({weight:.3f} جم × {invoice.gold24k_commission_per_gram:.2f} ر.س)',
        reference_type='invoice',
        reference_id=invoice.id,
        is_posted=False,
        created_by=posted_by,
    )
    db.session.add(je_commission)
    db.session.flush()

    db.session.add(JournalEntryLine(
        journal_entry_id=je_commission.id,
        account_id=supplier_cash_account_id,
        cash_debit=commission_total,
        cash_credit=0,
    ))
    db.session.add(JournalEntryLine(
        journal_entry_id=je_commission.id,
        account_id=revenue_account_id,
        cash_debit=0,
        cash_credit=commission_total,
    ))

def _ensure_karat_diff_expense_account():
    """Find or create رسوم السداد بعيار أقل. Production number: 5240."""
    by_number = Account.query.filter_by(account_number='5240').first()
    if by_number:
        return by_number.id

    acct_name = 'رسوم السداد بعيار أقل'
    by_name = Account.query.filter_by(name=acct_name).first()
    if by_name:
        return by_name.id

    parent = Account.query.filter_by(account_number='52').first()
    if not parent:
        parent = Account.query.filter_by(account_number='5').first()

    chosen_number = None
    for candidate in ('5240', '5241', '5242', '5250'):
        if not Account.query.filter_by(account_number=candidate).first():
            chosen_number = candidate
            break
    if not chosen_number:
        chosen_number = '5240'

    account = Account(
        account_number=chosen_number,
        name=acct_name,
        type='expense',
        transaction_type='cash',
        tracks_weight=False,
        parent_id=parent.id if parent else None,
    )
    db.session.add(account)
    db.session.flush()
    return account.id

def _create_karat_diff_settlement_entries(invoice, posted_by: str = 'system'):
    """
    Create commission/fee JEs for per-line karat-difference settlement.

    earn_total > 0 → company EARNS  → Dr. Supplier / Cr. إيرادات 4110
    pay_total  > 0 → company PAYS   → Dr. مصروف 5240 / Cr. Supplier
    """
    earn_total = float(getattr(invoice, 'karat_diff_earn_total', 0) or 0)
    pay_total = float(getattr(invoice, 'karat_diff_pay_total', 0) or 0)

    if earn_total <= 0 and pay_total <= 0:
        return

    supplier = Supplier.query.get(invoice.supplier_id) if invoice.supplier_id else None
    if not supplier or not supplier.account_id:
        return
    supplier_cash_id = supplier.account_id

    if earn_total > 0:
        revenue_acct_id = _ensure_gold24k_commission_revenue_account()
        if revenue_acct_id:
            je = JournalEntry(
                entry_number=_generate_journal_entry_number(entry_date=invoice.date),
                date=invoice.date,
                description=f'عمولة فرق العيار — فاتورة #{invoice.invoice_number}',
                reference_type='invoice',
                reference_id=invoice.id,
                is_posted=False,
                created_by=posted_by,
            )
            db.session.add(je)
            db.session.flush()
            db.session.add(JournalEntryLine(
                journal_entry_id=je.id,
                account_id=supplier_cash_id,
                cash_debit=earn_total, cash_credit=0,
            ))
            db.session.add(JournalEntryLine(
                journal_entry_id=je.id,
                account_id=revenue_acct_id,
                cash_debit=0, cash_credit=earn_total,
            ))

    if pay_total > 0:
        expense_acct_id = _ensure_karat_diff_expense_account()
        if expense_acct_id:
            je = JournalEntry(
                entry_number=_generate_journal_entry_number(entry_date=invoice.date),
                date=invoice.date,
                description=f'رسوم فرق العيار — فاتورة #{invoice.invoice_number}',
                reference_type='invoice',
                reference_id=invoice.id,
                is_posted=False,
                created_by=posted_by,
            )
            db.session.add(je)
            db.session.flush()
            db.session.add(JournalEntryLine(
                journal_entry_id=je.id,
                account_id=expense_acct_id,
                cash_debit=pay_total, cash_credit=0,
            ))
            db.session.add(JournalEntryLine(
                journal_entry_id=je.id,
                account_id=supplier_cash_id,
                cash_debit=0, cash_credit=pay_total,
            ))

def calculate_profit_in_gold(items_sold):
    """
    حساب الربح بالذهب لأصناف مباعة
    
    Args:
        items_sold: قائمة الأصناف المباعة
        مثال: [{'karat': '24', 'weight': 2.0, 'subtotal': 800}, ...]
    
    Returns:
        dict: {
            'total_profit_cash': float,      # الربح النقدي الإجمالي
            'total_profit_gold': float,      # الربح بالذهب الإجمالي (جم)
            'total_cost': float,             # التكلفة الإجمالية
            'details_by_karat': {            # التفاصيل حسب العيار
                '24': {
                    'weight_sold': float,
                    'sale_price': float,
                    'avg_cost_per_gram': float,
                    'total_cost': float,
                    'profit_cash': float,
                    'profit_gold': float,
                    'profit_percentage': float
                }
            }
        }
        
    المعادلة:
        الربح النقدي يعتمد على متوسط تكلفة الجرام
        الربح بالذهب (جم) = الربح النقدي (ر.س) ÷ سعر البيع المباشر للفاتورة (ر.س/جم)
    """
    total_profit_cash = 0.0
    total_profit_gold = 0.0
    total_cost = 0.0
    details_by_karat = {}
    
    for item in items_sold:
        karat = str(item.get('karat', '24'))
        weight = float(item.get('weight', 0))
        sale_price = float(item.get('subtotal', 0))
        
        # 1. حساب متوسط سعر الشراء (تكلفة/جم)
        avg_cost_per_gram = get_inventory_average_cost(karat)
        
        # 2. حساب متوسط سعر البيع (سعر الفاتورة المباشر)
        sale_price_per_gram = (sale_price / weight) if weight > 0 else 0
        
        # 3. حساب التكلفة والربح النقدي باستخدام متوسط التكلفة/جم
        item_cost = weight * avg_cost_per_gram
        profit_cash = (sale_price_per_gram - avg_cost_per_gram) * weight if weight > 0 else 0
        
        # 4. حساب الربح بالذهب باستخدام سعر الفاتورة المباشر
        profit_gold = (profit_cash / sale_price_per_gram) if sale_price_per_gram > 0 else 0
        
        # 5. حساب نسبة الربح
        profit_percentage = (profit_cash / item_cost * 100) if item_cost > 0 else 0
        
        # 6. جمع الإجماليات
        total_profit_cash += profit_cash
        total_profit_gold += profit_gold
        total_cost += item_cost
        
        # 7. حفظ التفاصيل حسب العيار
        if karat not in details_by_karat:
            details_by_karat[karat] = {
                'weight_sold': 0,
                'sale_price': 0,
                'avg_cost_per_gram': avg_cost_per_gram,
                'total_cost': 0,
                'profit_cash': 0,
                'profit_gold': 0,
                'sale_price_per_gram': 0,
                'profit_percentage': 0
            }
        
        details = details_by_karat[karat]
        details['weight_sold'] += weight
        details['sale_price'] += sale_price
        details['total_cost'] += item_cost
        details['profit_cash'] += profit_cash
        details['profit_gold'] += profit_gold
        details['avg_cost_per_gram'] = avg_cost_per_gram
        details['sale_price_per_gram'] = (
            details['sale_price'] / details['weight_sold']
            if details['weight_sold'] > 0 else 0
        )
        details['profit_percentage'] = (
            (details['profit_cash'] / details['total_cost'] * 100)
            if details['total_cost'] > 0 else 0
        )
    
    return {
        'total_profit_cash': round(total_profit_cash, 2),
        'total_profit_gold': round(total_profit_gold, 3),
        'total_cost': round(total_cost, 2),
        'details_by_karat': details_by_karat
    }

@invoices_bp.route('/invoices', methods=['POST'])
def add_invoice():
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid or missing JSON body'}), 400

    # 🆕 خيار أمني: رفض إنشاء الفاتورة بدون توكن
    # يمكن تفعيله من (متغير البيئة) أو من (الإعدادات) عبر الواجهة
    current_user = get_current_user()
    auth_required = bool(REQUIRE_AUTH_FOR_INVOICE_CREATE)
    if not auth_required:
        try:
            settings = Settings.query.first()
            auth_required = bool(getattr(settings, 'require_auth_for_invoice_create', False)) if settings else False
        except Exception:
            auth_required = bool(REQUIRE_AUTH_FOR_INVOICE_CREATE)

    # 🆕 ضبط السماح بالدفع الجزئي/البيع الآجل
    allow_partial_payments = False
    try:
        env_flag = str(os.getenv('ALLOW_PARTIAL_INVOICE_PAYMENTS', '')).strip().lower()
        if env_flag in ('1', 'true', 'yes', 'on'):
            allow_partial_payments = True
    except Exception:
        allow_partial_payments = False

    if not allow_partial_payments:
        try:
            settings_row = Settings.query.first()
            allow_partial_payments = bool(getattr(settings_row, 'allow_partial_invoice_payments', False)) if settings_row else False
        except Exception:
            allow_partial_payments = False

    if auth_required and not current_user:
        return jsonify({'error': 'Authentication required to create invoices'}), 401

    # 🆕 الحصول على سعر الذهب الحالي في بداية الدالة (يُستخدم في عدة أماكن)
    gold_price_data = get_current_gold_price()

    # --- VAT policy helpers (server-side enforcement) ---
    def _normalize_tax_rate(raw_value, fallback=0.15):
        try:
            val = float(raw_value)
        except Exception:
            val = float(fallback)
        # Support both 0.15 and 15 representations.
        if val > 1.0:
            val = val / 100.0
        if val < 0:
            val = abs(val)
        return val

    def _parse_vat_exempt_karats(settings_row):
        allowed = {18, 21, 22, 24}
        default = {24}
        if not settings_row:
            return default
        raw = getattr(settings_row, 'vat_exempt_karats', None)
        if raw in (None, '', False):
            return default
        try:
            import json
            decoded = json.loads(raw) if isinstance(raw, str) else raw
            if isinstance(decoded, (list, tuple, set)):
                out = set()
                for v in decoded:
                    try:
                        k = int(str(v).strip())
                    except Exception:
                        continue
                    if k in allowed:
                        out.add(k)
                return out or default
        except Exception:
            pass

        if isinstance(raw, str):
            out = set()
            for part in raw.split(','):
                try:
                    k = int(part.strip())
                except Exception:
                    continue
                if k in allowed:
                    out.add(k)
            return out or default

        return default

    # Snapshot VAT settings once per request.
    settings_row = None
    try:
        settings_row = Settings.query.first()
    except Exception:
        settings_row = None

    # Snapshot posting policy early because payment validation happens before
    # we compute approval gates/unposted_mode later in the flow.
    auto_post_invoices_enabled = True
    try:
        auto_post_invoices_enabled = bool(getattr(settings_row, 'auto_post_invoices', True)) if settings_row else True
    except Exception:
        auto_post_invoices_enabled = True

    vat_enabled = True
    vat_rate = 0.15
    vat_exempt_karats = {24}
    try:
        vat_enabled = bool(getattr(settings_row, 'tax_enabled', True)) if settings_row else True
        vat_rate = _normalize_tax_rate(getattr(settings_row, 'tax_rate', 0.15) if settings_row else 0.15, fallback=0.15)
        vat_exempt_karats = _parse_vat_exempt_karats(settings_row)
    except Exception:
        vat_enabled = True
        vat_rate = 0.15
        vat_exempt_karats = {24}

    # دعم كل من invoice_type و transaction_type للتوافق مع الشاشات المختلفة
    invoice_type = data.get('invoice_type')
    transaction_type = data.get('transaction_type')
    gold_type = data.get('gold_type', 'new')
    
    if not invoice_type:
        # إذا كان transaction_type موجود، استخدمه لتحديد invoice_type
        transaction_type = transaction_type or 'sell'
        if transaction_type == 'sell':
            invoice_type = 'بيع'
        elif transaction_type == 'buy':
            # تحديد نوع الشراء بناءً على gold_type ووجود supplier_id
            if gold_type == 'new' or data.get('supplier_id'):
                invoice_type = 'شراء'
            else:
                invoice_type = 'شراء من عميل'
        else:
            invoice_type = 'بيع'  # افتراضي
    elif isinstance(invoice_type, str):
        invoice_type = invoice_type.strip()

        # Canonicalize legacy supplier purchase/return labels by keywords.
        if 'مورد' in invoice_type and 'شراء' in invoice_type:
            if 'مرتجع' in invoice_type:
                invoice_type = 'مرتجع شراء (مورد)'
            else:
                invoice_type = 'شراء'

        # Some legacy clients may send supplier returns as 'مرتجع شراء' but include supplier_id.
        # Canonicalize to supplier return so optional original_invoice_id rules apply.
        if (
            invoice_type == 'مرتجع شراء'
            and data.get('supplier_id')
            and (data.get('gold_type', 'new') == 'new')
        ):
            invoice_type = 'مرتجع شراء (مورد)'

        if invoice_type == 'شراء':
            # 'شراء' can represent either supplier purchase (worked gold) or
            # scrap purchase, depending on gold_type.
            # ملاحظة: Flutter قد يرسل customer_id حتى للمورد، لذا نعتمد على gold_type
            if gold_type != 'new':
                invoice_type = 'شراء من عميل'
            else:
                # نقل customer_id إلى supplier_id إذا لم يكن supplier_id موجوداً
                if not data.get('supplier_id') and data.get('customer_id'):
                    print("⚠️ Converting customer_id to supplier_id for supplier purchase")
                    data['supplier_id'] = data.pop('customer_id')
    
    if not invoice_type:
        return jsonify({'error': 'invoice_type or transaction_type is required'}), 400

    def _to_positive_int(raw_value):
        try:
            if raw_value in (None, '', False):
                return None
            parsed = int(raw_value)
            return parsed if parsed > 0 else None
        except Exception:
            return None

    def _normalize_text(raw_value):
        try:
            return ' '.join(str(raw_value or '').strip().split()).lower()
        except Exception:
            return str(raw_value or '').strip().lower()

    def _contains_cash_keyword(raw_value):
        value = _normalize_text(raw_value)
        if not value:
            return False
        return ('نقد' in value) or ('كاش' in value) or ('cash' in value)

    def _resolve_default_cash_customer_id():
        """Resolve a stable default cash customer id without creating new rows."""
        try:
            candidates = Customer.query.order_by(Customer.active.desc(), Customer.id.asc()).all()
        except Exception:
            candidates = []

        aliases = {
            'عميل نقدي',
            'نقدي',
            'عميل كاش',
            'cash customer',
            'cash',
        }

        for customer in candidates:
            try:
                cid = int(getattr(customer, 'id', 0) or 0)
            except Exception:
                cid = 0
            if cid <= 0:
                continue
            name_norm = _normalize_text(getattr(customer, 'name', ''))
            if name_norm in aliases:
                return cid

        for customer in candidates:
            try:
                cid = int(getattr(customer, 'id', 0) or 0)
            except Exception:
                cid = 0
            if cid <= 0:
                continue

            if _contains_cash_keyword(getattr(customer, 'name', '')) or _contains_cash_keyword(
                getattr(customer, 'customer_code', '')
            ):
                return cid

        return None

    # Normalize customer_id early to prevent FK errors from sentinel values (e.g. -1).
    _raw_customer_id = data.get('customer_id')
    _normalized_customer_id = _to_positive_int(_raw_customer_id)
    if _normalized_customer_id is not None:
        try:
            _customer_exists = Customer.query.get(int(_normalized_customer_id)) is not None
        except Exception:
            _customer_exists = False
        if not _customer_exists:
            _normalized_customer_id = None

    # Sales and customer-scrap purchases can safely fallback to a configured cash customer.
    if _normalized_customer_id is None and invoice_type in ('بيع', 'شراء من عميل'):
        _fallback_cash_customer_id = _resolve_default_cash_customer_id()
        if _fallback_cash_customer_id is None:
            return jsonify({
                'error': 'customer_required',
                'message': 'تعذر تحديد عميل صالح للفاتورة. يرجى اختيار عميل موجود أو إنشاء "عميل نقدي".',
            }), 400
        _normalized_customer_id = int(_fallback_cash_customer_id)

    if _normalized_customer_id is not None:
        data['customer_id'] = int(_normalized_customer_id)
    elif _raw_customer_id not in (None, '', False):
        # Clear invalid/unusable customer ids so we never hit DB-level FK failures later.
        data['customer_id'] = None
    
    # 🆕 Validation للمرتجعات
    return_types = ['مرتجع بيع', 'مرتجع شراء', 'مرتجع شراء (مورد)']
    if invoice_type in return_types:
        original_invoice = None

        # NOTE: Supplier purchase return can be created without original_invoice_id
        # for legacy/very old invoices that are not available for selection.
        require_original = not (
            invoice_type == 'مرتجع شراء (مورد)'
            and not data.get('original_invoice_id')
        )

        if require_original:
            # التحقق من وجود original_invoice_id
            if not data.get('original_invoice_id'):
                return jsonify({'error': 'original_invoice_id is required for return invoices'}), 400

            # التحقق من وجود الفاتورة الأصلية
            original_invoice = Invoice.query.get(data['original_invoice_id'])
            if not original_invoice:
                return jsonify({'error': f'Original invoice with ID {data["original_invoice_id"]} not found'}), 404

            # التحقق من تطابق العميل/المورد
            if invoice_type == 'مرتجع بيع' and original_invoice.invoice_type == 'بيع':
                if original_invoice.customer_id != data.get('customer_id'):
                    return jsonify({'error': 'Customer ID must match original invoice'}), 400
            elif invoice_type == 'مرتجع شراء' and original_invoice.invoice_type == 'شراء من عميل':
                if original_invoice.customer_id != data.get('customer_id'):
                    return jsonify({'error': 'Customer ID must match original invoice'}), 400
            elif invoice_type == 'مرتجع شراء (مورد)':
                original_type = (original_invoice.invoice_type or '').strip()
                is_supplier_purchase = (
                    original_type == 'شراء'
                    or (
                        'مورد' in original_type
                        and 'شراء' in original_type
                        and 'مرتجع' not in original_type
                    )
                )
                if is_supplier_purchase and original_invoice.supplier_id != data.get('supplier_id'):
                    return jsonify({'error': 'Supplier ID must match original invoice'}), 400

        # 🛡️ Server-side validation (when original invoice exists): return items must reference
        # original invoice items and cannot exceed the original quantities/weights.
        if original_invoice is not None:
            try:
                def _to_float_local(value, default=0.0):
                    if value in (None, '', False):
                        return default
                    try:
                        normalized = normalize_number(str(value))
                        return float(normalized)
                    except Exception:
                            try:
                                return float(value)
                            except Exception:
                                return default

                items_payload = data.get('items', [])
                if items_payload in (None, '', False):
                    items_payload = []
                if not isinstance(items_payload, list):
                    return jsonify({'error': 'items must be a list'}), 400

                # Aggregate by original_invoice_item_id to prevent splitting to bypass limits.
                aggregated = {}  # id -> {'qty': float, 'total_weight': float}
                for idx, it in enumerate(items_payload, start=1):
                    if not isinstance(it, dict):
                        continue
                    raw_orig_item_id = it.get('original_invoice_item_id')
                    if raw_orig_item_id in (None, '', False):
                        return jsonify({
                            'error': 'missing_original_invoice_item_id',
                            'message': 'يجب ربط كل صنف مرتجع بصنف من الفاتورة الأصلية (original_invoice_item_id).',
                            'line_index': idx,
                        }), 400
                    try:
                        orig_item_id = int(raw_orig_item_id)
                    except Exception:
                        return jsonify({
                            'error': 'invalid_original_invoice_item_id',
                            'message': 'original_invoice_item_id غير صحيح',
                            'line_index': idx,
                            'value': raw_orig_item_id,
                        }), 400

                    req_qty = _to_float_local(it.get('quantity', 1), 1.0) or 0.0
                    if req_qty < 0:
                        req_qty = abs(req_qty)
                    # Return payload uses per-unit weight.
                    req_weight_per_unit = _to_float_local(it.get('weight', it.get('total_weight', 0.0)), 0.0)
                    if req_weight_per_unit < 0:
                        req_weight_per_unit = abs(req_weight_per_unit)

                    req_total_weight = float(req_weight_per_unit) * float(req_qty if req_qty > 0 else 0.0)
                    current = aggregated.get(orig_item_id) or {'qty': 0.0, 'total_weight': 0.0}
                    current['qty'] = float(current.get('qty', 0.0) or 0.0) + float(req_qty)
                    current['total_weight'] = float(current.get('total_weight', 0.0) or 0.0) + float(req_total_weight)
                    aggregated[orig_item_id] = current

                # Validate aggregated quantities/weights against DB original items.
                eps_qty = 1e-6
                eps_weight = 1e-3
                for orig_item_id, agg in aggregated.items():
                    orig_item = InvoiceItem.query.get(orig_item_id)
                    if not orig_item or int(getattr(orig_item, 'invoice_id', 0) or 0) != int(original_invoice.id):
                        return jsonify({
                            'error': 'original_invoice_item_not_found',
                            'message': 'أحد أصناف المرتجع غير موجود ضمن الفاتورة الأصلية',
                            'original_invoice_item_id': orig_item_id,
                            'original_invoice_id': int(original_invoice.id),
                        }), 400

                    try:
                        orig_qty = int(getattr(orig_item, 'quantity', 1) or 1)
                    except Exception:
                        orig_qty = 1
                    if orig_qty < 0:
                        orig_qty = abs(orig_qty)

                    try:
                        orig_weight_per_unit = float(getattr(orig_item, 'weight', 0.0) or 0.0)
                    except Exception:
                        orig_weight_per_unit = 0.0

                    orig_total_weight = float(orig_weight_per_unit) * float(orig_qty if orig_qty > 0 else 0.0)

                    req_qty_sum = float(agg.get('qty', 0.0) or 0.0)
                    req_weight_sum = float(agg.get('total_weight', 0.0) or 0.0)

                    if req_qty_sum > float(orig_qty) + eps_qty:
                        return jsonify({
                            'error': 'return_quantity_exceeds_original',
                            'message': 'كمية المرتجع أكبر من كمية الفاتورة الأصلية',
                            'original_invoice_item_id': orig_item_id,
                            'original_quantity': int(orig_qty),
                            'returned_quantity': float(req_qty_sum),
                        }), 400

                    # Only enforce weight if the original invoice item has a meaningful weight.
                    if orig_total_weight > eps_weight and req_weight_sum > orig_total_weight + eps_weight:
                        return jsonify({
                            'error': 'return_weight_exceeds_original',
                            'message': 'وزن المرتجع أكبر من وزن الفاتورة الأصلية',
                            'original_invoice_item_id': orig_item_id,
                            'original_total_weight': round(float(orig_total_weight), 3),
                            'returned_total_weight': round(float(req_weight_sum), 3),
                        }), 400
            except Exception:
                # Do not crash invoice creation on validation failures; fall back to existing behavior.
                pass
    
    # 🆕 Validation لنوع الذهب
    gold_type = data.get('gold_type', 'new')
    # For return invoices: inherit gold_type from the original invoice when not explicitly provided.
    # This ensures مرتجع بيع on a scrap sale uses the scrap inventory account (1310),
    # and مرتجع شراء correctly marks the safe-box weight transaction as scrap.
    if invoice_type in ('مرتجع بيع', 'مرتجع شراء', 'مرتجع شراء (مورد)') and not data.get('gold_type'):
        try:
            _orig_inv = original_invoice or (
                Invoice.query.get(data['original_invoice_id']) if data.get('original_invoice_id') else None
            )
            if _orig_inv:
                _inherited = str(getattr(_orig_inv, 'gold_type', 'new') or 'new').strip().lower()
                if _inherited in ('new', 'scrap'):
                    gold_type = _inherited
        except Exception:
            pass
    if gold_type not in ['new', 'scrap']:
        return jsonify({'error': 'gold_type must be either "new" or "scrap"'}), 400
    
    # 🆕 دعم وسائل دفع متعددة في الفاتورة الواحدة
    # يمكن إرسال إما:
    # 1. payment_method_id (وسيلة واحدة - للتوافق)
    # 2. payments (array من وسائل متعددة - الميزة الجديدة)

    def _to_float_request(value, default=0.0):
        if value in (None, '', False):
            return default
        try:
            normalized = normalize_number(str(value))
            return float(normalized)
        except (TypeError, ValueError):
            try:
                return float(value)
            except (TypeError, ValueError):
                return default
    
    payment_method_id = data.get('payment_method_id')  # للتوافق مع الكود القديم
    safe_box_id = data.get('safe_box_id')
    payments_data = data.get('payments', [])  # 🆕 دعم وسائل متعددة
    payment_method_obj = None  # نستخدمه لاحقاً عند الحاجة للخزينة الافتراضية
    karat_lines_data = data.get('karat_lines', [])

    # 🧩 Backward-compat: resolve legacy payment_method string into payment_method_id.
    # Some Flutter screens still send payment_method as a label (cash/card/transfer/deferred)
    # without payment_method_id. We map best-effort to an active PaymentMethod.
    if (not payment_method_id) and (not payments_data) and isinstance(data.get('payment_method'), str):
        legacy_label = (data.get('payment_method') or '').strip()
        if legacy_label:
            legacy_norm = legacy_label.lower()

            def _pick_payment_method(q):
                try:
                    return (
                        q.filter_by(is_active=True)
                        .order_by(PaymentMethod.display_order.asc(), PaymentMethod.id.asc())
                        .first()
                    )
                except Exception:
                    return None

            pm_guess = None
            try:
                # Arabic keywords
                if ('آجل' in legacy_label) or ('اجل' in legacy_label) or (legacy_norm in {'deferred', 'credit', 'receivable', 'on_account'}):
                    pm_guess = _pick_payment_method(PaymentMethod.query.filter(PaymentMethod.payment_type.in_(['receivable', 'credit', 'on_account', 'ar'])))
                elif ('نقد' in legacy_label) or (legacy_norm in {'cash', 'cash_payment'}):
                    pm_guess = _pick_payment_method(PaymentMethod.query.filter(PaymentMethod.payment_type.in_(['cash'])))
                    if not pm_guess:
                        pm_guess = _pick_payment_method(PaymentMethod.query.filter(PaymentMethod.name.contains('نقد')))
                elif ('تحويل' in legacy_label) or ('بنك' in legacy_label) or (legacy_norm in {'transfer', 'bank', 'bank_transfer'}):
                    pm_guess = _pick_payment_method(PaymentMethod.query.filter(PaymentMethod.name.contains('تحويل')))
                    if not pm_guess:
                        pm_guess = _pick_payment_method(PaymentMethod.query.filter(PaymentMethod.name.contains('بنك')))
                elif ('بطاق' in legacy_label) or (legacy_norm in {'card', 'mada', 'visa', 'mastercard'}):
                    pm_guess = _pick_payment_method(PaymentMethod.query.filter(PaymentMethod.payment_type.in_(['mada', 'visa', 'mastercard', 'card'])))
            except Exception:
                pm_guess = None

            if pm_guess and getattr(pm_guess, 'id', None):
                payment_method_id = int(pm_guess.id)
                # Keep data consistent for downstream logic.
                data['payment_method_id'] = payment_method_id

    # 🛡️ Asset Protection: strict payload validation (weight integrity)
    # منع إرسال items و karat_lines معاً في الحالات التي تؤثر على الأوزان
    # (خصوصاً شراء الكسر/المقايضة) لتفادي التلاعب أو الازدواجية.
    def _has_any_weight_in_items(items_list):
        if not items_list or not isinstance(items_list, list):
            return False
        for it in items_list:
            if not isinstance(it, dict):
                continue
            w = _to_float_request(it.get('weight', it.get('total_weight')), 0.0)
            q = _to_float_request(it.get('quantity', 1), 1.0)
            if w > 0 and (q if q > 0 else 1.0) > 0:
                return True
        return False

    def _has_any_weight_in_karat_lines(lines_list):
        if not lines_list or not isinstance(lines_list, list):
            return False
        for ln in lines_list:
            if not isinstance(ln, dict):
                continue
            w = _to_float_request(
                ln.get('weight_grams', ln.get('weight', ln.get('total_weight'))),
                0.0,
            )
            k = _to_float_request(ln.get('karat'), 0.0)
            if w > 0 and k > 0:
                return True
        return False

    def _is_weight_sensitive_context(payload: dict) -> bool:
        try:
            gt = str((payload.get('gold_type') or 'new')).strip().lower()
        except Exception:
            gt = 'new'
        try:
            inv_t = str((payload.get('invoice_type') or '')).strip()
        except Exception:
            inv_t = ''
        try:
            bt = _to_float_request(payload.get('barter_total', 0.0), 0.0)
        except Exception:
            bt = 0.0
        # Any barter/counterflow flags we know about.
        has_barter_link = payload.get('barter_sale_invoice_id') not in (None, '', False)
        try:
            settled_gold_w = _to_float_request(payload.get('settled_gold_weight', 0.0), 0.0)
        except Exception:
            settled_gold_w = 0.0

        # New: allow multiple gold settlements (multi-safe).
        # Payload example: gold_settlements: [{'safe_box_id': 1, 'karat': 21, 'weight': 1.234}, ...]
        gold_settlements_w = 0.0
        try:
            raw = payload.get('gold_settlements')
            if isinstance(raw, list):
                for ln in raw:
                    if isinstance(ln, dict):
                        gold_settlements_w += _to_float_request(
                            ln.get('weight', ln.get('weight_grams', 0.0)),
                            0.0,
                        )
        except Exception:
            gold_settlements_w = 0.0

        # Scrap purchase and barter are the primary high-risk weight contexts.
        if gt == 'scrap':
            return True
        if inv_t == 'شراء من عميل':
            return True
        if bt > 0.01 or has_barter_link or settled_gold_w > 0.0 or gold_settlements_w > 0.0:
            return True
        return False

    try:
        items_payload = data.get('items', [])
        weight_sensitive = _is_weight_sensitive_context(data)
        if weight_sensitive and _has_any_weight_in_items(items_payload) and _has_any_weight_in_karat_lines(karat_lines_data):
            return jsonify({
                'error': 'payload_conflict_weight_sources',
                'message': 'لا يمكن الجمع بين items و karat_lines في الفواتير الحساسة للأوزان (كسر/مقايضة). اختر مصدراً واحداً للوزن.',
            }), 400
    except Exception:
        # Do not block invoice creation if validation itself fails unexpectedly.
        pass

    # If the client supplied valid karat lines, treat them as the single source of truth
    # for gold weight totals to avoid double-counting with items (some clients send both).
    has_valid_karat_lines = False
    try:
        if karat_lines_data and isinstance(karat_lines_data, list):
            for _line in karat_lines_data:
                if not isinstance(_line, dict):
                    continue
                _k = _to_float_request(_line.get('karat'), 0.0)
                _w = _to_float_request(
                    _line.get('weight_grams', _line.get('weight', _line.get('total_weight'))),
                    0.0,
                )
                if _k > 0 and _w > 0:
                    has_valid_karat_lines = True
                    break
    except Exception:
        has_valid_karat_lines = False

    # 🆕 Branch dimension (separate from offices; offices are closing offices/suppliers)
    branch_id = data.get('branch_id')
    if branch_id not in (None, '', False):
        try:
            branch_id = int(branch_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'branch_id must be numeric'}), 400
        try:
            from models import Branch
            branch_row = Branch.query.get(branch_id)
            if not branch_row:
                return jsonify({'error': f'Branch with ID {branch_id} not found'}), 404
            if hasattr(branch_row, 'active') and not bool(getattr(branch_row, 'active', True)):
                return jsonify({'error': 'Selected branch is not active'}), 400
        except Exception:
            # In case branch subsystem is unavailable, still allow invoice creation.
            pass

    # 🆕 Office (closing office) - used for gold closing/reservations, not branch.
    office_id = data.get('office_id')
    if office_id not in (None, '', False):
        try:
            office_id = int(office_id)
        except (TypeError, ValueError):
            return jsonify({'error': 'office_id must be numeric'}), 400
        try:
            office_row = Office.query.get(office_id)
            if not office_row:
                return jsonify({'error': f'Office with ID {office_id} not found'}), 404
            if hasattr(office_row, 'active') and not bool(getattr(office_row, 'active', True)):
                return jsonify({'error': 'Selected office is not active'}), 400
        except Exception:
            # If offices subsystem is unavailable for some reason, still allow invoice creation.
            pass

    # Gate B — POS claim protocol (T2.2: INV-4 ENFORCED, ADR-016 §H1).
    # Replaces the pre-transaction availability read with an atomic claim request.
    # Commerce grants exclusive intent inside its own transaction; the ERP may
    # not write until the grant is received.
    # Fail-open: Commerce API timeout → sale proceeds + WARNING (legacy H2 path).
    _pos_claims: list[tuple[int, str]] = []   # [(item_id, claim_id), ...]
    _pos_claims_confirmed = False

    if invoice_type == 'بيع':
        try:
            from services.commerce_availability import (
                request_pos_claim as _request_pos_claim,
                _release_pos_claims_best_effort,
            )
            _items_to_claim = [
                item.get('item_id')
                for item in (data.get('items') or [])
                if item.get('item_id') is not None
            ]
            for _item_id in _items_to_claim:
                _result = _request_pos_claim(int(_item_id))
                if _result.denied:
                    # Release any claims already granted for this invoice, then
                    # return 409 with ZERO writes (C3b on the ERP side).
                    _release_pos_claims_best_effort(_pos_claims)
                    _pos_claims.clear()
                    return jsonify({
                        'error': 'item_pos_blocked',
                        'message': _result.blocked_reason,
                        'item_id': _item_id,
                        'block_type': _result.block_type,
                        'reserved_until': _result.reserved_until,
                    }), 409
                elif _result.claim_id:
                    _pos_claims.append((int(_item_id), _result.claim_id))
        except Exception as _gate_b_exc:
            # Fail-open: release partial claims; proceed without any claim.
            # The TOCTOU risk re-appears briefly (same as pre-T2.2 Gate B).
            import logging as _log_module
            _log_module.getLogger(__name__).warning(
                "gate_b: pos-claim error — failing open: %s", _gate_b_exc
            )
            try:
                from services.commerce_availability import _release_pos_claims_best_effort
                _release_pos_claims_best_effort(_pos_claims)
            except Exception:
                pass
            _pos_claims.clear()

    commission_amount = 0.0
    commission_vat_total = 0.0
    data_total = _to_float_request(data.get('total', 0.0))
    net_amount = data_total  # قد يكون محسوباً مسبقاً أو سيحسب من items

    # 🆕 Barter support: allow partial cash payments when part of the sale is settled via gold barter (offset).
    # The client should send `barter_total` (cash-equivalent) when barter is used.
    try:
        barter_total = _to_float_request(data.get('barter_total', 0.0))
    except Exception:
        barter_total = 0.0

    # 🆕 Scrap custody identity (who holds received scrap gold).
    # We intentionally do NOT use SafeBox for this.
    scrap_holder_employee_id = None
    try:
        raw_she = data.get('scrap_holder_employee_id')
        if raw_she not in (None, '', False):
            scrap_holder_employee_id = int(raw_she)
    except Exception:
        scrap_holder_employee_id = None

    # If this is a scrap purchase invoice and custody holder isn't specified,
    # default to the invoice employee_id (who executed the transaction).
    try:
        inv_type_key = str((data.get('invoice_type') or '')).strip()
        gold_type_key = str((data.get('gold_type') or '')).strip().lower()
        if scrap_holder_employee_id is None and inv_type_key == 'شراء من عميل' and gold_type_key == 'scrap':
            raw_emp = data.get('employee_id')
            if raw_emp not in (None, '', False):
                scrap_holder_employee_id = int(raw_emp)
    except Exception:
        pass

    allow_partial_for_barter_sale = False
    try:
        raw_invoice_type = data.get('invoice_type')
        if isinstance(raw_invoice_type, str):
            raw_invoice_type = raw_invoice_type.strip()
        bt = _to_float_request(data.get('barter_total', 0.0))
        allow_partial_for_barter_sale = (raw_invoice_type == 'بيع') and (bt > 0.01)
    except Exception:
        allow_partial_for_barter_sale = False
    
    # إذا كانت هناك وسائل دفع متعددة
    if payments_data and isinstance(payments_data, list) and len(payments_data) > 0:
        total_payments = sum(_to_float_request(p.get('amount', 0.0)) for p in payments_data)
        effective_settled = total_payments + (barter_total or 0.0)
        # In unposted workflows (auto-post disabled), allow saving partial settlements
        # even if partial payments are disabled globally.
        partial_allowed_for_request = bool(allow_partial_payments) or (not bool(auto_post_invoices_enabled))
        # التحقق من الدفعات مقابل إجمالي الفاتورة
        if data_total > 0:
            if partial_allowed_for_request:
                # ✅ السماح بالدفع الجزئي طالما لا يوجد تجاوز
                if (effective_settled - data_total) > 0.01:  # tolerance للفواصل العشرية
                    return jsonify({
                        'error': f'مجموع المبالغ ({effective_settled}) أكبر من إجمالي الفاتورة ({data_total})'
                    }), 400
            else:
                # ❌ الوضع الافتراضي: يجب أن يساوي مجموع الدفعات إجمالي الفاتورة
                if abs(effective_settled - data_total) > 0.01:  # tolerance للفواصل العشرية
                    return jsonify({
                        'error': f'مجموع المبالغ ({effective_settled}) لا يساوي إجمالي الفاتورة ({data_total})'
                    }), 400

        # 🆕 مزامنة amount_paid مع مجموع الدفعات إذا لم يُرسل أو كان غير متطابق.
        if 'amount_paid' not in data or data.get('amount_paid') in (None, '', False):
            data['amount_paid'] = total_payments
        else:
            body_paid = _to_float_request(data.get('amount_paid', 0.0))
            if abs(body_paid - total_payments) > 0.01:
                data['amount_paid'] = total_payments
        
        # حساب إجمالي العمولات
        for payment in payments_data:
            pm_id = payment.get('payment_method_id')
            pm_amount = _to_float_request(payment.get('amount', 0.0))
            
            if not pm_id:
                return jsonify({'error': 'payment_method_id is required for each payment'}), 400
            
            pm_obj = PaymentMethod.query.get(pm_id)
            if not pm_obj:
                return jsonify({'error': f'Payment method with ID {pm_id} not found'}), 404
            
            if not pm_obj.is_active:
                return jsonify({'error': f'Payment method "{pm_obj.name}" is not active'}), 400

            # Commission policy:
            # - invoice (default): commission is recorded at invoice time
            # - settlement: do not record commission at invoice time
            try:
                pm_commission_timing = str(getattr(pm_obj, 'commission_timing', 'invoice') or 'invoice').strip().lower()
            except Exception:
                pm_commission_timing = 'invoice'
            
            # حساب عمولة هذه الدفعة
            pm_commission_rate = _to_float_request(
                payment.get('commission_rate', pm_obj.commission_rate if pm_obj else 0.0)
            )

            pm_commission_fixed_amount = _to_float_request(
                payment.get(
                    'commission_fixed_amount',
                    getattr(pm_obj, 'commission_fixed_amount', 0.0) if pm_obj else 0.0,
                )
            )

            if pm_commission_timing == 'settlement':
                pm_commission_amount = 0.0
                pm_commission_vat = 0.0
            else:
                if 'commission_amount' in payment:
                    pm_commission_amount = _to_float_request(payment.get('commission_amount', 0.0))
                else:
                    pm_commission_amount = (
                        (pm_commission_fixed_amount if pm_commission_fixed_amount > 0 else 0.0)
                        + (pm_amount * (pm_commission_rate / 100) if pm_commission_rate > 0 else 0.0)
                    )

                pm_commission_vat = _to_float_request(
                    payment.get('commission_vat', pm_commission_amount * 0.15)
                )

            commission_amount += pm_commission_amount
            commission_vat_total += pm_commission_vat

        # ملاحظة: net_amount تاريخياً يمثل صافي قيمة الفاتورة بعد العمولات.
        # عند تفعيل الدفع الجزئي، لا يمكن معرفة عمولة الجزء غير المدفوع (وسيلة الدفع غير معروفة بعد)
        # لذلك نترك net_amount = إجمالي الفاتورة إذا كانت الدفعات أقل من الإجمالي.
        gross_amount = data_total if data_total > 0 else total_payments
        if partial_allowed_for_request and data_total > 0 and effective_settled < (data_total - 0.01):
            net_amount = data_total
        else:
            net_amount = gross_amount - commission_amount - commission_vat_total
    
    # وسيلة دفع واحدة (للتوافق مع الكود القديم)
    elif payment_method_id:
        payment_method_obj = PaymentMethod.query.get(payment_method_id)
        if not payment_method_obj:
            return jsonify({'error': f'Payment method with ID {payment_method_id} not found'}), 404
        
        if not payment_method_obj.is_active:
            return jsonify({'error': f'Payment method "{payment_method_obj.name}" is not active'}), 400
        
        # حساب العمولة
        try:
            pm_commission_timing = str(getattr(payment_method_obj, 'commission_timing', 'invoice') or 'invoice').strip().lower()
        except Exception:
            pm_commission_timing = 'invoice'

        if pm_commission_timing != 'settlement':
            try:
                fixed_amount = float(getattr(payment_method_obj, 'commission_fixed_amount', 0.0) or 0.0)
            except Exception:
                fixed_amount = 0.0

            rate_amount = 0.0
            try:
                if payment_method_obj.commission_rate and payment_method_obj.commission_rate > 0:
                    rate_amount = data_total * (payment_method_obj.commission_rate / 100)
            except Exception:
                rate_amount = 0.0

            commission_amount = (fixed_amount if fixed_amount > 0 else 0.0) + (rate_amount if rate_amount > 0 else 0.0)
            if commission_amount > 0:
                commission_vat_total = commission_amount * 0.15
                net_amount = data_total - commission_amount - commission_vat_total
    
    wage_mode_snapshot = _get_manufacturing_wage_mode()
    # Capture the current max SafeBoxTransaction id before we create any new ones.
    # This is used below to filter phantom old SBTs that happen to share the same
    # invoice_id as the new invoice (coincidental ID collision from old data).
    try:
        _max_sbt_id_before = db.session.query(db.func.max(SafeBoxTransaction.id)).scalar() or 0
    except Exception:
        _max_sbt_id_before = 0
    try:
        # --- 1. Create Invoice and Items ---
        next_invoice_type_id = _next_invoice_type_id([invoice_type])

        def _extract_float(key, default=0.0):
            if key not in data:
                return default
            try:
                normalized = normalize_number(str(data.get(key, default)))
                return float(normalized)
            except Exception:
                try:
                    return float(data.get(key, default))
                except Exception:
                    return default

        def _to_float(value, default=0.0):
            if value in (None, '', False):
                return default
            try:
                normalized = normalize_number(str(value))
                return float(normalized)
            except (TypeError, ValueError):
                try:
                    return float(value)
                except (TypeError, ValueError):
                    return default

        # 🆕 الحصول على المستخدم الحالي وتعيينه لـ posted_by
        # عند تفعيل auth_required لا نسمح بـ fallback من body
        # ملاحظة: نُفضل اسم الموظف المربوط بالحساب (إن وجد) وإلا نستخدم الاسم/المستخدم.
        posted_by_username = None
        employee_id_for_invoice = None

        # Admin-only import override: allow setting employee_id/posted_by from payload
        # but only when explicitly requested by a guarded flag.
        allow_employee_override = False
        try:
            allow_employee_override = bool(data.get('allow_employee_override', False))
        except Exception:
            allow_employee_override = False
        if current_user:
            employee_display_name = None

            # 1) Direct relationship: current_user.employee (AppUser)
            linked_employee = getattr(current_user, 'employee', None)
            if linked_employee:
                if getattr(linked_employee, 'id', None):
                    employee_id_for_invoice = linked_employee.id
                if getattr(linked_employee, 'name', None):
                    employee_display_name = linked_employee.name

            # 2) employee_id attribute (AppUser)
            if employee_id_for_invoice is None:
                try:
                    employee_id_value = getattr(current_user, 'employee_id', None)
                    if employee_id_value not in (None, '', 0, '0'):
                        employee_id_for_invoice = int(employee_id_value)
                except Exception:
                    employee_id_for_invoice = None

            if employee_display_name is None and employee_id_for_invoice:
                try:
                    emp = Employee.query.get(employee_id_for_invoice)
                    if emp and emp.name:
                        employee_display_name = emp.name
                except Exception:
                    employee_display_name = None

            # 3) Legacy User: map username -> AppUser (case-insensitive/trim) -> Employee
            if employee_id_for_invoice is None or employee_display_name is None:
                try:
                    from models import AppUser
                    from sqlalchemy import func

                    username_value = getattr(current_user, 'username', None)
                    username_key = (str(username_value).strip().lower() if username_value else '')
                    if username_key:
                        app_user = AppUser.query.filter(
                            func.lower(func.trim(AppUser.username)) == username_key
                        ).first()
                        if app_user:
                            if employee_id_for_invoice is None and getattr(app_user, 'employee_id', None):
                                try:
                                    employee_id_for_invoice = int(app_user.employee_id)
                                except Exception:
                                    employee_id_for_invoice = employee_id_for_invoice
                            if app_user.employee and app_user.employee.name:
                                employee_display_name = app_user.employee.name
                                employee_id_for_invoice = app_user.employee.id
                except Exception:
                    pass

            posted_by_username = (
                employee_display_name
                or getattr(current_user, 'full_name', None)
                or getattr(current_user, 'username', None)
            )

            # If this is an admin-triggered import, preserve historical employee attribution.
            if allow_employee_override and getattr(current_user, 'is_admin', False):
                try:
                    raw_emp_id = data.get('employee_id')
                    if raw_emp_id not in (None, '', 0, '0', False):
                        employee_id_for_invoice = int(raw_emp_id)
                except Exception:
                    pass

                try:
                    pb = data.get('posted_by')
                    if pb not in (None, '', False):
                        posted_by_username = str(pb).strip() or posted_by_username
                except Exception:
                    pass
        elif not auth_required:
            posted_by_username = (
                data.get('posted_by')
                or data.get('created_by')
                or data.get('username')
                or data.get('user')
            )

            # Optional: accept employee_id from the client when auth isn't required
            try:
                raw_emp_id = data.get('employee_id')
                employee_id_for_invoice = int(raw_emp_id) if raw_emp_id not in (None, '') else None
            except Exception:
                employee_id_for_invoice = None

        # Always store a stable posted_by label for audit + leaderboard fallback.
        try:
            posted_by_username = str(posted_by_username).strip() if posted_by_username else ''
        except Exception:
            posted_by_username = ''
        if not posted_by_username:
            try:
                posted_by_username = str(
                    getattr(current_user, 'full_name', None)
                    or getattr(current_user, 'name', None)
                    or getattr(current_user, 'username', None)
                    or 'system'
                ).strip()
            except Exception:
                posted_by_username = 'system'

        new_invoice = Invoice(
            invoice_type_id=next_invoice_type_id,
            customer_id=data.get('customer_id'),
            supplier_id=data.get('supplier_id'),
            employee_id=employee_id_for_invoice,
            branch_id=branch_id,
            office_id=office_id,
            date=datetime.fromisoformat(data['date']) if data.get('date') else datetime.now(),
            total=_extract_float('total', 0.0),
            invoice_type=invoice_type,
            total_weight=_extract_float('total_weight', 0.0),
            total_tax=_extract_float('total_tax'),
            total_cost=_extract_float('total_cost'),
            gold_subtotal=_extract_float('gold_subtotal'),
            wage_subtotal=_extract_float('wage_subtotal'),
            gold_tax_total=_extract_float('gold_tax_total'),
            wage_tax_total=_extract_float('wage_tax_total'),
            apply_gold_tax=bool(data.get('apply_gold_tax', False)),
            settlement_method=data.get('settlement_method'),
            payment_method=data.get('payment_method'),  # للتوافق مع الفواتير القديمة
            payment_method_id=payment_method_id,  # 🆕 Foreign key
            commission_amount=commission_amount,  # 🆕 العمولة المحسوبة
            net_amount=net_amount,  # 🆕 المبلغ الصافي
            amount_paid=_extract_float('amount_paid', 0.0),
            barter_total=barter_total,  # 🆕 قيمة المقايضة (تسوية غير نقدية)
            scrap_holder_employee_id=scrap_holder_employee_id,
            safe_box_id=data.get('safe_box_id'),  # 🆕 الخزينة المستخدمة
            barter_sale_invoice_id=data.get('barter_sale_invoice_id'),  # 🆕 ربط فاتورة شراء الكسر بفاتورة البيع (المقايضة)
            posted_by=posted_by_username,  # 🆕 تعيين المستخدم الذي أنشأ الفاتورة
            # 🆕 الحقول الجديدة
            original_invoice_id=data.get('original_invoice_id'),
            return_reason=data.get('return_reason'),
            gold_type=gold_type
        )
        db.session.add(new_invoice)
        db.session.flush()

        computed_total_weight = 0.0

        is_customer_scrap_purchase = (
            str(invoice_type).strip() in ('شراء من عميل', 'مرتجع شراء')
            and str(gold_type).strip().lower() == 'scrap'
        )

        # 🧮 Profit for customer scrap purchase (used by rewards)
        # الربح = (الوزن القائم - وزن الأحجار - الوزن) * سعر الشراء المباشر للعيار
        purchase_profit_cash = 0.0
        gold_price_data = None
        price_per_gram_24k = 0.0
        if invoice_type == 'شراء من عميل':
            gold_price_data = get_current_gold_price()
            price_per_gram_24k = _to_float(gold_price_data.get('price_per_gram_24k') if gold_price_data else 0.0, 0.0)
            if price_per_gram_24k <= 0:
                price_per_gram_24k = 400.0

        # --- Discount intelligence (audit) ---
        total_discount_cash = 0.0
        total_gross_cash = 0.0
        large_discount_pct_threshold = 10.0
        try:
            raw_thr = str(os.getenv('LARGE_DISCOUNT_PCT', '')).strip()
            if raw_thr:
                large_discount_pct_threshold = float(raw_thr)
        except Exception:
            large_discount_pct_threshold = 10.0

        for item_data in data.get('items', []):
            item_id = item_data.get('item_id')
            item = Item.query.get(item_id) if item_id else None

            if (item_data.get('create_inline') or False) and not item:
                try:
                    item = create_item_from_invoice_payload(item_data)
                    item_id = item.id
                except InlineItemCreationError as exc:
                    db.session.rollback()
                    return jsonify({'error': str(exc)}), 400

            if item_id and not item:
                return jsonify({'error': f"Item {item_id} not found"}), 404

            # Extract base attributes (prefer request values when provided)
            item_name = (item.name if item else item_data.get('name')) or 'صنف بدون اسم'
            item_karat = (
                item_data.get('karat')
                if item_data.get('karat') not in (None, '')
                else (item.karat if item else None)
            )
            item_weight = item_data.get('weight') if item_data.get('weight') is not None else (item.weight if item else None)
            # Prefer the wage sent from the frontend (per-invoice override),
            # then fall back to manufacturing_wage_per_gram, then catalog item wage.
            item_wage = (
                item_data.get('wage')
                or item_data.get('manufacturing_wage_per_gram')
                or (item.wage if item else None)
                or 0
            )

            if item_weight is None:
                item_weight = item_data.get('total_weight', 0)

            # 💵 Get values from request
            selling_price_raw = (
                item_data.get('selling_price')
                or item_data.get('price')
                or item_data.get('subtotal')
                or 0
            )
            tax_amount_raw = item_data.get('tax_amount', item_data.get('tax', 0)) or 0
            discount_amount_raw = item_data.get('discount_amount', 0)
            quantity_raw = item_data.get('quantity', 1)

            quantity_value = _to_float(quantity_raw, 1.0) or 1.0
            quantity_int = int(round(quantity_value)) if quantity_value > 0 else 1

            selling_price_val = _to_float(selling_price_raw, 0.0)
            tax_amount_val = _to_float(tax_amount_raw, 0.0)
            discount_amount_val = _to_float(discount_amount_raw, 0.0)

            # Track discount for audit purposes (sales only)
            try:
                if str(invoice_type).strip() == 'بيع':
                    total_discount_cash += max(0.0, float(discount_amount_val))
                    total_gross_cash += max(0.0, float(selling_price_val))
            except Exception:
                pass

            print(f"   💵 selling_price={selling_price_val}, tax_amount={tax_amount_val}, discount={discount_amount_val}")

            if tax_amount_val < 0:
                    tax_amount_val = abs(tax_amount_val)

            net_price = selling_price_val - tax_amount_val - discount_amount_val
            total_price = selling_price_val

            weight_per_item = _to_float(item_weight, 0.0)
            if weight_per_item <= 0:
                weight_per_item = _to_float(item_data.get('total_weight'), 0.0)

            standing_weight_val = _to_float(item_data.get('standing_weight'), 0.0)
            stones_weight_val = _to_float(item_data.get('stones_weight'), 0.0)
            direct_purchase_price_per_gram_val = _to_float(item_data.get('direct_purchase_price_per_gram'), 0.0)

            if invoice_type == 'شراء من عميل' and weight_per_item > 0 and standing_weight_val > 0:
                # Prefer purchase direct price from client (lower than market). Fallback to market-derived if missing.
                direct_price_per_gram = direct_purchase_price_per_gram_val
                if direct_price_per_gram <= 0:
                    karat_float = _to_float(item_karat, get_main_karat())
                    if karat_float <= 0:
                        karat_float = get_main_karat()
                    direct_price_per_gram = (price_per_gram_24k * karat_float) / 24.0
                diff_weight = standing_weight_val - stones_weight_val - weight_per_item
                purchase_profit_cash += diff_weight * direct_price_per_gram

            qty_multiplier = 1.0 if is_customer_scrap_purchase else quantity_value
            # weight_per_item = الوزن الكلي للسطر — لا نضربه في الكمية
            # (الكمية للعرض التوضيحي فقط)
            item_total_weight = weight_per_item
            if item_total_weight > 0:
                if not has_valid_karat_lines:
                    computed_total_weight += item_total_weight

            item_wage_val = _to_float(item_wage, 0.0)

            db.session.add(InvoiceItem(
                invoice_id=new_invoice.id,
                item_id=item.id if item else None,
                category_id=item_data.get('category_id'),
                name=item_name,
                karat=item_karat,
                weight=weight_per_item,
                standing_weight=standing_weight_val,
                stones_weight=stones_weight_val,
                direct_purchase_price_per_gram=direct_purchase_price_per_gram_val,
                wage=item_wage_val,
                net=net_price,
                tax=tax_amount_val,
                price=total_price,
                quantity=quantity_int
            ))

        if invoice_type == 'شراء من عميل':
            new_invoice.profit_cash = round(_to_float(purchase_profit_cash, 0.0), 2)

        processed_karat_lines = 0
        # Server-side enforced tax totals for karat_lines payloads
        enforced_gold_tax_total = 0.0
        enforced_wage_tax_total = 0.0
        if karat_lines_data and isinstance(karat_lines_data, list):
            for idx, line_data in enumerate(karat_lines_data, start=1):
                karat_value = _to_float(line_data.get('karat'))
                weight_value = _to_float(
                    line_data.get('weight_grams',
                                   line_data.get('weight',
                                                 line_data.get('total_weight')))
                )

                if karat_value <= 0 or weight_value <= 0:
                    continue

                gold_value_cash = _to_float(line_data.get('gold_value_cash', line_data.get('gold_value')))
                wage_cash = _to_float(line_data.get('manufacturing_wage_cash', line_data.get('wage_cash')))

                # Enforce VAT policy on server for karat lines.
                karat_int = int(round(_to_float(karat_value, 0.0))) if karat_value else 0
                is_exempt = karat_int in vat_exempt_karats
                apply_gold_tax_flag = bool(data.get('apply_gold_tax', False))

                # If client provided tax fields, validate them strictly.
                def _extract_optional_float(obj, key):
                    if not isinstance(obj, dict):
                        return None
                    if key not in obj:
                        return None
                    raw = obj.get(key)
                    if raw in (None, '', False):
                        return None
                    return _to_float(raw, 0.0)

                received_gold_tax = _extract_optional_float(line_data, 'gold_tax')
                received_wage_tax = _extract_optional_float(line_data, 'wage_tax')

                if not vat_enabled:
                    gold_tax_val = 0.0
                    wage_tax_val = 0.0
                else:
                    expected_wage_tax = wage_cash * vat_rate if wage_cash > 0 else 0.0
                    expected_gold_tax = 0.0
                    if apply_gold_tax_flag and not is_exempt and gold_value_cash > 0:
                        expected_gold_tax = gold_value_cash * vat_rate

                    # Strict validation (when provided): reject mismatches.
                    tol = 0.01
                    if received_gold_tax is not None and abs(received_gold_tax - expected_gold_tax) > tol:
                        db.session.rollback()
                        return jsonify({
                            'error': 'tax_policy_mismatch',
                            'message': 'Gold VAT does not match current VAT policy',
                            'line_index': idx,
                            'karat': karat_int,
                            'expected_gold_tax': round(expected_gold_tax, 2),
                            'received_gold_tax': round(received_gold_tax, 2),
                            'vat_rate': vat_rate,
                            'gold_vat_exempt': bool(is_exempt),
                        }), 400

                    if received_wage_tax is not None and abs(received_wage_tax - expected_wage_tax) > tol:
                        db.session.rollback()
                        return jsonify({
                            'error': 'tax_policy_mismatch',
                            'message': 'Wage VAT does not match current VAT policy',
                            'line_index': idx,
                            'karat': karat_int,
                            'expected_wage_tax': round(expected_wage_tax, 2),
                            'received_wage_tax': round(received_wage_tax, 2),
                            'vat_rate': vat_rate,
                            'gold_vat_exempt': bool(is_exempt),
                        }), 400

                    # Store expected values (always enforce exemption).
                    gold_tax_val = expected_gold_tax
                    wage_tax_val = expected_wage_tax

                enforced_gold_tax_total += _to_float(gold_tax_val, 0.0)
                enforced_wage_tax_total += _to_float(wage_tax_val, 0.0)
                description = line_data.get('description') or line_data.get('notes')

                db.session.add(InvoiceKaratLine(
                    invoice_id=new_invoice.id,
                    karat=karat_value,
                    weight_grams=weight_value,
                    gold_value_cash=gold_value_cash,
                    manufacturing_wage_cash=wage_cash,
                    gold_tax=gold_tax_val,
                    wage_tax=wage_tax_val,
                    description=description
                ))

                computed_total_weight += weight_value
                processed_karat_lines += 1

            # Override invoice tax totals from enforced karat-line calculation.
            try:
                new_invoice.gold_tax_total = round(enforced_gold_tax_total, 2)
                new_invoice.wage_tax_total = round(enforced_wage_tax_total, 2)
                new_invoice.total_tax = round(enforced_gold_tax_total + enforced_wage_tax_total, 2)
            except Exception:
                pass
        else:
            pass  # no karat lines

        if computed_total_weight > 0:
            new_invoice.total_weight = round(computed_total_weight, 4)
        elif data.get('items'):
            fallback_weight = sum(
                _to_float(item.get('weight'))
                or _to_float(item.get('total_weight'))
                or 0.0 for item in data.get('items', [])
            )
            fallback_weight = fallback_weight if fallback_weight > 0 else len(data.get('items', [])) * 0.001
            new_invoice.total_weight = round(max(fallback_weight, 0.001), 4)

        new_invoice.manufacturing_wage_mode_snapshot = wage_mode_snapshot
        db.session.add(new_invoice)
        db.session.flush()

        # --- Approval gates (sales) ---
        # Allow saving but prevent posting/safebox effects until approved.
        # Reasons:
        # - large_discount
        # - below_cost
        approval_required = False
        approval_reason = None
        approval_reasons = []
        discount_pct = None

        below_cost_details = {
            'enabled': False,
            'cost_basis': None,
            'cost_cash': 0.0,
            'effective_sale_cash_ex_vat': 0.0,
            'weight_main_grams_estimate': 0.0,
            'avg_total_cost_per_gram': 0.0,
            'profit_cash_estimate': 0.0,
        }
        purchase_above_live_price_details = {
            'enabled': False,
            'price_24k': 0.0,
            'tolerance_pct': 0.5,
            'items': [],
        }

        def _safe_float(v, default=0.0):
            try:
                if v in (None, '', False):
                    return float(default)
                return float(v)
            except Exception:
                return float(default)

        def _estimate_weight_main_from_payload() -> float:
            """Estimate total weight in MAIN_KARAT from payload (items or karat_lines)."""
            try:
                main_k = _safe_float(get_main_karat(), 21.0)
                if main_k <= 0:
                    main_k = 21.0
            except Exception:
                main_k = 21.0

            total_main = 0.0
            try:
                if has_valid_karat_lines and karat_lines_data and isinstance(karat_lines_data, list):
                    for ln in karat_lines_data:
                        if not isinstance(ln, dict):
                            continue
                        w = _to_float_request(
                            ln.get('weight_grams', ln.get('weight', ln.get('total_weight'))),
                            0.0,
                        )
                        k = _to_float_request(ln.get('karat'), 0.0)
                        if w <= 0 or k <= 0:
                            continue
                        total_main += float(w) * (float(k) / float(main_k))
                else:
                    for it in (data.get('items') or []):
                        if not isinstance(it, dict):
                            continue
                        w = _to_float_request(it.get('weight', it.get('total_weight')), 0.0)
                        q = _to_float_request(it.get('quantity', 1), 1.0) or 1.0
                        k = _to_float_request(it.get('karat'), 0.0)
                        if w <= 0:
                            continue
                        if k <= 0:
                            # If karat unknown, assume already in main karat.
                            total_main += float(w) * float(q)
                        else:
                            total_main += float(w) * float(q) * (float(k) / float(main_k))
            except Exception:
                return 0.0

            return round(max(total_main, 0.0), 4)

        # Gate 0: Customer scrap purchase above live direct gold price.
        try:
            if str(invoice_type).strip() == 'شراء من عميل':
                purchase_above_live_price_details['enabled'] = True
                purchase_above_live_price_details['price_24k'] = float(price_per_gram_24k or 0.0)

                tolerance_pct = 0.5
                tol_factor = 1.0 + (tolerance_pct / 100.0)

                for item in (data.get('items') or []):
                    if not isinstance(item, dict):
                        continue

                    weight_val = _to_float_request(item.get('weight', item.get('total_weight')), 0.0)
                    standing_weight_val = _to_float_request(item.get('standing_weight'), 0.0)
                    stones_weight_val = _to_float_request(item.get('stones_weight'), 0.0)
                    karat_val = _to_float_request(item.get('karat'), 0.0)
                    if karat_val <= 0:
                        karat_val = float(get_main_karat() or 21.0)

                    effective_weight = weight_val
                    if standing_weight_val > 0:
                        effective_weight = max(0.0, standing_weight_val - stones_weight_val)
                    if effective_weight <= 0:
                        continue

                    total_paid = _to_float_request(
                        item.get('net', item.get('price', item.get('subtotal'))),
                        0.0,
                    )
                    paid_per_gram = (total_paid / effective_weight) if effective_weight > 0 else 0.0
                    live_per_gram = (float(price_per_gram_24k or 0.0) * float(karat_val) / 24.0) if price_per_gram_24k > 0 else 0.0
                    max_allowed = live_per_gram * tol_factor

                    if live_per_gram > 0 and paid_per_gram > max_allowed:
                        purchase_above_live_price_details['items'].append({
                            'name': str(item.get('name') or 'صنف'),
                            'karat': float(karat_val),
                            'weight': round(float(effective_weight), 4),
                            'paid_per_gram': round(float(paid_per_gram), 2),
                            'live_per_gram': round(float(live_per_gram), 2),
                            'max_allowed_per_gram': round(float(max_allowed), 2),
                            'difference_per_gram': round(float(paid_per_gram - live_per_gram), 2),
                        })

                if purchase_above_live_price_details['items']:
                    approval_reasons.append('above_live_price')
        except Exception:
            pass

        # Gate 1: Large discount
        try:
            if str(invoice_type).strip() == 'بيع' and total_gross_cash > 0 and total_discount_cash > 0:
                discount_pct = (float(total_discount_cash) / float(total_gross_cash)) * 100.0
                if discount_pct >= float(large_discount_pct_threshold or 0.0):
                    approval_reasons.append('large_discount')
        except Exception:
            discount_pct = None

        # Gate 2: Sale below cost
        try:
            if str(invoice_type).strip() == 'بيع':
                # Effective sale value excluding VAT (and excluding commissions already in net_amount).
                effective_sale_cash_ex_vat = _safe_float(new_invoice.net_amount, 0.0) - _safe_float(new_invoice.total_tax, 0.0)
                effective_sale_cash_ex_vat = round(max(effective_sale_cash_ex_vat, 0.0), 2)

                # Prefer client-provided total_cost if available (it is later used for profit computation as well).
                provided_cost = _safe_float(getattr(new_invoice, 'total_cost', 0.0), 0.0)
                cost_cash = 0.0
                cost_basis = None

                if provided_cost and provided_cost > 0:
                    cost_cash = round(provided_cost, 2)
                    cost_basis = 'provided_total_cost'
                else:
                    # Fallback to moving-average estimate when cost is not provided.
                    weight_main = _estimate_weight_main_from_payload()
                    snapshot = None
                    # Isolate snapshot failures so they do not poison the outer transaction.
                    if weight_main > 0:
                        try:
                            with db.session.no_autoflush:
                                with db.session.begin_nested():
                                    snapshot = GoldCostingService.snapshot()
                        except Exception:
                            snapshot = None
                    avg_total = _safe_float(getattr(snapshot, 'avg_total', 0.0), 0.0)
                    if weight_main > 0 and avg_total > 0:
                        cost_cash = round(avg_total * weight_main, 2)
                        cost_basis = 'moving_average'
                    else:
                        cost_cash = 0.0
                        cost_basis = None

                    below_cost_details['weight_main_grams_estimate'] = float(weight_main)
                    below_cost_details['avg_total_cost_per_gram'] = float(avg_total)

                below_cost_details['enabled'] = True
                below_cost_details['cost_basis'] = cost_basis
                below_cost_details['cost_cash'] = float(cost_cash)
                below_cost_details['effective_sale_cash_ex_vat'] = float(effective_sale_cash_ex_vat)
                below_cost_details['profit_cash_estimate'] = float(round(effective_sale_cash_ex_vat - cost_cash, 2))

                # Trigger approval if we have a non-zero cost basis and sale is below it.
                tol = 0.01
                if cost_cash > 0 and (effective_sale_cash_ex_vat + tol) < cost_cash:
                    approval_reasons.append('below_cost')
        except Exception:
            # Do not break invoice creation if costing gate fails unexpectedly.
            pass

        # 🆕 force_post: allow admin/import callers to bypass approval gates
        # (historical imports should not be held for approval).
        force_post = False
        try:
            force_post = bool(data.get('force_post', False))
        except Exception:
            force_post = False

        if force_post:
            # Only honour force_post when auth is not required (CLI import)
            # or when the caller is an admin.
            is_admin_caller = bool(getattr(current_user, 'is_admin', False)) if current_user else False
            if not (is_admin_caller or not auth_required):
                force_post = False

        if force_post:
            approval_reasons = []

        approval_required = bool(approval_reasons)
        approval_reason = approval_reasons[0] if approval_reasons else None

        # 🆕 Check server-side posting settings
        # If auto_post_invoices is False, ALL invoices go to unposted mode
        # (unless force_post is set by admin/import).
        _posting_auto_post = True
        try:
            _posting_settings = Settings.query.first()
            if _posting_settings:
                _posting_auto_post = bool(getattr(_posting_settings, 'auto_post_invoices', True))
        except Exception:
            _posting_auto_post = True

        unposted_mode = bool(approval_required) or (not _posting_auto_post and not force_post)

        if unposted_mode:
            # Keep the invoice unposted, but do not overwrite payment status/amounts.
            # Payment rows are persisted and can be reflected in the invoice list,
            # while ledger/safebox effects remain gated by posting/approval.
            try:
                new_invoice.is_posted = False
                _reason_labels_ar = {
                    'above_live_price': 'السعر المدفوع يتجاوز السعر الحي للذهب',
                    'large_discount':   'خصم كبير يتجاوز الحد المسموح',
                    'below_cost':       'سعر البيع أقل من التكلفة',
                }
                if approval_reasons:
                    _parts = []
                    if 'above_live_price' in approval_reasons:
                        _items = (purchase_above_live_price_details or {}).get('items') or []
                        if _items:
                            _f = _items[0]
                            _parts.append(
                                f"⚠️ شراء أعلى من السعر المباشر: {_f.get('name','صنف')} "
                                f"بسعر/جرام {_f.get('paid_per_gram',0):.2f} "
                                f"مقابل مباشر {_f.get('live_per_gram',0):.2f}"
                            )
                        else:
                            _parts.append('⚠️ شراء أعلى من السعر المباشر')
                    if 'below_cost' in approval_reasons:
                        _bc = below_cost_details or {}
                        _sale = _bc.get('effective_sale_cash_ex_vat', 0) or 0
                        _cost = _bc.get('cost_cash', 0) or 0
                        _diff = _bc.get('profit_cash_estimate', 0) or 0
                        _parts.append(
                            f"⚠️ بيع تحت التكلفة: صافي {float(_sale):.2f} "
                            f"مقابل تكلفة {float(_cost):.2f} "
                            f"(فرق {float(_diff):.2f})"
                        )
                    if 'large_discount' in approval_reasons:
                        _dp = discount_pct or 0
                        _th = large_discount_pct_threshold or 0
                        _parts.append(
                            f"⚠️ خصم كبير: {float(_dp):.2f}% "
                            f"(الحد {float(_th):.2f}%)"
                        )
                    new_invoice.pending_approval_reason = (
                        '\n'.join(_parts) if _parts
                        else ' | '.join(approval_reasons)
                    )
                db.session.add(new_invoice)
                db.session.flush()
            except Exception:
                pass

        # 🆕 --- 1.5. Create Invoice Payments (وسائل دفع متعددة) ---

        def _is_cash_payment_method(pm) -> bool:
            """Best-effort check whether a PaymentMethod represents cash."""
            if pm is None:
                return False
            try:
                pt = str(getattr(pm, 'payment_type', '') or '').strip().lower()
                name = str(getattr(pm, 'name', '') or '').strip()
                if pt in {'cash'}:
                    return True
                return 'نقد' in name
            except Exception:
                return False

        def _is_receivable_payment_method(pm) -> bool:
            """Receivable means on-account; no safe box movement should be created."""
            if pm is None:
                return False
            try:
                pt = str(getattr(pm, 'payment_type', '') or '').strip().lower()
                if pt in {'receivable', 'credit', 'on_account', 'ar'}:
                    return True
                name = str(getattr(pm, 'name', '') or '').strip()
                # Common Arabic labels for on-account/receivables.
                if 'آجل' in name or 'اجل' in name:
                    return True
                return False
            except Exception:
                return False

        def _fallback_cash_safe_box_id() -> int | None:
            """Fallback cash SafeBox when none is supplied/configured.

            Precedence:
            - If Settings.employee_cash_safes_enabled and invoice employee has cash_safe_box_id -> use it
            - Else Settings.main_cash_safe_box_id
            - Else default cash safe
            """
            try:
                settings_row = Settings.query.first()
            except Exception:
                settings_row = None

            if bool(getattr(settings_row, 'employee_cash_safes_enabled', False)):
                try:
                    emp = getattr(new_invoice, 'employee', None)
                    if not emp and getattr(new_invoice, 'employee_id', None):
                        emp = Employee.query.get(int(new_invoice.employee_id))
                    emp_cash = getattr(emp, 'cash_safe_box_id', None) if emp else None
                    if emp_cash not in (None, '', 0, '0', False):
                        return int(emp_cash)
                except Exception:
                    pass

            try:
                main_cash = getattr(settings_row, 'main_cash_safe_box_id', None) if settings_row else None
                if main_cash not in (None, '', 0, '0', False):
                    return int(main_cash)
            except Exception:
                pass

            try:
                sb = SafeBox.get_default_by_type('cash')
                if sb and sb.id:
                    return int(sb.id)
            except Exception:
                pass

            # If no explicit default is configured, but there is exactly one active
            # cash safe box, use it as a conservative fallback.
            try:
                safes = SafeBox.query.filter_by(safe_type='cash', is_active=True).all()
                if isinstance(safes, list) and len(safes) == 1 and getattr(safes[0], 'id', None):
                    return int(safes[0].id)
            except Exception:
                pass

            # If multiple active cash safes exist (common in production), pick a stable
            # fallback instead of failing: prefer default then lowest id.
            try:
                sb = (
                    SafeBox.query.filter_by(safe_type='cash', is_active=True)
                    .order_by(SafeBox.is_default.desc(), SafeBox.id.asc())
                    .first()
                )
                if sb and getattr(sb, 'id', None):
                    return int(sb.id)
            except Exception:
                pass

            # Last resort: if there is exactly one active safe box in the system,
            # use it rather than failing (helps when safe_type is misconfigured).
            try:
                safes = SafeBox.query.filter_by(is_active=True).all()
                if isinstance(safes, list) and len(safes) == 1 and getattr(safes[0], 'id', None):
                    return int(safes[0].id)
            except Exception:
                pass

            # If multiple active safes exist, pick a stable fallback rather than failing.
            try:
                sb = SafeBox.query.filter_by(is_active=True).order_by(SafeBox.id.asc()).first()
                if sb and getattr(sb, 'id', None):
                    return int(sb.id)
            except Exception:
                pass

            return None

        def _fallback_non_cash_safe_box_id(pm) -> int | None:
            """Fallback SafeBox for non-cash payment methods when none is supplied/configured.

            Precedence:
            - If auto settlement enabled: default clearing safe, then default bank safe
            - Otherwise: default bank safe, then default clearing safe
            """
            try:
                if pm is None:
                    return None
                auto_settle = bool(getattr(pm, 'auto_settlement_enabled', False))
            except Exception:
                auto_settle = False

            def pick_default(t: str) -> int | None:
                try:
                    sb = SafeBox.get_default_by_type(t)
                    if sb and sb.id:
                        return int(sb.id)
                except Exception:
                    return None

                # If no explicit default is configured, but there is exactly one
                # active safe box of this type, use it as a conservative fallback.
                try:
                    safes = SafeBox.query.filter_by(safe_type=t, is_active=True).all()
                    if isinstance(safes, list) and len(safes) == 1 and getattr(safes[0], 'id', None):
                        return int(safes[0].id)
                except Exception:
                    return None

                # If multiple active safes exist and none is marked default, pick a
                # stable fallback to avoid blocking invoice creation.
                try:
                    sb = (
                        SafeBox.query.filter_by(safe_type=t, is_active=True)
                        .order_by(SafeBox.is_default.desc(), SafeBox.id.asc())
                        .first()
                    )
                    if sb and getattr(sb, 'id', None):
                        return int(sb.id)
                except Exception:
                    return None
                return None

            if auto_settle:
                return pick_default('clearing') or pick_default('bank')
            return pick_default('bank') or pick_default('clearing')

        def _coerce_cash_payment_safe_box_id(candidate_safe_box_id):
            """Ensure cash payments never route into a gold safe box.

            Scrap-sale invoices may carry a gold safe at invoice level for weight movement,
            but cash settlement vouchers must always hit a cash safe account.
            """
            if candidate_safe_box_id in (None, '', 0, '0', False):
                return candidate_safe_box_id

            try:
                sb = SafeBox.query.get(int(candidate_safe_box_id))
            except Exception:
                sb = None

            if sb and str(getattr(sb, 'safe_type', '') or '').strip().lower() == 'gold':
                try:
                    _settings_cash = Settings.query.first()
                except Exception:
                    _settings_cash = None

                main_cash = getattr(_settings_cash, 'main_cash_safe_box_id', None) if _settings_cash else None
                if main_cash not in (None, '', 0, '0', False):
                    try:
                        return int(main_cash)
                    except Exception:
                        pass

                fallback_cash = _fallback_cash_safe_box_id()
                if fallback_cash not in (None, '', 0, '0', False):
                    try:
                        return int(fallback_cash)
                    except Exception:
                        return fallback_cash

            return candidate_safe_box_id

        if payments_data and isinstance(payments_data, list) and len(payments_data) > 0:
            # إنشاء سجل لكل وسيلة دفع
            for payment in payments_data:
                pm_id = payment.get('payment_method_id')
                pm_amount = _to_float(payment.get('amount', 0.0))
                # Ignore zero/empty payment lines (common UI artifacts)
                if pm_amount <= 0.0001:
                    continue
                pm_obj = PaymentMethod.query.get(pm_id)

                is_receivable = _is_receivable_payment_method(pm_obj)

                # Resolve safe box per payment (single source of truth).
                # NOTE: Receivable methods are on-account and should not require a safe box.
                resolved_safe_box_id = None
                safe_box_obj = None
                if not is_receivable:
                    try:
                        raw_safe_box_id = payment.get('safe_box_id')
                        if raw_safe_box_id not in (None, '', False):
                            resolved_safe_box_id = int(raw_safe_box_id)
                    except Exception:
                        resolved_safe_box_id = None
                    # invoice.safe_box_id is the invoice-level default (usually cash).
                    # For non-cash payment methods (bank transfer, mada, etc.) we must NOT
                    # use the invoice safe_box because it would route to cash instead of bank.
                    if resolved_safe_box_id is None and _is_cash_payment_method(pm_obj):
                        resolved_safe_box_id = new_invoice.safe_box_id

                    if resolved_safe_box_id is None and _is_cash_payment_method(pm_obj):
                        try:
                            settings_row = Settings.query.first()
                        except Exception:
                            settings_row = None
                        if bool(getattr(settings_row, 'employee_cash_safes_enabled', False)):
                            resolved_safe_box_id = _fallback_cash_safe_box_id()

                    if resolved_safe_box_id is None and pm_obj is not None:
                        resolved_safe_box_id = getattr(pm_obj, 'default_safe_box_id', None)

                    if resolved_safe_box_id is None:
                        if _is_cash_payment_method(pm_obj):
                            resolved_safe_box_id = _fallback_cash_safe_box_id()
                        else:
                            resolved_safe_box_id = _fallback_non_cash_safe_box_id(pm_obj)

                    # Ultimate fallback: use cash safe as last resort for any payment
                    # method rather than failing the entire invoice.
                    if resolved_safe_box_id is None:
                        resolved_safe_box_id = _fallback_cash_safe_box_id()

                    # For cash methods, never allow routing to a gold safe.
                    if _is_cash_payment_method(pm_obj):
                        resolved_safe_box_id = _coerce_cash_payment_safe_box_id(resolved_safe_box_id)

                    # Enforce employee cash safe toggle: when disabled, do not route
                    # payments into the employee cash safe (fallback to main cash safe).
                    try:
                        emp_cash_safe_id = None
                        emp = getattr(new_invoice, 'employee', None)
                        if not emp and getattr(new_invoice, 'employee_id', None):
                            emp = Employee.query.get(int(new_invoice.employee_id))
                        raw_emp_cash = getattr(emp, 'cash_safe_box_id', None) if emp else None
                        if raw_emp_cash not in (None, '', 0, '0', False):
                            emp_cash_safe_id = int(raw_emp_cash)
                        if (
                            emp_cash_safe_id
                            and resolved_safe_box_id is not None
                            and int(resolved_safe_box_id) == int(emp_cash_safe_id)
                            and not bool(getattr(settings_row, 'employee_cash_safes_enabled', False))
                        ):
                            main_cash = getattr(settings_row, 'main_cash_safe_box_id', None) if settings_row else None
                            if main_cash not in (None, '', 0, '0', False):
                                resolved_safe_box_id = int(main_cash)
                            else:
                                try:
                                    sb = SafeBox.get_default_by_type('cash')
                                    if sb and sb.id:
                                        resolved_safe_box_id = int(sb.id)
                                except Exception:
                                    pass
                    except Exception:
                        pass

                    # شراء من عميل / مرتجع شراء: safe_box_id قد يكون خزينة ذهبية (لتتبع الوزن).
                    # سند الصرف النقدي يجب أن يستخدم دائماً خزينة نقدية.
                    if (
                        str(getattr(new_invoice, 'invoice_type', '') or '').strip() in ('شراء من عميل', 'مرتجع شراء')
                        and resolved_safe_box_id is not None
                    ):
                        try:
                            _pmt_sb = SafeBox.query.get(resolved_safe_box_id)
                            if _pmt_sb and ((_pmt_sb.safe_type or '').lower() == 'gold'):
                                _pmt_settings = Settings.query.first()
                                _pmt_main = getattr(_pmt_settings, 'main_cash_safe_box_id', None) if _pmt_settings else None
                                if _pmt_main not in (None, '', 0, '0', False):
                                    resolved_safe_box_id = int(_pmt_main)
                                else:
                                    _pmt_cs = SafeBox.get_default_by_type('cash')
                                    if _pmt_cs and _pmt_cs.id:
                                        resolved_safe_box_id = int(_pmt_cs.id)
                        except Exception:
                            pass

                    if resolved_safe_box_id is None:
                        db.session.rollback()
                        return jsonify({
                            'error': 'missing_safe_box_for_payment_method',
                            'message': 'يجب تحديد خزينة (SafeBox) لوسيلة الدفع أو ضبط خزينة افتراضية لها',
                            'payment_method_id': pm_id,
                            'payment_method_name': getattr(pm_obj, 'name', None) if pm_obj else None,
                        }), 400

                    # Persist resolved safe_box_id back into the payment dict so
                    # the journal-entry section (Phase 2) can read it later.
                    payment['safe_box_id'] = resolved_safe_box_id

                    # Validate safe box exists (avoid FK failures and keep atomicity explicit)
                    safe_box_obj = SafeBox.query.get(resolved_safe_box_id)
                    if not safe_box_obj:
                        db.session.rollback()
                        return jsonify({
                            'error': 'safe_box_not_found',
                            'message': 'الخزينة المحددة غير موجودة',
                            'safe_box_id': resolved_safe_box_id,
                            'payment_method_id': pm_id,
                            'payment_method_name': getattr(pm_obj, 'name', None) if pm_obj else None,
                        }), 400

                def _direction_for_invoice_type(t: str) -> str:
                    t = (t or '').strip()
                    if not t:
                        return 'in'
                    if t == 'بيع':
                        return 'in'
                    if t == 'مرتجع بيع':
                        return 'out'
                    if t in ('شراء من عميل', 'شراء') or (
                        'شراء' in t and 'مورد' in t and 'مرتجع' not in t
                    ):
                        return 'out'
                    if t in ('مرتجع شراء', 'مرتجع شراء (مورد)') or (
                        'مرتجع' in t and 'شراء' in t and 'مورد' in t
                    ):
                        return 'in'
                    return 'in'

                created_by_name = posted_by_username
                
                # حساب العمولة وضريبتها لهذه الدفعة
                pm_commission_rate = _to_float(payment.get('commission_rate', pm_obj.commission_rate if pm_obj else 0.0))

                if 'commission_amount' in payment:
                    pm_commission_amount = _to_float(payment.get('commission_amount', 0.0))
                else:
                    pm_commission_amount = pm_amount * (pm_commission_rate / 100) if pm_commission_rate > 0 else 0.0

                pm_commission_vat = _to_float(payment.get('commission_vat', pm_commission_amount * 0.15))  # 🆕 ضريبة 15%
                pm_net_amount = _to_float(payment.get('net_amount', pm_amount - pm_commission_amount - pm_commission_vat))

                payment_notes = payment.get('notes')
                if unposted_mode:
                    # Persist resolved safe_box_id for later approval/posting.
                    try:
                        payment_notes = json.dumps({
                            'user_notes': payment_notes,
                            **({'safe_box_id': resolved_safe_box_id} if (not is_receivable and resolved_safe_box_id is not None) else {}),
                        }, ensure_ascii=False)
                    except Exception:
                        payment_notes = payment.get('notes')
                
                payment_row = InvoicePayment(
                    invoice_id=new_invoice.id,
                    payment_method_id=pm_id,
                    amount=pm_amount,
                    commission_rate=pm_commission_rate,
                    commission_amount=pm_commission_amount,
                    commission_vat=pm_commission_vat,
                    net_amount=pm_net_amount,
                    notes=payment_notes
                )
                db.session.add(payment_row)
                db.session.flush()

                if (not unposted_mode) and (not is_receivable):
                    safe_account_id = getattr(safe_box_obj, 'account_id', None)
                    _warn_if_safe_account_mismatches_payment_method(pm_obj, safe_account_id, context='add_invoice:multi_payment')
                    if not safe_account_id:
                        db.session.rollback()
                        return jsonify({
                            'error': 'safe_box_missing_account_id',
                            'message': 'الخزينة المحددة لا تحتوي على حساب مرتبط (account_id)',
                            'safe_box_id': resolved_safe_box_id,
                        }), 400

                    direction = _direction_for_invoice_type(new_invoice.invoice_type)
                    voucher_type = 'receipt' if direction == 'in' else 'payment'

                    party_type = None
                    party_id = None
                    party_account_id = None
                    if getattr(new_invoice, 'supplier_id', None):
                        party_type = 'supplier'
                        party_id = int(new_invoice.supplier_id)
                        supplier = Supplier.query.get(party_id)
                        if not supplier:
                            db.session.rollback()
                            return jsonify({'error': 'supplier_not_found'}), 404
                        party_account_id = int(ensure_supplier_accounts(supplier).financial.id)
                    elif getattr(new_invoice, 'customer_id', None):
                        party_type = 'customer'
                        party_id = int(new_invoice.customer_id)
                        customer = Customer.query.get(party_id)
                        if not customer:
                            db.session.rollback()
                            return jsonify({'error': 'customer_not_found'}), 404
                        party_account_id = int(ensure_customer_accounts(customer).financial.id)
                    else:
                        db.session.rollback()
                        return jsonify({'error': 'missing_party_for_payment_voucher'}), 400

                    voucher_number = generate_voucher_number(voucher_type)
                    voucher_date = datetime.now()
                    try:
                        voucher_date = new_invoice.date or voucher_date
                    except Exception:
                        pass

                    voucher_notes = None
                    try:
                        voucher_notes = json.dumps({
                            'source': 'invoice_payment',
                            'invoice_id': int(new_invoice.id),
                            'invoice_payment_id': int(payment_row.id),
                            'payment_method_id': int(pm_id),
                        }, ensure_ascii=False)
                    except Exception:
                        voucher_notes = None

                    voucher = Voucher(
                        voucher_number=voucher_number,
                        voucher_type=voucher_type,
                        date=voucher_date,
                        party_type=party_type,
                        customer_id=party_id if party_type == 'customer' else None,
                        supplier_id=party_id if party_type == 'supplier' else None,
                        amount_cash=float(pm_amount),
                        amount_gold=0.0,
                        description=f"دفعة فاتورة {getattr(new_invoice, 'invoice_type_id', '')}".strip(),
                        reference_type='invoice',
                        reference_id=int(new_invoice.id),
                        reference_number=str(getattr(new_invoice, 'invoice_type_id', '') or '') or None,
                        notes=voucher_notes,
                        created_by=created_by_name or 'system',
                        status='pending',
                    )
                    db.session.add(voucher)
                    db.session.flush()

                    safe_line_type = 'debit' if direction == 'in' else 'credit'
                    party_line_type = 'credit' if direction == 'in' else 'debit'

                    db.session.add(VoucherAccountLine(
                        voucher_id=voucher.id,
                        account_id=int(safe_account_id),
                        line_type=safe_line_type,
                        amount_type='cash',
                        amount=float(pm_amount),
                        description=payment.get('notes'),
                    ))
                    db.session.add(VoucherAccountLine(
                        voucher_id=voucher.id,
                        account_id=int(party_account_id),
                        line_type=party_line_type,
                        amount_type='cash',
                        amount=float(pm_amount),
                        description=payment.get('notes'),
                    ))
                    db.session.flush()

                    consolidated_je = _add_payment_lines_to_consolidated_je(
                        invoice=new_invoice,
                        voucher=voucher,
                        voucher_number=voucher_number,
                        safe_account_id=int(safe_account_id),
                        party_account_id=int(party_account_id),
                        amount=float(pm_amount),
                        payment_id=payment_row.id,
                        direction=direction,
                        voucher_date=voucher_date,
                        created_by=created_by_name or 'system',
                    )

                    voucher.status = 'approved'
                    voucher.approved_at = datetime.now()
                    voucher.approved_by = created_by_name or 'system'
                    voucher.journal_entry_id = consolidated_je.id

                    _append_safe_transactions_for_voucher(voucher, created_by=voucher.approved_by)
        
        # وسيلة دفع واحدة (للتوافق مع الكود القديم)
        elif payment_method_id:
            pm_obj = PaymentMethod.query.get(payment_method_id)
            pm_commission_rate = pm_obj.commission_rate if pm_obj else 0.0

            is_receivable = _is_receivable_payment_method(pm_obj)

            resolved_safe_box_id = None
            safe_box_obj = None
            if not is_receivable:
                # invoice.safe_box_id is the cash-level default — only use it for cash PMs.
                # Non-cash PMs (bank transfer, mada, bnpl) must go to their own default_safe_box_id.
                if _is_cash_payment_method(pm_obj):
                    resolved_safe_box_id = new_invoice.safe_box_id
                if resolved_safe_box_id is None:
                    resolved_safe_box_id = getattr(pm_obj, 'default_safe_box_id', None) if pm_obj else None
                if resolved_safe_box_id is None:
                    if _is_cash_payment_method(pm_obj):
                        resolved_safe_box_id = _fallback_cash_safe_box_id()
                    else:
                        resolved_safe_box_id = _fallback_non_cash_safe_box_id(pm_obj)
                # Ultimate fallback: use cash safe as last resort.
                if resolved_safe_box_id is None:
                    resolved_safe_box_id = _fallback_cash_safe_box_id()

                # For cash methods, never allow routing to a gold safe.
                if _is_cash_payment_method(pm_obj):
                    resolved_safe_box_id = _coerce_cash_payment_safe_box_id(resolved_safe_box_id)

                # شراء من عميل / مرتجع شراء: safe_box_id قد يكون خزينة ذهبية (لتتبع الوزن).
                # سند الصرف النقدي يجب أن يستخدم دائماً خزينة نقدية.
                if (
                    str(getattr(new_invoice, 'invoice_type', '') or '').strip() in ('شراء من عميل', 'مرتجع شراء')
                    and resolved_safe_box_id is not None
                ):
                    try:
                        _pmt_sb = SafeBox.query.get(resolved_safe_box_id)
                        if _pmt_sb and ((_pmt_sb.safe_type or '').lower() == 'gold'):
                            _pmt_settings = Settings.query.first()
                            _pmt_main = getattr(_pmt_settings, 'main_cash_safe_box_id', None) if _pmt_settings else None
                            if _pmt_main not in (None, '', 0, '0', False):
                                resolved_safe_box_id = int(_pmt_main)
                            else:
                                _pmt_cs = SafeBox.get_default_by_type('cash')
                                if _pmt_cs and _pmt_cs.id:
                                    resolved_safe_box_id = int(_pmt_cs.id)
                    except Exception:
                        pass

                if resolved_safe_box_id is None:
                    db.session.rollback()
                    return jsonify({
                        'error': 'missing_safe_box_for_payment_method',
                        'message': 'يجب تحديد خزينة (SafeBox) لوسيلة الدفع أو ضبط خزينة افتراضية لها',
                        'payment_method_id': payment_method_id,
                        'payment_method_name': getattr(pm_obj, 'name', None) if pm_obj else None,
                    }), 400

                # Persist resolved safe_box_id so the journal-entry section can read it.
                safe_box_id = resolved_safe_box_id

                safe_box_obj = SafeBox.query.get(resolved_safe_box_id)
                if not safe_box_obj:
                    db.session.rollback()
                    return jsonify({
                        'error': 'safe_box_not_found',
                        'message': 'الخزينة المحددة غير موجودة',
                        'safe_box_id': resolved_safe_box_id,
                        'payment_method_id': payment_method_id,
                        'payment_method_name': getattr(pm_obj, 'name', None) if pm_obj else None,
                    }), 400

            def _direction_for_invoice_type(t: str) -> str:
                t = (t or '').strip()
                if not t:
                    return 'in'
                if t == 'بيع':
                    return 'in'
                if t == 'مرتجع بيع':
                    return 'out'
                if t in ('شراء من عميل', 'شراء') or (
                    'شراء' in t and 'مورد' in t and 'مرتجع' not in t
                ):
                    return 'out'
                if t in ('مرتجع شراء', 'مرتجع شراء (مورد)') or (
                    'مرتجع' in t and 'شراء' in t and 'مورد' in t
                ):
                    return 'in'
                return 'in'

            payment_notes = None
            if unposted_mode:
                try:
                    payment_notes = json.dumps({
                        'user_notes': None,
                        **({'safe_box_id': resolved_safe_box_id} if (not is_receivable and resolved_safe_box_id is not None) else {}),
                    }, ensure_ascii=False)
                except Exception:
                    payment_notes = None

            # Legacy single-payment path:
            # - If the client explicitly provided amount_paid, respect it.
            #   (Important for deferred/credit returns where amount_paid=0.)
            # - Otherwise, fall back to the invoice total (historical behavior).
            if 'amount_paid' in data and data.get('amount_paid') not in (None, '', False):
                single_payment_amount = _extract_float('amount_paid', 0.0)
            else:
                single_payment_amount = _extract_float('total', 0.0)

            # If explicit amount is zero, do not create an InvoicePayment row.
            if single_payment_amount <= 0.0001:
                single_payment_amount = 0.0
            
            payment_row = None
            if single_payment_amount > 0.0001:
                payment_row = InvoicePayment(
                    invoice_id=new_invoice.id,
                    payment_method_id=payment_method_id,
                    amount=single_payment_amount,
                    commission_rate=pm_commission_rate,
                    commission_amount=commission_amount,
                    net_amount=net_amount,
                    notes=payment_notes,
                )
                db.session.add(payment_row)
                db.session.flush()

            if payment_row is not None and (not unposted_mode) and (not is_receivable):
                safe_account_id = getattr(safe_box_obj, 'account_id', None)
                _warn_if_safe_account_mismatches_payment_method(pm_obj, safe_account_id, context='add_invoice:legacy_single_payment')
                if not safe_account_id:
                    db.session.rollback()
                    return jsonify({
                        'error': 'safe_box_missing_account_id',
                        'message': 'الخزينة المحددة لا تحتوي على حساب مرتبط (account_id)',
                        'safe_box_id': resolved_safe_box_id,
                    }), 400

                direction = _direction_for_invoice_type(new_invoice.invoice_type)
                voucher_type = 'receipt' if direction == 'in' else 'payment'
                amount_value = float(single_payment_amount or 0.0)

                party_type = None
                party_id = None
                party_account_id = None
                if getattr(new_invoice, 'supplier_id', None):
                    party_type = 'supplier'
                    party_id = int(new_invoice.supplier_id)
                    supplier = Supplier.query.get(party_id)
                    if not supplier:
                        db.session.rollback()
                        return jsonify({'error': 'supplier_not_found'}), 404
                    party_account_id = int(ensure_supplier_accounts(supplier).financial.id)
                elif getattr(new_invoice, 'customer_id', None):
                    party_type = 'customer'
                    party_id = int(new_invoice.customer_id)
                    customer = Customer.query.get(party_id)
                    if not customer:
                        db.session.rollback()
                        return jsonify({'error': 'customer_not_found'}), 404
                    party_account_id = int(ensure_customer_accounts(customer).financial.id)
                else:
                    db.session.rollback()
                    return jsonify({'error': 'missing_party_for_payment_voucher'}), 400

                voucher_number = generate_voucher_number(voucher_type)
                voucher_date = datetime.now()
                try:
                    voucher_date = new_invoice.date or voucher_date
                except Exception:
                    pass

                voucher_notes = None
                try:
                    voucher_notes = json.dumps({
                        'source': 'invoice_payment',
                        'invoice_id': int(new_invoice.id),
                        'invoice_payment_id': int(payment_row.id),
                        'payment_method_id': int(payment_method_id),
                    }, ensure_ascii=False)
                except Exception:
                    voucher_notes = None

                voucher = Voucher(
                    voucher_number=voucher_number,
                    voucher_type=voucher_type,
                    date=voucher_date,
                    party_type=party_type,
                    customer_id=party_id if party_type == 'customer' else None,
                    supplier_id=party_id if party_type == 'supplier' else None,
                    amount_cash=float(amount_value),
                    amount_gold=0.0,
                    description=f"دفعة فاتورة {getattr(new_invoice, 'invoice_type_id', '')}".strip(),
                    reference_type='invoice',
                    reference_id=int(new_invoice.id),
                    reference_number=str(getattr(new_invoice, 'invoice_type_id', '') or '') or None,
                    notes=voucher_notes,
                    created_by=posted_by_username or 'system',
                    status='pending',
                )
                db.session.add(voucher)
                db.session.flush()

                safe_line_type = 'debit' if direction == 'in' else 'credit'
                party_line_type = 'credit' if direction == 'in' else 'debit'

                db.session.add(VoucherAccountLine(
                    voucher_id=voucher.id,
                    account_id=int(safe_account_id),
                    line_type=safe_line_type,
                    amount_type='cash',
                    amount=float(amount_value),
                    description=None,
                ))
                db.session.add(VoucherAccountLine(
                    voucher_id=voucher.id,
                    account_id=int(party_account_id),
                    line_type=party_line_type,
                    amount_type='cash',
                    amount=float(amount_value),
                    description=None,
                ))
                db.session.flush()

                consolidated_je = _add_payment_lines_to_consolidated_je(
                    invoice=new_invoice,
                    voucher=voucher,
                    voucher_number=voucher_number,
                    safe_account_id=int(safe_account_id),
                    party_account_id=int(party_account_id),
                    amount=float(amount_value),
                    payment_id=payment_row.id,
                    direction=direction,
                    voucher_date=voucher_date,
                    created_by=posted_by_username or 'system',
                )

                voucher.status = 'approved'
                voucher.approved_at = datetime.now()
                voucher.approved_by = posted_by_username or 'system'
                voucher.journal_entry_id = consolidated_je.id
                _append_safe_transactions_for_voucher(voucher, created_by=voucher.approved_by)

        # --- Gold settlement (barter/partial) ---
        try:
            settlement_method = data.get('settlement_method') or data.get('settlement_mode')
            if settlement_method is not None:
                new_invoice.settlement_method = str(settlement_method).strip() or None
        except Exception:
            pass

        # --- عمولة السداد بذهب صافي (عيار 24) — قديم، يُبقى للتوافق ---
        try:
            if data.get('gold24k_settlement'):
                new_invoice.gold24k_settlement = True
                new_invoice.gold24k_weight = _to_float_request(data.get('gold24k_weight', 0.0))
                new_invoice.gold24k_commission_per_gram = _to_float_request(data.get('gold24k_commission_per_gram', 0.0))
                new_invoice.gold24k_commission_total = _to_float_request(data.get('gold24k_commission_total', 0.0))
        except Exception:
            pass

        # --- عمولة / رسوم فرق العيار — per-line aggregation ---
        try:
            earn = _to_float_request(data.get('karat_diff_earn_total', 0.0))
            pay = _to_float_request(data.get('karat_diff_pay_total', 0.0))
            if earn > 0:
                new_invoice.karat_diff_earn_total = earn
            if pay > 0:
                new_invoice.karat_diff_pay_total = pay
        except Exception:
            pass

        settled_gold_weight = _to_float_request(data.get('settled_gold_weight', 0.0))
        settled_gold_karat = _to_float_request(data.get('settled_gold_karat', 0.0))
        settled_gold_safe_box_id = data.get('settled_gold_safe_box_id')

        gold_settlements_data = data.get('gold_settlements')
        if gold_settlements_data in (None, '', False):
            # Backward-compat aliases (if any older client uses a different key)
            gold_settlements_data = data.get('settled_gold_lines')
        if not isinstance(gold_settlements_data, list):
            gold_settlements_data = []

        # We may need to reflect gold settlements in statements (built off JournalEntryLine).
        # We now prefer generating an approved Voucher (payment/receipt) for gold settlements
        # so the movement is visible both as a سند and as JournalEntryLine (via voucher posting).
        resolved_gold_settlements_lines = []  # [{safe_box_id, karat, weight}]
        gold_settlement_voucher_created = False

        # For unposted invoices (drafts/approval-required), do not allow gold settlement inputs (would require safe movements).
        if unposted_mode:
            if (
                settled_gold_weight > 0
                or settled_gold_karat > 0
                or settled_gold_safe_box_id not in (None, '', False)
                or (gold_settlements_data and len(gold_settlements_data) > 0)
            ):
                db.session.rollback()
                return jsonify({
                    'error': 'unposted_no_gold_settlement',
                    'message': 'لا يمكن حفظ سداد/مقايضة ذهب عندما تكون الفاتورة غير مرحلة (مسودة/بحاجة اعتماد). قم بإزالة بيانات سداد الذهب ثم احفظ، وبعد الإكمال/الترحيل قم بالترحيل.',
                    'approval_required': bool(approval_required),
                    'reason': approval_reason,
                }), 400

        # Multi-safe gold settlements.
        # If provided, this takes precedence over the legacy settled_gold_* fields.
        if gold_settlements_data and isinstance(gold_settlements_data, list):
            # Resolve/override rules depend on Settings and employee gold safe configuration.
            try:
                srow = Settings.query.first()
            except Exception:
                srow = None

            employee_gold_enabled = bool(getattr(srow, 'employee_gold_safes_enabled', False)) if srow else False
            emp_gold_safe_id = None
            try:
                emp = getattr(new_invoice, 'employee', None)
                if not emp and getattr(new_invoice, 'employee_id', None):
                    emp = Employee.query.get(int(new_invoice.employee_id))
                raw_emp_gold = getattr(emp, 'gold_safe_box_id', None) if emp else None
                if raw_emp_gold not in (None, '', 0, '0', False):
                    emp_gold_safe_id = int(raw_emp_gold)
            except Exception:
                emp_gold_safe_id = None

            def _resolve_gold_safe_id(requested_safe_id):
                """Resolve the target gold safe.

                Rules (compatible with legacy behavior):
                - If employee gold safes enabled and employee has a linked gold safe: enforce using it.
                  If client explicitly selected a different safe, fail.
                - Otherwise: use requested_safe_id, falling back to main scrap gold safe.
                - If employee gold safes disabled: prevent routing into employee gold safe (reroute to main scrap).
                """
                if employee_gold_enabled and emp_gold_safe_id:
                    if requested_safe_id not in (None, '', False):
                        try:
                            if int(requested_safe_id) != int(emp_gold_safe_id):
                                db.session.rollback()
                                return None, jsonify({
                                    'error': 'gold_settlement_forced_employee_safe',
                                    'message': 'تم تفعيل خزائن ذهب الموظفين، ويجب أن يتم سداد الذهب على خزينة ذهب الموظف فقط',
                                    'employee_gold_safe_box_id': emp_gold_safe_id,
                                    'requested_safe_box_id': requested_safe_id,
                                }), 400
                        except Exception:
                            db.session.rollback()
                            return None, jsonify({
                                'error': 'invalid_gold_safe_box_id',
                                'message': 'خزينة الذهب غير صحيحة',
                            }), 400
                    return int(emp_gold_safe_id), None

                # Not forced: honor request if present.
                resolved = requested_safe_id
                if resolved in (None, '', False):
                    resolved = getattr(srow, 'main_scrap_gold_safe_box_id', None) if srow else None

                try:
                    resolved = int(resolved)
                except Exception:
                    db.session.rollback()
                    return None, jsonify({
                        'error': 'missing_or_invalid_gold_safe_box',
                        'message': 'يجب تحديد خزينة ذهب صحيحة للمقايضة/السداد بالذهب',
                    }), 400

                # If employee gold safes are disabled, prevent routing into employee gold safe.
                try:
                    if (not employee_gold_enabled) and emp_gold_safe_id and int(resolved) == int(emp_gold_safe_id):
                        main_scrap = getattr(srow, 'main_scrap_gold_safe_box_id', None) if srow else None
                        if main_scrap not in (None, '', 0, '0', False):
                            resolved = int(main_scrap)
                except Exception:
                    pass

                return resolved, None

            total_main_equiv = 0.0
            created_any = False

            for idx, line in enumerate(gold_settlements_data):
                if not isinstance(line, dict):
                    continue

                line_weight = _to_float_request(line.get('weight', line.get('weight_grams', 0.0)), 0.0)
                line_karat = _to_float_request(line.get('karat', line.get('gold_karat', 0.0)), 0.0)
                if line_weight <= 0 or line_karat <= 0:
                    continue

                safe_id, err = _resolve_gold_safe_id(line.get('safe_box_id'))
                if err is not None:
                    return err
                if safe_id is None:
                    db.session.rollback()
                    return jsonify({
                        'error': 'missing_or_invalid_gold_safe_box',
                        'message': 'يجب تحديد خزينة ذهب صحيحة للمقايضة/السداد بالذهب',
                    }), 400

                gold_safe = SafeBox.query.get(safe_id)
                if not gold_safe:
                    db.session.rollback()
                    return jsonify({
                        'error': 'gold_safe_box_not_found',
                        'message': 'خزينة الذهب المحددة غير موجودة',
                        'safe_box_id': safe_id,
                    }), 400

                safe_type = getattr(gold_safe, 'safe_type', None) or getattr(gold_safe, 'safeType', None)
                if safe_type != 'gold':
                    db.session.rollback()
                    return jsonify({
                        'error': 'safe_box_not_gold',
                        'message': 'الخزينة المحددة ليست خزينة ذهب',
                        'safe_box_id': safe_id,
                        'safe_type': safe_type,
                    }), 400

                karat_int = int(round(float(line_karat)))

                # Validation: if the safe is fixed to a single karat, enforce it.
                try:
                    safe_karat = int(getattr(gold_safe, 'karat', None) or 0)
                except Exception:
                    safe_karat = 0
                if safe_karat in (18, 21, 22, 24) and karat_int != safe_karat:
                    db.session.rollback()
                    return jsonify({
                        'error': 'karat_mismatch_for_safe_box',
                        'message': f'الخزينة المحددة مخصصة لعيار {safe_karat} ولا تقبل عيار {karat_int}',
                        'safe_box_id': safe_id,
                        'allowed_karat': safe_karat,
                        'karat': karat_int,
                    }), 400

                weight_kwargs = {
                    'weight_18k': 0.0,
                    'weight_21k': 0.0,
                    'weight_22k': 0.0,
                    'weight_24k': 0.0,
                }
                if karat_int == 18:
                    weight_kwargs['weight_18k'] = float(line_weight)
                elif karat_int == 21:
                    weight_kwargs['weight_21k'] = float(line_weight)
                elif karat_int == 22:
                    weight_kwargs['weight_22k'] = float(line_weight)
                elif karat_int == 24:
                    weight_kwargs['weight_24k'] = float(line_weight)
                else:
                    db.session.rollback()
                    return jsonify({
                        'error': 'invalid_gold_karat',
                        'message': 'عيار الذهب غير مدعوم للمقايضة/السداد',
                        'karat': line_karat,
                    }), 400

                # Aggregate settled weight in MAIN karat on the invoice.
                try:
                    main_equiv = convert_to_main_karat(float(line_weight), karat_int)
                except Exception:
                    main_equiv = float(line_weight) * (karat_int / 21.0) if karat_int > 0 else 0.0
                total_main_equiv += float(main_equiv or 0.0)

                try:
                    resolved_gold_settlements_lines.append({
                        'safe_box_id': int(safe_id),
                        'karat': int(karat_int),
                        'weight': float(line_weight),
                    })
                except Exception:
                    pass

                created_any = True

            if created_any:
                new_invoice.settled_gold_weight = round(float(total_main_equiv), 3)

            # Avoid double-posting via legacy settled_gold_* fields.
            settled_gold_weight = 0.0
            settled_gold_karat = 0.0
            settled_gold_safe_box_id = None

        if settled_gold_weight > 0 and settled_gold_karat > 0:
            # Resolve the target gold safe — always use employee vault when available.
            try:
                srow = Settings.query.first()
            except Exception:
                srow = None

            emp_gold_safe_id = None
            try:
                emp = getattr(new_invoice, 'employee', None)
                if not emp and getattr(new_invoice, 'employee_id', None):
                    emp = Employee.query.get(int(new_invoice.employee_id))
                raw_emp_gold = getattr(emp, 'gold_safe_box_id', None) if emp else None
                if raw_emp_gold not in (None, '', 0, '0', False):
                    emp_gold_safe_id = int(raw_emp_gold)
            except Exception:
                emp_gold_safe_id = None

            # الأولوية: خزينة ذهب الموظف دائماً (بغض النظر عن employee_gold_safes_enabled)
            if emp_gold_safe_id:
                settled_gold_safe_box_id = emp_gold_safe_id
            elif settled_gold_safe_box_id in (None, '', False):
                settled_gold_safe_box_id = getattr(srow, 'main_scrap_gold_safe_box_id', None) if srow else None

            try:
                settled_gold_safe_box_id = int(settled_gold_safe_box_id)
            except Exception:
                db.session.rollback()
                return jsonify({
                    'error': 'missing_or_invalid_gold_safe_box',
                    'message': 'يجب تحديد خزينة ذهب صحيحة للمقايضة/السداد بالذهب',
                }), 400

            gold_safe = SafeBox.query.get(settled_gold_safe_box_id)
            if not gold_safe:
                db.session.rollback()
                return jsonify({
                    'error': 'gold_safe_box_not_found',
                    'message': 'خزينة الذهب المحددة غير موجودة',
                    'safe_box_id': settled_gold_safe_box_id,
                }), 400

            safe_type = getattr(gold_safe, 'safe_type', None) or getattr(gold_safe, 'safeType', None)
            if safe_type != 'gold':
                db.session.rollback()
                return jsonify({
                    'error': 'safe_box_not_gold',
                    'message': 'الخزينة المحددة ليست خزينة ذهب',
                    'safe_box_id': settled_gold_safe_box_id,
                    'safe_type': safe_type,
                }), 400

            karat_int = int(round(float(settled_gold_karat)))

            # Validation: if the safe is fixed to a single karat, enforce it.
            try:
                safe_karat = int(getattr(gold_safe, 'karat', None) or 0)
            except Exception:
                safe_karat = 0
            if safe_karat in (18, 21, 22, 24) and karat_int != safe_karat:
                db.session.rollback()
                return jsonify({
                    'error': 'karat_mismatch_for_safe_box',
                    'message': f'الخزينة المحددة مخصصة لعيار {safe_karat} ولا تقبل عيار {karat_int}',
                    'safe_box_id': settled_gold_safe_box_id,
                    'allowed_karat': safe_karat,
                    'karat': karat_int,
                }), 400

            weight_kwargs = {
                'weight_18k': 0.0,
                'weight_21k': 0.0,
                'weight_22k': 0.0,
                'weight_24k': 0.0,
            }
            if karat_int == 18:
                weight_kwargs['weight_18k'] = float(settled_gold_weight)
            elif karat_int == 21:
                weight_kwargs['weight_21k'] = float(settled_gold_weight)
            elif karat_int == 22:
                weight_kwargs['weight_22k'] = float(settled_gold_weight)
            elif karat_int == 24:
                weight_kwargs['weight_24k'] = float(settled_gold_weight)
            else:
                db.session.rollback()
                return jsonify({
                    'error': 'invalid_gold_karat',
                    'message': 'عيار الذهب غير مدعوم للمقايضة/السداد',
                    'karat': settled_gold_karat,
                }), 400

            # Store settled weight in MAIN karat on the invoice.
            try:
                main_equiv = convert_to_main_karat(float(settled_gold_weight), karat_int)
            except Exception:
                # Fallback: simple proportion to 21k if conversion helper unavailable
                main_equiv = float(settled_gold_weight) * (karat_int / 21.0) if karat_int > 0 else 0.0

            new_invoice.settled_gold_weight = round(float(main_equiv), 3)

            try:
                resolved_gold_settlements_lines.append({
                    'safe_box_id': int(settled_gold_safe_box_id),
                    'karat': int(karat_int),
                    'weight': float(settled_gold_weight),
                })
            except Exception:
                pass

            # SafeBoxTransaction is written via the auto-created settlement voucher.

        # If we have gold settlement lines, generate an approved voucher so the movement
        # exists as (سند صرف/قبض) + produces JournalEntryLines + SafeBoxTransaction.
        if (not unposted_mode) and resolved_gold_settlements_lines and (not gold_settlement_voucher_created):
            try:
                settlement_direction = _direction_for_invoice_type(new_invoice.invoice_type)
            except Exception:
                settlement_direction = 'out'

            voucher_type = 'receipt' if settlement_direction == 'in' else 'payment'

            party_type = None
            party_id = None
            party_account_id = None
            try:
                if getattr(new_invoice, 'supplier_id', None):
                    party_type = 'supplier'
                    party_id = int(new_invoice.supplier_id)
                    supplier = Supplier.query.get(party_id)
                    if not supplier:
                        db.session.rollback()
                        return jsonify({'error': 'supplier_not_found'}), 404
                    party_account_id = int(ensure_supplier_accounts(supplier).financial.id)
                elif getattr(new_invoice, 'customer_id', None):
                    party_type = 'customer'
                    party_id = int(new_invoice.customer_id)
                    customer = Customer.query.get(party_id)
                    if not customer:
                        db.session.rollback()
                        return jsonify({'error': 'customer_not_found'}), 404
                    party_account_id = int(ensure_customer_accounts(customer).financial.id)
                else:
                    db.session.rollback()
                    return jsonify({'error': 'missing_party_for_gold_settlement_voucher'}), 400
            except Exception:
                db.session.rollback()
                return jsonify({'error': 'missing_party_for_gold_settlement_voucher'}), 400

            # Aggregate by (safe_box_id, karat) for safe lines and by karat for party line(s)
            safe_agg = {}  # (safe_id, karat) -> weight
            party_agg = {}  # karat -> weight
            for row in resolved_gold_settlements_lines:
                try:
                    safe_id = int(row.get('safe_box_id'))
                    karat_int = int(row.get('karat'))
                    weight_val = float(row.get('weight') or 0.0)
                except Exception:
                    continue
                if safe_id <= 0 or weight_val <= 0 or karat_int not in (18, 21, 22, 24):
                    continue
                safe_agg[(safe_id, karat_int)] = float(safe_agg.get((safe_id, karat_int), 0.0) or 0.0) + float(weight_val)
                party_agg[karat_int] = float(party_agg.get(karat_int, 0.0) or 0.0) + float(weight_val)

            if safe_agg and party_agg:
                voucher_number = generate_voucher_number(voucher_type)
                voucher_date = datetime.now()
                try:
                    voucher_date = new_invoice.date or voucher_date
                except Exception:
                    pass

                voucher_notes = None
                try:
                    voucher_notes = json.dumps({
                        'source': 'invoice_gold_settlement',
                        'invoice_id': int(new_invoice.id),
                        'settlement_method': (getattr(new_invoice, 'settlement_method', None) or data.get('settlement_method') or data.get('settlement_mode')),
                        'gold_settlements': resolved_gold_settlements_lines,
                    }, ensure_ascii=False)
                except Exception:
                    voucher_notes = None

                total_gold_weight = round(sum(float(v or 0.0) for v in party_agg.values()), 3)

                voucher = Voucher(
                    voucher_number=voucher_number,
                    voucher_type=voucher_type,
                    date=voucher_date,
                    party_type=party_type,
                    customer_id=party_id if party_type == 'customer' else None,
                    supplier_id=party_id if party_type == 'supplier' else None,
                    amount_cash=0.0,
                    amount_gold=float(total_gold_weight),
                    gold_karat=None,
                    description=f"سداد ذهب للمورد - فاتورة {getattr(new_invoice, 'invoice_type_id', '')}".strip(),
                    reference_type='invoice',
                    reference_id=int(new_invoice.id),
                    reference_number=str(getattr(new_invoice, 'invoice_type_id', '') or '') or None,
                    notes=voucher_notes,
                    created_by=posted_by_username or 'system',
                    status='pending',
                )
                db.session.add(voucher)
                db.session.flush()

                # Determine line types based on direction
                safe_line_type = 'debit' if settlement_direction == 'in' else 'credit'
                party_line_type = 'credit' if settlement_direction == 'in' else 'debit'

                # Safe lines (gold)
                for (safe_id, karat_int), weight_val in safe_agg.items():
                    if weight_val <= 0:
                        continue
                    gold_safe = SafeBox.query.get(int(safe_id))
                    if not gold_safe:
                        db.session.rollback()
                        return jsonify({'error': 'gold_safe_box_not_found', 'safe_box_id': safe_id}), 400
                    safe_account_id = getattr(gold_safe, 'account_id', None)
                    if not safe_account_id:
                        db.session.rollback()
                        return jsonify({'error': 'safe_box_missing_account_id', 'safe_box_id': safe_id}), 400
                    db.session.add(VoucherAccountLine(
                        voucher_id=voucher.id,
                        account_id=int(safe_account_id),
                        line_type=safe_line_type,
                        amount_type='gold',
                        amount=round(float(weight_val), 3),
                        karat=int(karat_int),
                        description=f'خزينة ذهب: {getattr(gold_safe, "name", "")}'.strip() or None,
                    ))

                # Party lines (gold) - aggregated by karat
                for karat_int, weight_val in party_agg.items():
                    if weight_val <= 0:
                        continue
                    db.session.add(VoucherAccountLine(
                        voucher_id=voucher.id,
                        account_id=int(party_account_id),
                        line_type=party_line_type,
                        amount_type='gold',
                        amount=round(float(weight_val), 3),
                        karat=int(karat_int),
                        description='طرف السداد',
                    ))

                db.session.flush()

                journal_entry = create_journal_entry_from_voucher(voucher)
                if not journal_entry:
                    db.session.rollback()
                    return jsonify({'error': 'gold_settlement_voucher_post_failed', 'message': 'فشل إنشاء القيد من سند سداد الذهب'}), 500

                voucher.status = 'approved'
                voucher.approved_at = datetime.now()
                voucher.approved_by = posted_by_username or 'system'
                voucher.journal_entry_id = journal_entry.id
                _append_safe_transactions_for_voucher(voucher, created_by=voucher.approved_by)
                gold_settlement_voucher_created = True

        # --- 2. Aggregate Gold and Cash Totals ---
        total_cash = new_invoice.total
        
        # Aggregate weights by karat from invoice items (using DB data)
        gold_by_karat = {'18': 0.0, '21': 0.0, '22': 0.0, '24': 0.0}

        is_customer_scrap_purchase = (
            ((new_invoice.invoice_type or '').strip() in ('شراء من عميل', 'مرتجع شراء'))
            and (str(getattr(new_invoice, 'gold_type', '') or '').strip().lower() == 'scrap')
        )

        def _register_gold_weight(karat_val, weight_val):
            karat_float = _to_float(karat_val, 0.0)
            weight_float = _to_float(weight_val, 0.0)
            if karat_float <= 0 or weight_float <= 0:
                return

            karat_key = str(int(round(karat_float)))
            if karat_key not in gold_by_karat:
                gold_by_karat[karat_key] = 0.0
            gold_by_karat[karat_key] += weight_float

        if not has_valid_karat_lines:
            for item_data in data.get('items', []):
                item_id = item_data.get('item_id')
                item = Item.query.get(item_id) if item_id else None

                karat_value = item_data.get('karat') if item_data.get('karat') not in (None, '') else (item.karat if item else None)
                weight_value = item_data.get('weight') if item_data.get('weight') is not None else (item.weight if item else None)

                if weight_value is None:
                    weight_value = item_data.get('total_weight')

                # gold_by_karat = الوزن الصافي (بدون فصوص) → للقيود المحاسبية وحساب الربح
                total_weight_value = _to_float(weight_value, 0.0)
                _register_gold_weight(karat_value, total_weight_value)

        if karat_lines_data and isinstance(karat_lines_data, list):
            for line_data in karat_lines_data:
                karat_val = line_data.get('karat')
                weight_val = line_data.get('weight_grams', line_data.get('weight', line_data.get('total_weight')))
                _register_gold_weight(karat_val, weight_val)

        # حساب وزن الفصوص الإجمالي من الأصناف — للتتبع المعلوماتي في SafeBoxTransaction فقط
        # لا يدخل في القيود المحاسبية
        _invoice_stones_weight = 0.0
        _invoice_stones_by_karat = {'18': 0.0, '21': 0.0, '22': 0.0, '24': 0.0}
        if str(invoice_type).strip() == 'شراء من عميل' and not has_valid_karat_lines:
            for _item_d in (data.get('items') or []):
                _sw = _to_float(_item_d.get('stones_weight'), 0.0)
                _invoice_stones_weight += _sw
                if _sw > 0:
                    try:
                        _k = str(int(float(_item_d.get('karat') or 0)))
                        if _k in _invoice_stones_by_karat:
                            _invoice_stones_by_karat[_k] += _sw
                    except Exception:
                        pass

        # --- Customer scrap purchase/return: move physical gold through a gold safe (employee or main) ---
        # Also expose the resolved gold safe account for weight journal entries.
        scrap_purchase_gold_safe_account_id = None
        try:
            inv_type_tmp = (new_invoice.invoice_type or '').strip()
            inv_gold_type_tmp = (str(getattr(new_invoice, 'gold_type', '') or '').strip().lower())
            # شراء من عميل دائماً يُعامَل كحركة كسر بغض النظر عن gold_type
            # (inventory resolver hardcodes scrap for this type)
            is_customer_scrap_move = (
                inv_type_tmp == 'شراء من عميل'
                or (inv_type_tmp == 'مرتجع شراء' and inv_gold_type_tmp == 'scrap')
            )
        except Exception:
            is_customer_scrap_move = False

        if is_customer_scrap_move:
            target_gold_safe_id = None
            settings_row = None
            try:
                settings_row = Settings.query.first()
            except Exception:
                settings_row = None
            
            # 1) If an explicit scrap holder employee is provided, honor it.
            if target_gold_safe_id is None:
                try:
                    holder_id = getattr(new_invoice, 'scrap_holder_employee_id', None)
                    if holder_id not in (None, '', False):
                        try:
                            emp_id = int(holder_id)
                        except Exception:
                            emp_id = None
                        if emp_id:
                            holder = Employee.query.get(emp_id)
                            if holder and getattr(holder, 'gold_safe_box_id', None):
                                target_gold_safe_id = int(holder.gold_safe_box_id)
                                print(f"\n🔍 SCRAP_RECEIPT: Using explicit holder employee {emp_id} gold safe: {target_gold_safe_id}")
                except Exception:
                    pass

            # 2) Invoice employee's gold safe — always check, regardless of global flag.
            # Rule: إذا كان الموظف يملك خزينة ذهب → استخدمها تلقائياً.
            if target_gold_safe_id is None:
                try:
                    employee_id_fallback = getattr(new_invoice, 'employee_id', None)
                    if employee_id_fallback not in (None, '', False):
                        try:
                            emp_id = int(employee_id_fallback)
                        except Exception:
                            emp_id = None
                        if emp_id:
                            holder = Employee.query.get(emp_id)
                            if holder and getattr(holder, 'gold_safe_box_id', None):
                                target_gold_safe_id = int(holder.gold_safe_box_id)
                                print(f"\n🔍 SCRAP_RECEIPT: Using invoice employee {emp_id} gold safe: {target_gold_safe_id}")
                except Exception:
                    pass

            # 2b) Fallback: original invoice's employee gold safe (for مرتجع شراء).
            # The return invoice may not carry employee_id, so check original.
            if target_gold_safe_id is None:
                try:
                    _orig = None
                    _orig_id = data.get('original_invoice_id')
                    if _orig_id:
                        _orig = Invoice.query.get(_orig_id)
                    orig_emp_id = getattr(_orig, 'employee_id', None) if _orig else None
                    if orig_emp_id not in (None, '', False):
                        try:
                            orig_emp_id = int(orig_emp_id)
                        except Exception:
                            orig_emp_id = None
                    if orig_emp_id:
                        orig_holder = Employee.query.get(orig_emp_id)
                        if orig_holder and getattr(orig_holder, 'gold_safe_box_id', None):
                            target_gold_safe_id = int(orig_holder.gold_safe_box_id)
                            print(f"\n🔍 SCRAP_RECEIPT: Using ORIGINAL invoice employee {orig_emp_id} gold safe: {target_gold_safe_id}")
                except Exception:
                    pass

            # 3) Fall back to configured main scrap safe.
            if target_gold_safe_id is None:
                try:
                    scrap_sb_id = getattr(settings_row, 'main_scrap_gold_safe_box_id', None) if settings_row else None
                    if scrap_sb_id not in (None, '', 0, '0', False):
                        target_gold_safe_id = int(scrap_sb_id)
                        print(f"\n🔍 SCRAP_RECEIPT: Using main scrap safe from settings: {target_gold_safe_id}")
                except Exception:
                    target_gold_safe_id = None

            # 4) Last resort: any active gold safe.
            if target_gold_safe_id is None:
                try:
                    default_gold_safe = SafeBox.get_default_by_type('gold')
                    if default_gold_safe and default_gold_safe.id:
                        target_gold_safe_id = int(default_gold_safe.id)
                        print(f"\n🔍 SCRAP_RECEIPT: Using default gold safe: {target_gold_safe_id}")
                except Exception:
                    target_gold_safe_id = None

            if target_gold_safe_id is None:
                try:
                    any_gold_safe = SafeBox.query.filter_by(safe_type='gold', is_active=True).order_by(SafeBox.is_default.desc(), SafeBox.id.asc()).first()
                    if any_gold_safe and any_gold_safe.id:
                        target_gold_safe_id = int(any_gold_safe.id)
                except Exception:
                    target_gold_safe_id = None

            # Always resolve the weight target account (used by weight journal entries)
            # even when approval is required, so draft invoices don't fail balance checks.
            if target_gold_safe_id is not None:
                try:
                    gold_safe = SafeBox.query.get(target_gold_safe_id)
                    if gold_safe and (gold_safe.safe_type or '').lower() == 'gold' and bool(getattr(gold_safe, 'is_active', True)):
                        try:
                            if getattr(gold_safe, 'account', None) is not None:
                                scrap_purchase_gold_safe_account_id = int(gold_safe.account.id)
                        except Exception:
                            scrap_purchase_gold_safe_account_id = None

                        # NOTE: avoid double-counting; safebox gold movements are applied only when posting is allowed.
                        if not approval_required:
                            # الخزنة تسجل الوزن الصافي (ذهب فقط) + الفصوص معلوماتياً
                            weight_kwargs = {
                                'weight_18k': float(gold_by_karat.get('18', 0.0) or 0.0),
                                'weight_21k': float(gold_by_karat.get('21', 0.0) or 0.0),
                                'weight_22k': float(gold_by_karat.get('22', 0.0) or 0.0),
                                'weight_24k': float(gold_by_karat.get('24', 0.0) or 0.0),
                                'stones_weight': round(_invoice_stones_weight, 6),
                                'stones_18k':   round(_invoice_stones_by_karat.get('18', 0.0), 6),
                                'stones_21k':   round(_invoice_stones_by_karat.get('21', 0.0), 6),
                                'stones_22k':   round(_invoice_stones_by_karat.get('22', 0.0), 6),
                                'stones_24k':   round(_invoice_stones_by_karat.get('24', 0.0), 6),
                            }
                            has_any_weight = any(v > 0 for v in weight_kwargs.values())
                            if has_any_weight:
                                try:
                                    inv_type_tmp_2 = (new_invoice.invoice_type or '').strip()
                                except Exception:
                                    inv_type_tmp_2 = ''
                                direction = 'out' if inv_type_tmp_2 == 'مرتجع شراء' else 'in'
                                ref_type = 'invoice_scrap_return' if inv_type_tmp_2 == 'مرتجع شراء' else 'invoice_scrap_receipt'
                                notes = 'scrap return' if inv_type_tmp_2 == 'مرتجع شراء' else 'scrap receipt'
                                db.session.add(
                                    SafeBoxTransaction(
                                        safe_box_id=target_gold_safe_id,
                                        ref_type=ref_type,
                                        ref_id=new_invoice.id,
                                        invoice_id=new_invoice.id,
                                        direction=direction,
                                        amount_cash=0.0,
                                        notes=notes,
                                        created_by=posted_by_username,
                                        **weight_kwargs,
                                    )
                                )
                except Exception:
                    pass

        # --- Sale gold movements: withdraw/return physical gold from the configured sale safe ---
        # This complements the weight journal lines and provides an audit ledger for gold safes.
        try:
            inv_type_for_gold = (new_invoice.invoice_type or '').strip()
            inv_gold_type = (str(getattr(new_invoice, 'gold_type', '') or '').strip().lower() or 'new')
        except Exception:
            inv_type_for_gold = ''
            inv_gold_type = 'new'

        # Scrap gold type is handled separately below (invoice_scrap_sale SBT).
        # Avoid creating a duplicate invoice_sale_gold_movement SBT for scrap invoices.
        if (not approval_required) and inv_type_for_gold in ('بيع', 'مرتجع بيع') and inv_gold_type != 'scrap':
            try:
                settings_row = Settings.query.first()
            except Exception:
                settings_row = None

            target_gold_safe_id = None
            try:
                target_gold_safe_id = getattr(settings_row, 'sale_gold_safe_box_id', None) if settings_row else None
            except Exception:
                target_gold_safe_id = None

            try:
                if target_gold_safe_id not in (None, '', 0, '0', False):
                    sb = SafeBox.query.get(int(target_gold_safe_id))
                else:
                    sb = None
            except Exception:
                sb = None

            if sb and (sb.safe_type or '').lower() == 'gold' and bool(getattr(sb, 'is_active', True)):
                try:
                    weight_kwargs = {
                        'weight_18k': float(gold_by_karat.get('18', 0.0) or 0.0),
                        'weight_21k': float(gold_by_karat.get('21', 0.0) or 0.0),
                        'weight_22k': float(gold_by_karat.get('22', 0.0) or 0.0),
                        'weight_24k': float(gold_by_karat.get('24', 0.0) or 0.0),
                    }
                    has_any_weight = any(v > 0 for v in weight_kwargs.values())
                    if has_any_weight:
                        db.session.add(
                            SafeBoxTransaction(
                                safe_box_id=int(sb.id),
                                ref_type='invoice_sale_gold_movement',
                                ref_id=new_invoice.id,
                                invoice_id=new_invoice.id,
                                direction=('out' if inv_type_for_gold == 'بيع' else 'in'),
                                amount_cash=0.0,
                                notes=('sale gold out' if inv_type_for_gold == 'بيع' else 'sale return gold in'),
                                created_by=posted_by_username,
                                **weight_kwargs,
                            )
                        )
                except Exception:
                    pass

        # --- 3. Determine Accounts and Journal Entry Logic ---
        # 🆕 منطق محدث لدعم 6 أنواع من الفواتير
        
        # الحسابات الأساسية
        cash_account = Account.query.filter_by(name='صندوق النقدية').first()
        inventory_account = Account.query.filter_by(name='المخزون').first()
        sales_account = Account.query.filter(Account.name.like('مبيعات%')).first()
        revenue_account = Account.query.filter(Account.name.like('الإيرادات%')).first()
        purchases_account = Account.query.filter_by(name='تكلفة البضاعة المباعة').first()
        
        # حساب الطرف (عميل أو مورد)
        party_account = None
        if new_invoice.customer_id:
            customer = Customer.query.get(new_invoice.customer_id)
            if customer:
                try:
                    if not customer.account_id or not Account.query.get(customer.account_id):
                        ensure_customer_accounts(customer)
                except Exception as exc:
                    return jsonify({
                        'error': 'customer_account_missing',
                        'message': 'تعذر إنشاء/ربط حساب العميل (مالي + مذكرة وزنية).',
                        'details': str(exc),
                    }), 400

                if customer.account_id:
                    party_account = Account.query.get(customer.account_id)
        elif new_invoice.supplier_id:
            supplier = Supplier.query.get(new_invoice.supplier_id)
            if supplier:
                try:
                    if not supplier.account_id or not Account.query.get(supplier.account_id):
                        ensure_supplier_accounts(supplier)
                except Exception as exc:
                    return jsonify({
                        'error': 'supplier_account_missing',
                        'message': 'تعذر إنشاء/ربط حساب المورد (مالي + مذكرة وزنية).',
                        'details': str(exc),
                    }), 400

                if supplier.account_id:
                    party_account = Account.query.get(supplier.account_id)
        
        # إذا لم يكن هناك طرف، استخدم الصندوق
        if not party_account:
            party_account = cash_account

        # معرف حساب العميل/الطرف المستخدم في القيود اللاحقة (مثل القيود الوزنية)
        customer_account_id = None
        # ✅ الصحيح: حساب النقدية الوزني هو 71100 (وليس 7100)
        default_memo_cash_account = Account.query.filter_by(account_number='71100').first()
        default_memo_cash_account_id = default_memo_cash_account.id if default_memo_cash_account else None

        memo_party_account = None
        if party_account and party_account.memo_account_id:
            memo_party_account = Account.query.get(party_account.memo_account_id)
            if not memo_party_account:
                print(
                    f"⚠️ Linked memo account {party_account.memo_account_id} for account {party_account.account_number} not found. "
                    "Falling back to default memo cash account."
                )

        if memo_party_account:
            customer_account_id = memo_party_account.id
        elif default_memo_cash_account_id:
            customer_account_id = default_memo_cash_account_id
        elif party_account and party_account.tracks_weight:
            customer_account_id = party_account.id

        # --- 4. Create Journal Entry ---
        journal_desc = f"فاتورة {invoice_type} رقم #{new_invoice.invoice_type_id}"
        if new_invoice.original_invoice_id:
            journal_desc += f" (مرتبطة بفاتورة #{new_invoice.original_invoice_id})"
        
        # 🔧 توليد رقم القيد (موحّد وآمن ضد التكرار)
        entry_number_str = _generate_journal_entry_number(entry_date=new_invoice.date)
        
        journal_entry = JournalEntry(
            entry_number=entry_number_str,
            date=new_invoice.date,
            description=journal_desc,
            reference_type='invoice',
            reference_id=new_invoice.id,
            created_by=posted_by_username,
            posted_by=posted_by_username,
        )
        db.session.add(journal_entry)
        db.session.flush()

        # --- 5. Create Journal Entry Lines ---
        # 🆕 منطق محدث لدعم 6 أنواع من الفواتير
        
        # تحضير حقول الذهب
        gold_debit_fields = {f"debit_{k}k": v for k, v in gold_by_karat.items() if v > 0}
        gold_credit_fields = {f"credit_{k}k": v for k, v in gold_by_karat.items() if v > 0}
        
        # 🆕 دالة مساعدة لإضافة قيد العمولة وضريبتها
        def add_commission_entry(journal_entry_id, payment_method_obj, commission_amount, commission_vat=0.0):
            """
            ملاحظة: قيود العمولات تُعالج الآن في قسم multi-payment أدناه
            """
        
        # --- القيود حسب نوع الفاتورة ---
        
        # 🆕 الحصول على سعر الذهب الحالي (يلزم لجميع أنواع الفواتير)
        gold_price_data = get_current_gold_price()
        
        if invoice_type == 'بيع':
            # ============================================
            # 1. فاتورة بيع - النظام المحاسبي الصحيح
            # ============================================
            # القيد الأول: إثبات الإيراد الكامل
            #     من حـ/ النقدية [مدين نقد]
            #         إلى حـ/ مبيعات الذهب الجديد [دائن نقد بالمبلغ الكامل]
            # 
            # القيد الثاني: إثبات التكلفة (متوسط سعر الشراء)
            #     من حـ/ تكلفة المبيعات [مدين نقد + وزن]
            #         إلى حـ/ مخزون الذهب عيار XX [دائن نقد + وزن]
            #
            # الربح = الإيراد - التكلفة
            # الربح بالذهب = الربح النقدي ÷ متوسط سعر الشراء
            # ============================================
            
            # الحصول على الحسابات من الربط المحاسبي
            cash_acc_id = get_account_id_for_mapping('بيع', 'cash')
            _sales_new_acc_id  = get_account_id_for_mapping('بيع', 'sales_gold_new') or get_account_id_for_mapping('بيع', 'revenue')
            _sales_scrap_acc_id = get_account_id_for_mapping('بيع', 'sales_gold_scrap') or _sales_new_acc_id
            # اختر حساب المبيعات حسب نوع الذهب
            sales_gold_new_acc_id = _sales_scrap_acc_id if gold_type == 'scrap' else _sales_new_acc_id
            vat_payable_acc_id = get_account_id_for_mapping('بيع', 'vat_payable')
            commission_acc_id = get_account_id_for_mapping('بيع', 'commission')
            commission_vat_acc_id = get_account_id_for_mapping('بيع', 'commission_vat')
            
            # حسابات المخزون: دعم التوحيد (حساب واحد لكل العيارات)
            inventory_accounts = {}
            unified_inventory_acc_id = _resolve_inventory_account_id_for_invoice(invoice_type, gold_type)
            if unified_inventory_acc_id:
                inventory_accounts = {k: unified_inventory_acc_id for k in ['18', '21', '22', '24']}
            else:
                for karat in ['18', '21', '22', '24']:
                    inv_acc_id = get_account_id_for_mapping('بيع', f'inventory_{karat}k')
                    if inv_acc_id:
                        inventory_accounts[karat] = inv_acc_id

            # بيع كسر: استخدام حساب صندوق الكسر للوزن بدلاً من 7130001 (مخزون وزني)
            # يضمن التناسق مع سطر الشراء الذي يُدبن 71310000 (صندوق الكسر الرئيسي وزني)
            scrap_sale_safe_account_id = None
            _scrap_sale_target_sb_id = None
            if gold_type == 'scrap':
                try:
                    _sale_settings = Settings.query.first()
                    _scrap_sale_sb_id = getattr(_sale_settings, 'main_scrap_gold_safe_box_id', None) if _sale_settings else None
                    if not _scrap_sale_sb_id:
                        _fallback_sale_sb = SafeBox.query.filter_by(safe_type='gold', is_active=True).order_by(SafeBox.is_default.desc(), SafeBox.id.asc()).first()
                        if _fallback_sale_sb:
                            _scrap_sale_sb_id = _fallback_sale_sb.id
                    if _scrap_sale_sb_id:
                        _sale_gold_safe = SafeBox.query.get(int(_scrap_sale_sb_id))
                        if _sale_gold_safe and getattr(_sale_gold_safe, 'account', None):
                            scrap_sale_safe_account_id = int(_sale_gold_safe.account.id)
                            _scrap_sale_target_sb_id = int(_sale_gold_safe.id)
                            # Override: كل العيارات تشير لحساب صندوق الكسر المالي
                            # _resolve_weight_account_id سيحوله لـ 71310000 عند إنشاء قيد الوزن
                            inventory_accounts = {k: scrap_sale_safe_account_id for k in ['18', '21', '22', '24']}
                except Exception:
                    scrap_sale_safe_account_id = None
                    _scrap_sale_target_sb_id = None

            # ✅ تحقق مبكر: منع إنشاء قيود بحساب None وإرجاع رسالة واضحة
            missing = []
            if not sales_gold_new_acc_id:
                missing.append({'mapping': 'sales_gold_new/revenue', 'operation_type': 'بيع'})

            # If VAT exists on items, VAT payable mapping is required to keep JE balanced.
            try:
                _tax_total_for_check = sum(
                    _to_float(
                        it.get('tax_amount', it.get('tax', 0.0)),
                        0.0,
                    )
                    for it in (data.get('items') or [])
                    if isinstance(it, dict)
                )
                if _tax_total_for_check < 0:
                    _tax_total_for_check = abs(_tax_total_for_check)
            except Exception:
                _tax_total_for_check = 0.0

            if _tax_total_for_check > 0.009 and not vat_payable_acc_id:
                missing.append({'mapping': 'vat_payable', 'operation_type': 'بيع'})

            # If commissions are being applied at invoice time, commission mappings are required.
            try:
                if float(commission_amount or 0.0) > 0.009 and not commission_acc_id:
                    missing.append({'mapping': 'commission', 'operation_type': 'بيع'})
            except Exception:
                pass

            try:
                if float(commission_vat_total or 0.0) > 0.009 and not (commission_vat_acc_id or commission_acc_id):
                    missing.append({'mapping': 'commission_vat', 'operation_type': 'بيع'})
            except Exception:
                pass

            # تحقق من حساب المخزون المطلوب فعلياً حسب عناصر الفاتورة
            try:
                required_karats = {
                    str(_to_float(item.get('karat', 0), 0.0)).split('.')[0]
                    for item in (data.get('items') or [])
                    if _to_float(item.get('weight', 0), 0.0) > 0
                }
            except Exception:
                required_karats = set()

            for k in sorted(required_karats):
                if k in {'18', '21', '22', '24'} and not inventory_accounts.get(k):
                    missing.append({'mapping': f'inventory_{k}k', 'operation_type': 'بيع'})

            if missing:
                db.session.rollback()
                try:
                    missing_keys = ', '.join(sorted({str(m.get('mapping')) for m in missing if isinstance(m, dict) and m.get('mapping')}))
                except Exception:
                    missing_keys = ''
                return jsonify({
                    'error': 'account_mapping_missing',
                    'message': (
                        'نقص في ربط الحسابات المطلوبة لإنشاء قيد فاتورة البيع. '
                        'الرجاء ضبط Accounting Mapping أو التأكد من وجود الحسابات الافتراضية.'
                        + (f' (المطلوب: {missing_keys})' if missing_keys else '')
                    ),
                    'missing': missing,
                    'resolved': {
                        'cash_acc_id': cash_acc_id,
                        'sales_gold_new_acc_id': sales_gold_new_acc_id,
                        'inventory_accounts': inventory_accounts,
                    },
                }), 400
            
            # ─── حساب الضريبة من العناصر ───
            _total_tax_for_je = sum(
                _to_float(
                    item_data.get('tax_amount', item_data.get('tax', 0.0)),
                    0.0,
                )
                for item_data in data.get('items', [])
            )
            if _total_tax_for_je < 0:
                _total_tax_for_je = abs(_total_tax_for_je)
            # Fallback: total_tax المحسوب من karat_lines ومخزّن في الفاتورة
            if _total_tax_for_je == 0.0:
                _total_tax_for_je = float(new_invoice.total_tax or 0.0)

            # ─── مجموع المدفوع + العمولات (تُخصم من مبلغ الذمم لتوازن القيد) ───
            paid_amount_total = 0.0
            _total_commission_je = 0.0   # إجمالي العمولات المُدرجة في قيد الفاتورة

            if payments_data and len(payments_data) > 0:
                for payment in payments_data:
                    pm_obj = PaymentMethod.query.get(payment['payment_method_id'])
                    pm_amount = _to_float(payment.get('amount', 0.0))
                    paid_amount_total += pm_amount
                    try:
                        pm_commission_timing = str(
                            getattr(pm_obj, 'commission_timing', 'invoice') or 'invoice'
                        ).strip().lower()
                    except Exception:
                        pm_commission_timing = 'invoice'

                    if pm_commission_timing == 'settlement':
                        pm_commission = 0.0
                        pm_commission_vat = 0.0
                    else:
                        pm_commission = _to_float(payment.get('commission_amount', 0.0))
                        pm_commission_vat = _to_float(payment.get('commission_vat', 0.0))

                    if pm_commission > 0 and commission_acc_id:
                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=commission_acc_id,
                            cash_debit=pm_commission,
                            description=f"عمولة {pm_obj.name if pm_obj else ''}",
                            apply_golden_rule=False,
                        )
                        _total_commission_je += pm_commission
                    _vat_debit_pm = commission_vat_acc_id or commission_acc_id
                    if pm_commission_vat > 0 and _vat_debit_pm:
                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=_vat_debit_pm,
                            cash_debit=pm_commission_vat,
                            description=f"ضريبة عمولة {pm_obj.name if pm_obj else ''}",
                            apply_golden_rule=False,
                        )
                        _total_commission_je += pm_commission_vat

            elif payment_method_id:
                try:
                    pm_commission_timing = str(
                        getattr(payment_method_obj, 'commission_timing', 'invoice') or 'invoice'
                    ).strip().lower()
                except Exception:
                    pm_commission_timing = 'invoice'

                if pm_commission_timing == 'settlement':
                    commission_amount = 0.0
                    commission_vat_total = 0.0

                paid_amount_total = total_cash

                if commission_amount > 0 and commission_acc_id:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=commission_acc_id,
                        cash_debit=commission_amount,
                        description="عمولة الدفع",
                        apply_golden_rule=False,
                    )
                    _total_commission_je += commission_amount
                _vat_debit_single = commission_vat_acc_id or commission_acc_id
                if commission_vat_total > 0 and _vat_debit_single:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=_vat_debit_single,
                        cash_debit=commission_vat_total,
                        description="ضريبة عمولة الدفع",
                        apply_golden_rule=False,
                    )
                    _total_commission_je += commission_vat_total
            else:
                paid_amount_total = total_cash

            # ─── je_engine_v2: قيد البيع الكامل ───
            # (ذمم العميل + مبيعات + ضريبة + مخزون وزني + وزن العميل)
            # الذمم = total_cash − العمولات المُدرجة أعلاه (لتوازن القيد)
            import je_adapter as _je_adapter  # noqa: local import

            _je_party = (
                party_account
                or (Account.query.get(customer_account_id) if customer_account_id else None)
                or (Account.query.get(default_memo_cash_account_id) if default_memo_cash_account_id else None)
            )
            _ar_cash_for_je = round(total_cash - _total_commission_je, 2)

            if _je_party:
                _je_adapter.sale_je_for_invoice(
                    journal_entry_id=journal_entry.id,
                    invoice_type=invoice_type,
                    gold_type=gold_type,
                    get_mapping_fn=get_account_id_for_mapping,
                    inventory_accounts=inventory_accounts,
                    gold_by_karat=gold_by_karat,
                    sales_account_id=sales_gold_new_acc_id,
                    vat_payable_account_id=vat_payable_acc_id,
                    ar_account_id=_je_party.id,
                    customer_account_obj=_je_party,
                    total_cash=_ar_cash_for_je,
                    total_tax=_total_tax_for_je,
                    customer_id=new_invoice.customer_id,
                )
            else:
                # Fallback: لا يوجد حساب عميل — ندين الصندوق مباشرة
                _fb_cash = cash_acc_id or (cash_account.id if cash_account else None)
                if _fb_cash:
                    _sales_amt_fb = total_cash - _total_tax_for_je
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=_fb_cash,
                        cash_debit=total_cash,
                        description="استلام نقدي (بدون حساب عميل)",
                        apply_golden_rule=False,
                    )
                    if sales_gold_new_acc_id:
                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=sales_gold_new_acc_id,
                            cash_credit=_sales_amt_fb,
                            description="مبيعات ذهب",
                            apply_golden_rule=False,
                        )
                    if _total_tax_for_je > 0 and vat_payable_acc_id:
                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=vat_payable_acc_id,
                            cash_credit=_total_tax_for_je,
                            description="ضريبة القيمة المضافة",
                            apply_golden_rule=False,
                        )

            # بيع كسر: تسجيل خروج الوزن من صندوق الكسر (يقابل دخول عند الشراء)
            if _scrap_sale_target_sb_id and not approval_required:
                _sale_sbt_weights = {
                    'weight_18k': float(gold_by_karat.get('18', 0.0) or 0.0),
                    'weight_21k': float(gold_by_karat.get('21', 0.0) or 0.0),
                    'weight_22k': float(gold_by_karat.get('22', 0.0) or 0.0),
                    'weight_24k': float(gold_by_karat.get('24', 0.0) or 0.0),
                }
                if any(v > 0 for v in _sale_sbt_weights.values()):
                    db.session.add(SafeBoxTransaction(
                        safe_box_id=_scrap_sale_target_sb_id,
                        ref_type='invoice_scrap_sale',
                        ref_id=new_invoice.id,
                        invoice_id=new_invoice.id,
                        direction='out',
                        amount_cash=0.0,
                        notes='scrap sale - gold out',
                        created_by=posted_by_username,
                        **_sale_sbt_weights,
                    ))

            # ─── المصنعية: استهلاك منفصل ───
            _wage_cash_sale = 0.0
            for _wd in data.get('items', []):
                _wage_cash_sale += (
                    _to_float(_wd.get('wage', 0), 0.0)
                    * _to_float(_wd.get('weight', 0), 0.0)
                )
            if karat_lines_data and isinstance(karat_lines_data, list):
                for _kl in karat_lines_data:
                    _wr = _to_float(_kl.get('manufacturing_wage_cash', 0), 0.0)
                    _ww = _to_float(
                        _kl.get('weight_grams', _kl.get('weight', _kl.get('total_weight'))), 0.0
                    )
                    _wage_cash_sale += _wr * _ww

            if _wage_cash_sale > 0:
                _wage_inv_acc = (
                    _get_manufacturing_wage_inventory_account_id()
                    or get_account_id_by_number('1320')
                    or get_account_id_by_number('1350')
                )
                _wage_exp_acc = (
                    get_account_id_for_mapping('بيع', 'manufacturing_wage')
                    or _ensure_manufacturing_wage_expense_account()
                    or get_account_id_for_mapping('بيع', 'operating_expenses')
                    or get_account_id_by_number('51')
                )
                if _wage_inv_acc and _wage_exp_acc:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=_wage_exp_acc,
                        cash_debit=round(_wage_cash_sale, 2),
                        description="استهلاك أجور المصنعية - مصروفات",
                        apply_golden_rule=False,
                    )
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=_wage_inv_acc,
                        cash_credit=round(_wage_cash_sale, 2),
                        description="خصم من مخزون أجور المصنعية",
                        apply_golden_rule=False,
                    )

            # ─── حساب الربح + تسكير الوزن ───
            _gold_price_now = get_current_gold_price()
            _direct_price_main = _gold_price_now.get(
                'price_per_gram_main_karat',
                _gold_price_now.get('price_main_karat', 350.0),
            )
            total_weight_sold = sum(w for w in gold_by_karat.values() if w > 0)
            invoice_total_tax = new_invoice.total_tax or 0.0

            # رسوم وسائل الدفع: نقرأ من InvoicePayment مباشرة لأن timing=settlement
            # يُصفّر commission_amount المحلي لكن الرسوم مُخزنة بشكل صحيح في InvoicePayment
            _total_pm_fees = round(
                sum(float(p.commission_amount or 0) for p in (new_invoice.payments or [])), 2
            )
            new_invoice.commission_amount = _total_pm_fees

            profit_cash = (
                new_invoice.total
                - invoice_total_tax
                - (new_invoice.total_cost or 0.0)
                - _total_pm_fees
            )
            profit_gold = (profit_cash / _direct_price_main) if _direct_price_main > 0 else 0
            new_invoice.profit_cash = round(profit_cash, 2)
            new_invoice.profit_gold = round(profit_gold, 3)
            new_invoice.profit_weight_price_per_gram = round(_direct_price_main, 4)

            try:
                closing_price = _coerce_float(
                    data.get('weight_closing_price')
                    or data.get('close_price_per_gram')
                    or new_invoice.profit_weight_price_per_gram,
                    0.0,
                )
                if closing_price <= 0:
                    price_snapshot = get_current_gold_price()
                    closing_price = price_snapshot.get('price_per_gram_24k', 0.0)
                if closing_price > 0:
                    _upsert_weight_closing_order(
                        new_invoice,
                        close_price_per_gram=closing_price,
                        settings=_load_weight_closing_settings(),
                    )
            except Exception as exc:
                print(f"⚠️ Failed to initialize weight closing order for invoice {new_invoice.id}: {exc}")
        
        elif invoice_type == 'شراء من عميل':
            # ============================================
            # 2. شراء كسر من عميل - تطبيق القاعدة الذهبية
            # ============================================
            # القاعدة:
            # - المشتريات: مدين بالقيمة النقدية (حساب مشتريات)
            # - العميل/الصندوق: دائن بالقيمة النقدية
            # - المخزون الوزني: مدين بالجرام (حساب وزني)
            # - وزن العميل: دائن بالجرام (حساب وزني العميل)
            # ============================================

            # الحصول على الحسابات
            cash_acc_id = get_account_id_for_mapping('شراء من عميل', 'cash')
            vat_receivable_acc_id = get_account_id_for_mapping('شراء من عميل', 'vat_receivable')

            # حساب المشتريات (مدين نقداً) — 512 كسر / 511 جديد
            # ملاحظة: لا نستخدم 510 فهو أجور مصنعية وليس مشتريات
            _purchases_key = 'purchases_gold_scrap' if gold_type == 'scrap' else 'purchases_gold_new'
            purchases_acc_id = (
                get_account_id_for_mapping('شراء من عميل', _purchases_key)
                or get_account_id_for_mapping('شراء من عميل', 'purchases_gold')
                or get_account_id_for_mapping('شراء من عميل', 'purchases')
                or get_account_id_by_number('512' if gold_type == 'scrap' else '511')
                or get_account_id_by_number('511')
            )

            # حسابات المخزون الوزني: دعم التوحيد (حساب واحد لكل العيارات)
            inventory_accounts = {}
            unified_inventory_acc_id = _resolve_inventory_account_id_for_invoice(invoice_type, gold_type)
            if unified_inventory_acc_id:
                inventory_accounts = {k: unified_inventory_acc_id for k in ['18', '21', '22', '24']}
            else:
                for karat in ['18', '21', '22', '24']:
                    inv_acc_id = get_account_id_for_mapping('شراء من عميل', f'inventory_{karat}k')
                    if inv_acc_id:
                        inventory_accounts[karat] = inv_acc_id
            
            # ✅ الحصول على السعر المباشر للذهب (العيار الرئيسي)
            gold_price_data = get_current_gold_price()
            direct_gold_price_main = gold_price_data.get('price_per_gram_main_karat', 
                                                         gold_price_data.get('price_main_karat', 350.0))
            
            # ============================================
            # A) القيود المالية (نقد فقط)
            # ============================================
            
            # 1. مدين المشتريات (نقد فقط - الوزن في حساب المذكرة الوزنية)
            _purchase_cash_acc = purchases_acc_id or unified_inventory_acc_id
            if _purchase_cash_acc:
                create_dual_journal_entry(
                    journal_entry_id=journal_entry.id,
                    account_id=_purchase_cash_acc,
                    cash_debit=total_cash,
                    apply_golden_rule=False,
                    description=f"شراء ذهب {'كسر' if gold_type == 'scrap' else 'جديد'} من عميل"
                )
            
            # 2. دائن حساب النقدية (من الخزينة)
            # Resolve payment credit account.
            # - Normal purchase: credit cash/bank safe box account
            # - Barter/offset: credit customer account (liability) so it can net with the sale receivable
            settlement_method_raw = (
                (getattr(new_invoice, 'settlement_method', None) or data.get('settlement_method') or '')
            )
            settlement_method_key = str(settlement_method_raw).strip().lower()
            is_offset_settlement = settlement_method_key in ('offset', 'barter', 'trade', 'swap')

            # ================================================================
            # نمط "البيع" المعكوس: قيد الفاتورة يُدائن حساب العميل (AR)،
            # وسند الصرف (Phase 1) يُدين AR ويُدائن الخزينة النقدية.
            # النتيجة: AR يصفر، والخزينة تُحرَّك مرة واحدة فقط من السند.
            # ================================================================
            acc_id = None
            if is_offset_settlement:
                # مقايضة/تقاص: استخدام AR مباشرة (يتقاطع مع الجانب الدائن لفاتورة البيع)
                acc_id = party_account.id if party_account else (cash_acc_id or (cash_account.id if cash_account else None))
            else:
                # دفع نقدي: دائن AR (كما في بيع يُدين AR) ← السند يُغلق AR مقابل الخزينة
                acc_id = party_account.id if party_account else (cash_acc_id or (cash_account.id if cash_account else None))

            if not acc_id:
                db.session.rollback()
                return jsonify({
                    'error': 'cash_account_missing',
                    'message': 'لا يوجد حساب لتسجيل مقابل الشراء (نقداً/تقاص). الرجاء ضبط ربط الحسابات أو حساب العميل.',
                }), 400

            if not Account.query.get(acc_id):
                db.session.rollback()
                return jsonify({
                    'error': 'account_not_found',
                    'message': 'الحساب غير موجود في شجرة الحسابات',
                    'account_id': acc_id,
                }), 400

            create_dual_journal_entry(
                journal_entry_id=journal_entry.id,
                account_id=acc_id,
                cash_credit=total_cash,
                apply_golden_rule=False,
                description=("تقاص/مقايضة شراء ذهب" if is_offset_settlement else "ذمم عميل - مستحق الدفع مقابل شراء الكسر")
            )
            
            # ─── قيود الوزن (je_engine_v2) ───
            # مدين: حسابات المخزون الوزنية (ذهب يدخل المخزون)
            # دائن: حساب وزن العميل (ذهب يخرج من العميل)
            import je_adapter as _je_adapter  # noqa: local import

            _je_party_purchase = (
                party_account
                or (Account.query.get(customer_account_id) if customer_account_id else None)
            )
            if _je_party_purchase:
                _purch_weight_override = (
                    _je_adapter._resolve_weight_account_id(purchases_acc_id)
                    if purchases_acc_id else None
                )
                _je_adapter.weight_entries_for_party(
                    journal_entry_id=journal_entry.id,
                    gold_by_karat=gold_by_karat,
                    inventory_accounts=inventory_accounts,
                    party_account_obj=_je_party_purchase,
                    direction='purchase',
                    customer_id=new_invoice.customer_id,
                    scrap_purchase_gold_safe_account_id=scrap_purchase_gold_safe_account_id,
                    party_weight_account_override=_purch_weight_override,
                )
            else:
                # Fallback: لا يوجد حساب عميل — قيود وزنية مباشرة
                for _karat, _w in gold_by_karat.items():
                    if _w > 0 and _karat in inventory_accounts:
                        _inv_id = inventory_accounts[_karat]
                        _target = (
                            scrap_purchase_gold_safe_account_id
                            or get_account_id_by_number('7521')
                            or _inv_id
                        )
                        if _target:
                            create_dual_journal_entry(
                                journal_entry_id=journal_entry.id,
                                account_id=_target,
                                **{f'weight_{_karat}k_debit': _w},
                                description=f"دخول وزني عيار {_karat}",
                                apply_golden_rule=False,
                            )

            # قيد ضريبة القيمة المضافة (إن وجدت)
            _total_vat = data.get('total_tax', 0)
            if _total_vat and float(_total_vat) > 0 and vat_receivable_acc_id:
                create_dual_journal_entry(
                    journal_entry_id=journal_entry.id,
                    account_id=vat_receivable_acc_id,
                    cash_debit=float(_total_vat),
                    description="ضريبة القيمة المضافة",
                    apply_golden_rule=False,
                )

            # الفصوص لا تدخل القيود — تُتابَع معلوماتياً عبر SafeBoxTransaction.stones_weight فقط

            # ─── نقاط السباق: profit_gold = وزن مشترى - مقابله بسعر السوق ───
            # المنطق: الموظف دفع X ريال واستلم Y جرام
            # مقابل X ريال بسعر اليوم = X / سعر الجرام (عيار رئيسي)
            # الربح الحقيقي = Y - (X / سعر_الجرام) → كمية الذهب الزائدة عن السعر
            # نفس مبدأ المبيعات: profit_gold = profit_cash / سعر_الجرام
            _mk = get_main_karat() or 21
            _weight_main_karat = sum(
                float(w or 0) * int(k) / _mk
                for k, w in gold_by_karat.items()
                if float(w or 0) > 0
            )
            _gold_price_now2 = get_current_gold_price()
            _price_main2 = _gold_price_now2.get(
                'price_per_gram_main_karat',
                _gold_price_now2.get('price_main_karat', 350.0),
            )
            _cash_paid = float(total_cash or 0)
            _cost_in_gold = (_cash_paid / _price_main2) if _price_main2 > 0 else _weight_main_karat
            _purchase_profit_gold = max(0.0, _weight_main_karat - _cost_in_gold)
            new_invoice.profit_gold = round(_purchase_profit_gold, 3)

        elif invoice_type == 'مرتجع بيع':
            # 3. مرتجع بيع — عكس كامل لفاتورة البيع الأصلية
            # ══════════════════════════════════════════════════════════
            # القيود النقدية:
            #   مدين  مردودات المبيعات (بدون ضريبة)
            #   مدين  ضريبة القيمة المضافة (عكس الدائن في البيع)
            #   دائن  العميل/الصندوق (المبلغ الكامل المسترد)
            # القيود الوزنية:
            #   مدين  المخزون الوزني (ذهب يعود)
            #   دائن  وزن العميل (يخرج من رصيده)
            # ══════════════════════════════════════════════════════════

            cash_acc_id          = get_account_id_for_mapping('مرتجع بيع', 'cash')
            customers_acc_id     = get_account_id_for_mapping('مرتجع بيع', 'customers')
            sales_returns_acc_id = get_account_id_for_mapping('مرتجع بيع', 'sales_returns')
            if not sales_returns_acc_id:
                sales_returns_acc_id = (
                    get_account_id_by_number('420')
                    or get_account_id_for_mapping('بيع', 'revenue')
                    or get_account_id_for_mapping('بيع', 'sales_gold_new')
                )
            vat_payable_acc_id = get_account_id_for_mapping('بيع', 'vat_payable')

            # حسابات المخزون
            inventory_accounts = {}
            unified_inventory_acc_id = _resolve_inventory_account_id_for_invoice(invoice_type, gold_type)
            if unified_inventory_acc_id:
                inventory_accounts = {k: unified_inventory_acc_id for k in ['18', '21', '22', '24']}
            else:
                for karat in ['18', '21', '22', '24']:
                    inv_acc_id = get_account_id_for_mapping('مرتجع بيع', f'inventory_{karat}k')
                    if inv_acc_id:
                        inventory_accounts[karat] = inv_acc_id

            # بيع كسر مرتجع: استخدام حساب صندوق الكسر للوزن بدلاً من مخزون الكسر
            # العام، ليطابق فاتورة البيع الأصلية (نفس منطق بيع الكسر في فرع 'بيع'
            # أعلاه). بدون هذا، الوزن المرتجع يدخل حساب مختلف عن الذي خرج منه فعلياً.
            _ret_scrap_sale_safe_account_id = None
            _ret_scrap_sale_target_sb_id = None
            if gold_type == 'scrap':
                try:
                    _ret_sale_settings = Settings.query.first()
                    _ret_scrap_sale_sb_id = getattr(_ret_sale_settings, 'main_scrap_gold_safe_box_id', None) if _ret_sale_settings else None
                    if not _ret_scrap_sale_sb_id:
                        _ret_fallback_sale_sb = SafeBox.query.filter_by(safe_type='gold', is_active=True).order_by(SafeBox.is_default.desc(), SafeBox.id.asc()).first()
                        if _ret_fallback_sale_sb:
                            _ret_scrap_sale_sb_id = _ret_fallback_sale_sb.id
                    if _ret_scrap_sale_sb_id:
                        _ret_sale_gold_safe = SafeBox.query.get(int(_ret_scrap_sale_sb_id))
                        if _ret_sale_gold_safe and getattr(_ret_sale_gold_safe, 'account', None):
                            _ret_scrap_sale_safe_account_id = int(_ret_sale_gold_safe.account.id)
                            _ret_scrap_sale_target_sb_id = int(_ret_sale_gold_safe.id)
                            # Override: كل العيارات تشير لحساب صندوق الكسر المالي
                            inventory_accounts = {k: _ret_scrap_sale_safe_account_id for k in ['18', '21', '22', '24']}
                except Exception:
                    _ret_scrap_sale_safe_account_id = None
                    _ret_scrap_sale_target_sb_id = None

            # ─── الضريبة ───
            # المصدر الأول: items المُرسلة من الفرونت (tax_amount)
            _total_tax_ret = sum(
                _to_float(item_data.get('tax_amount', item_data.get('tax', 0.0)), 0.0)
                for item_data in data.get('items', [])
            )
            if _total_tax_ret < 0:
                _total_tax_ret = abs(_total_tax_ret)
            # Fallback: إذا الفرونت لم يُرسل tax في items، استخدم total_tax المخزّن في الفاتورة
            # (يُحسب من karat_lines ومخزّن بشكل موثوق في new_invoice.total_tax)
            if _total_tax_ret == 0.0:
                _total_tax_ret = float(new_invoice.total_tax or 0.0)
            # Second fallback: من الفاتورة الأصلية إن وُجدت
            if _total_tax_ret == 0.0 and data.get('original_invoice_id'):
                try:
                    _orig_for_tax = Invoice.query.get(int(data['original_invoice_id']))
                    if _orig_for_tax:
                        _total_tax_ret = float(_orig_for_tax.total_tax or 0.0)
                except Exception:
                    pass

            # ─── استخدام je_adapter كامل (يشمل VAT + وزن) ───
            import je_adapter as _je_adapter  # noqa: local import
            _je_party_sale_ret = (
                party_account
                or (Account.query.get(customers_acc_id) if customers_acc_id else None)
            )
            acc_id = customers_acc_id or cash_acc_id or (party_account.id if party_account else None)

            if _je_party_sale_ret and inventory_accounts and sales_returns_acc_id:
                _je_adapter.sale_return_je_for_invoice(
                    journal_entry_id=journal_entry.id,
                    invoice_type=invoice_type,
                    gold_type=gold_type,
                    get_mapping_fn=get_account_id_for_mapping,
                    inventory_accounts=inventory_accounts,
                    gold_by_karat=gold_by_karat,
                    sales_returns_account_id=sales_returns_acc_id,
                    vat_payable_account_id=vat_payable_acc_id,
                    ar_account_id=_je_party_sale_ret.id,
                    customer_account_obj=_je_party_sale_ret,
                    return_cash=total_cash,
                    total_tax=_total_tax_ret,
                    cash_refunded=total_cash,
                    customer_id=new_invoice.customer_id,
                )
            else:
                # Fallback manual entries if adapter unavailable
                _net_ret = round(total_cash - _total_tax_ret, 2)
                if sales_returns_acc_id:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=sales_returns_acc_id,
                        cash_debit=_net_ret,
                        apply_golden_rule=False,
                        description="مردودات المبيعات",
                    )
                if _total_tax_ret > 0 and vat_payable_acc_id:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=vat_payable_acc_id,
                        cash_debit=_total_tax_ret,
                        apply_golden_rule=False,
                        description="عكس ضريبة القيمة المضافة - مرتجع",
                    )
                if acc_id:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=acc_id,
                        cash_credit=total_cash,
                        apply_golden_rule=False,
                        description="استرداد نقدي للعميل",
                    )

            # ─── عكس عمولات وسائل الدفع ───
            _commission_acc_id     = get_account_id_for_mapping('بيع', 'commission')
            _commission_vat_acc_id = get_account_id_for_mapping('بيع', 'commission_vat')
            if payments_data:
                for payment in payments_data:
                    pm_obj = PaymentMethod.query.get(payment['payment_method_id'])
                    try:
                        pm_commission_timing = str(
                            getattr(pm_obj, 'commission_timing', 'invoice') or 'invoice'
                        ).strip().lower()
                    except Exception:
                        pm_commission_timing = 'invoice'
                    if pm_commission_timing == 'settlement':
                        continue
                    pm_commission     = _to_float(payment.get('commission_amount', 0.0))
                    pm_commission_vat = _to_float(payment.get('commission_vat', 0.0))
                    if pm_commission > 0 and _commission_acc_id:
                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=_commission_acc_id,
                            cash_credit=pm_commission,
                            apply_golden_rule=False,
                            description=f"عكس عمولة {pm_obj.name if pm_obj else ''} - مرتجع",
                        )
                    _vat_credit_pm = _commission_vat_acc_id or _commission_acc_id
                    if pm_commission_vat > 0 and _vat_credit_pm:
                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=_vat_credit_pm,
                            cash_credit=pm_commission_vat,
                            apply_golden_rule=False,
                            description=f"عكس ضريبة عمولة {pm_obj.name if pm_obj else ''} - مرتجع",
                        )

            # ─── عكس أجور المصنعية ───
            _wage_cash_ret = 0.0
            for _wd in data.get('items', []):
                _wage_cash_ret += (
                    _to_float(_wd.get('wage', 0), 0.0)
                    * _to_float(_wd.get('weight', 0), 0.0)
                )
            if karat_lines_data and isinstance(karat_lines_data, list):
                for _kl in karat_lines_data:
                    _wage_cash_ret += (
                        _to_float(_kl.get('manufacturing_wage_cash', 0), 0.0)
                        * _to_float(_kl.get('weight_grams', _kl.get('weight', 0)), 0.0)
                    )
            if _wage_cash_ret > 0:
                _wage_inv_acc = (
                    _get_manufacturing_wage_inventory_account_id()
                    or get_account_id_by_number('1320')
                    or get_account_id_by_number('1350')
                )
                _wage_exp_acc = (
                    get_account_id_for_mapping('بيع', 'manufacturing_wage')
                    or _ensure_manufacturing_wage_expense_account()
                    or get_account_id_by_number('51')
                )
                if _wage_inv_acc and _wage_exp_acc:
                    # عكس: مدين مخزون المصنعية، دائن مصروف المصنعية
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=_wage_inv_acc,
                        cash_debit=round(_wage_cash_ret, 2),
                        apply_golden_rule=False,
                        description="إعادة أجور المصنعية للمخزون - مرتجع",
                    )
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=_wage_exp_acc,
                        cash_credit=round(_wage_cash_ret, 2),
                        apply_golden_rule=False,
                        description="عكس استهلاك أجور المصنعية - مرتجع",
                    )

            # ─── SBT صندوق الكسر (بيع كسر مرتجع → الكسر يعود للصندوق) ───
            if gold_type == 'scrap' and not approval_required:
                try:
                    # نعيد استخدام نفس الخزينة المُحدَّدة أعلاه لحساب المخزون
                    # (inventory_accounts override) لضمان التطابق بين القيد و SBT.
                    _ret_scrap_sb_id = _ret_scrap_sale_target_sb_id
                    if not _ret_scrap_sb_id:
                        _ret_settings = Settings.query.first()
                        _ret_scrap_sb_id = getattr(_ret_settings, 'main_scrap_gold_safe_box_id', None) if _ret_settings else None
                        if not _ret_scrap_sb_id:
                            _fb = SafeBox.query.filter_by(safe_type='gold', is_active=True).order_by(SafeBox.is_default.desc(), SafeBox.id.asc()).first()
                            if _fb:
                                _ret_scrap_sb_id = _fb.id
                    if _ret_scrap_sb_id:
                        _ret_sbt_weights = {
                            'weight_18k': float(gold_by_karat.get('18', 0.0) or 0.0),
                            'weight_21k': float(gold_by_karat.get('21', 0.0) or 0.0),
                            'weight_22k': float(gold_by_karat.get('22', 0.0) or 0.0),
                            'weight_24k': float(gold_by_karat.get('24', 0.0) or 0.0),
                        }
                        if any(v > 0 for v in _ret_sbt_weights.values()):
                            db.session.add(SafeBoxTransaction(
                                safe_box_id=int(_ret_scrap_sb_id),
                                ref_type='invoice_scrap_return_receipt',
                                ref_id=new_invoice.id,
                                invoice_id=new_invoice.id,
                                direction='in',
                                amount_cash=0.0,
                                notes='scrap sale return - gold back in safe',
                                created_by=posted_by_username,
                                **_ret_sbt_weights,
                            ))
                except Exception as _sbt_exc:
                    print(f"⚠️ SBT for sale return scrap failed: {_sbt_exc}")

            # ─── ملاحظة: سندات الصرف (Voucher payment) و SBTs لمرتجع بيع ───
            # تُنشأ تلقائياً في قسم معالجة الدفعات المشترك (أعلاه)
            # عبر _direction_for_invoice_type('مرتجع بيع') → 'out' → voucher_type='payment'
            # ثم _append_safe_transactions_for_voucher. لا حاجة لتكرارها هنا.

            # ─── إلغاء/تعديل أمر تسكير الوزن للفاتورة الأصلية ───
            # ملاحظة: total_weight_main_karat هو الحقل الذي يجب تقليصه، لا
            # remaining_weight_main_karat — لأن _auto_consume_weight_closing
            # يعيد حساب remaining دائماً من (total - executed) عند كل تنفيذ
            # لاحق، فأي تعديل مباشر عليه فقط يُمحى ولا يدوم.
            # كما أن أسماء الحالة الصحيحة هي 'open' / 'partially_closed' /
            # 'closed' (لا 'partial' التي لا تُستخدم في أي مكان آخر بالكود).
            # إن كان جزء من الكمية المرتجعة قد سُوّي نقدياً مسبقاً (قيد تكلفة
            # مبيعات COGS عبر تنفيذ تسكير سابق)، يجب عكس ذلك القيد بدلاً من
            # ترك نقد خارج حساب المخزون المالي لذهب رجع فعلياً للمحل.
            _orig_inv_id = data.get('original_invoice_id')
            if _orig_inv_id:
                try:
                    _orig_wco = WeightClosingOrder.query.filter_by(
                        invoice_id=int(_orig_inv_id)
                    ).first()
                    if _orig_wco and _orig_wco.status in ('open', 'partially_closed', 'closed'):
                        _ret_weight_mk = sum(
                            float(w or 0) * int(k) / (get_main_karat() or 21)
                            for k, w in gold_by_karat.items()
                            if float(w or 0) > 0
                        )
                        if _ret_weight_mk > 0:
                            _old_total = float(_orig_wco.total_weight_main_karat or 0.0)
                            _old_executed = float(_orig_wco.executed_weight_main_karat or 0.0)
                            _new_total = max(_old_total - _ret_weight_mk, 0.0)
                            _over_executed = max(_old_executed - _new_total, 0.0)

                            if _over_executed > 0.0001:
                                _remaining_to_reverse = _over_executed
                                _cogs_account = Account.query.filter_by(account_number='521').first()
                                _inv_fin_id = _get_inventory_account_by_karat(
                                    int(_orig_wco.main_karat or get_main_karat() or 21),
                                    kind=(_orig_wco.invoice.gold_type if _orig_wco.invoice else 'new'),
                                )
                                _past_executions = (
                                    WeightClosingExecution.query
                                    .filter_by(order_id=_orig_wco.id)
                                    .filter(WeightClosingExecution.weight_main_karat > 0)
                                    .order_by(WeightClosingExecution.created_at.desc(), WeightClosingExecution.id.desc())
                                    .all()
                                )
                                for _exec_row in _past_executions:
                                    if _remaining_to_reverse <= 0.0001:
                                        break
                                    _exec_weight = float(_exec_row.weight_main_karat or 0.0)
                                    if _exec_weight <= 0:
                                        continue
                                    _reverse_chunk = min(_exec_weight, _remaining_to_reverse)
                                    _exec_price = float(_exec_row.price_per_gram or 0.0)
                                    _reverse_24k = convert_from_main_karat(_reverse_chunk, 24)
                                    _reverse_cash = round(_reverse_24k * _exec_price, 2) if _exec_price else 0.0

                                    if _reverse_cash > 0 and _cogs_account and _inv_fin_id:
                                        create_dual_journal_entry(
                                            journal_entry_id=journal_entry.id,
                                            account_id=_inv_fin_id,
                                            cash_debit=_reverse_cash,
                                            description=(
                                                f"عكس تكلفة مبيعات مُسوّاة سابقاً - مرتجع بيع "
                                                f"(أمر {_orig_wco.order_number})"
                                            ),
                                        )
                                        create_dual_journal_entry(
                                            journal_entry_id=journal_entry.id,
                                            account_id=_cogs_account.id,
                                            cash_credit=_reverse_cash,
                                            description=(
                                                f"عكس تكلفة مبيعات مُسوّاة سابقاً - مرتجع بيع "
                                                f"(أمر {_orig_wco.order_number})"
                                            ),
                                        )

                                    db.session.add(WeightClosingExecution(
                                        order_id=_orig_wco.id,
                                        source_invoice_id=new_invoice.id,
                                        execution_type='sale_return_reversal',
                                        weight_main_karat=-_reverse_chunk,
                                        price_per_gram=_exec_price,
                                        journal_entry_id=journal_entry.id,
                                        notes=f'عكس جزئي لتنفيذ تسكير سابق بسبب مرتجع بيع #{new_invoice.id}',
                                    ))
                                    _remaining_to_reverse -= _reverse_chunk

                            _orig_wco.total_weight_main_karat = round(_new_total, 6)
                            _orig_wco.executed_weight_main_karat = round(max(_old_executed - _over_executed, 0.0), 6)
                            _orig_wco.remaining_weight_main_karat = round(
                                max(_orig_wco.total_weight_main_karat - _orig_wco.executed_weight_main_karat, 0.0), 6
                            )
                            if _orig_wco.total_weight_main_karat <= 0.001:
                                _orig_wco.status = 'cancelled'
                            elif _orig_wco.remaining_weight_main_karat <= 0.0001:
                                _orig_wco.status = 'closed'
                            elif _orig_wco.executed_weight_main_karat > 0:
                                _orig_wco.status = 'partially_closed'
                            else:
                                _orig_wco.status = 'open'

                            if _orig_wco.invoice:
                                _orig_wco.invoice.weight_closing_status = _orig_wco.status
                                _orig_wco.invoice.weight_closing_total_weight = _orig_wco.total_weight_main_karat
                                _orig_wco.invoice.weight_closing_executed_weight = _orig_wco.executed_weight_main_karat
                                _orig_wco.invoice.weight_closing_remaining_weight = _orig_wco.remaining_weight_main_karat

                        db.session.flush()
                except Exception as _wco_exc:
                    print(f"⚠️ WeightClosingOrder update for sale return failed: {_wco_exc}")

        elif invoice_type == 'مرتجع شراء':
            # 4. مرتجع شراء كسر (عكس الشراء من عميل)
            # من حـ/ العميل (أو الصندوق) [مدين]
            #     إلى حـ/ المخزون - كسر [دائن]
            
            # 🔥 استخدام الربط المحاسبي
            cash_acc_id = get_account_id_for_mapping('مرتجع شراء', 'cash')
            customers_acc_id = get_account_id_for_mapping('مرتجع شراء', 'customers')
            purchase_returns_acc_id = (
                get_account_id_for_mapping('مرتجع شراء', 'purchase_returns')
                or get_account_id_by_number('513')
                or get_account_id_by_number('512')
                or get_account_id_by_number('511')
            )

            # حسابات المخزون — يجب أن تكون مدركة لـ gold_type
            # (كسر → 1310، جديد → 1300) مثل سائر أنواع الفواتير الأخرى.
            inventory_acc_id = _resolve_inventory_account_id_for_invoice(invoice_type, gold_type)
            if not inventory_acc_id:
                # Fallback: scan generic karat mappings (legacy behaviour)
                for karat in ['18', '21', '22', '24']:
                    inv_acc_id = get_account_id_for_mapping('مرتجع شراء', f'inventory_{karat}k')
                    if inv_acc_id:
                        inventory_acc_id = inv_acc_id
                        break
            
            # VAT receivable to reverse
            vat_receivable_acc_id_ret = get_account_id_for_mapping('شراء من عميل', 'vat_receivable')
            _total_tax_purchase_ret = sum(
                _to_float(item_data.get('tax_amount', item_data.get('tax', 0.0)), 0.0)
                for item_data in data.get('items', [])
            )
            if _total_tax_purchase_ret < 0:
                _total_tax_purchase_ret = abs(_total_tax_purchase_ret)
            # Fallback: total_tax المخزّن في الفاتورة (يُحسب من karat_lines)
            if _total_tax_purchase_ret == 0.0:
                _total_tax_purchase_ret = float(new_invoice.total_tax or 0.0)
            # Second fallback: من الفاتورة الأصلية
            if _total_tax_purchase_ret == 0.0 and data.get('original_invoice_id'):
                try:
                    _orig_for_ptax = Invoice.query.get(int(data['original_invoice_id']))
                    if _orig_for_ptax:
                        _total_tax_purchase_ret = float(_orig_for_ptax.total_tax or 0.0)
                except Exception:
                    pass

            _net_purchase_ret = round(total_cash - _total_tax_purchase_ret, 2)

            # Line 1: مدين العميل/الصندوق (نقد فقط — الوزن عبر je_engine_v2 أدناه)
            acc_id = customers_acc_id or cash_acc_id or party_account.id
            create_dual_journal_entry(
                journal_entry_id=journal_entry.id,
                account_id=acc_id,
                cash_debit=total_cash,
                apply_golden_rule=False,
                description="استلام نقدي من مرتجع شراء"
            )

            # Line 2: دائن مردودات المشتريات (بدون ضريبة)
            _pr_credit_acc = purchase_returns_acc_id or inventory_acc_id
            if _pr_credit_acc:
                create_dual_journal_entry(
                    journal_entry_id=journal_entry.id,
                    account_id=_pr_credit_acc,
                    cash_credit=_net_purchase_ret,
                    apply_golden_rule=False,
                    description="مردودات مشتريات - مرتجع شراء"
                )

            # Line 2b: دائن ضريبة القيمة المضافة (عكس المدين في الشراء الأصلي)
            if _total_tax_purchase_ret > 0 and vat_receivable_acc_id_ret:
                create_dual_journal_entry(
                    journal_entry_id=journal_entry.id,
                    account_id=vat_receivable_acc_id_ret,
                    cash_credit=_total_tax_purchase_ret,
                    apply_golden_rule=False,
                    description="عكس ضريبة القيمة المضافة - مرتجع شراء"
                )

            # قيود الوزن (je_engine_v2): خروج وزن من المخزون + عودة وزن العميل
            import je_adapter as _je_adapter  # noqa: local import
            _inv_accounts_ret = {}
            if inventory_acc_id:
                for _k in ['18', '21', '22', '24']:
                    _inv_accounts_ret[_k] = inventory_acc_id
            # if scrap return, use scrap safe account for weight
            _is_scrap_ret = (
                str(getattr(new_invoice, 'gold_type', '') or '').strip().lower() == 'scrap'
                and scrap_purchase_gold_safe_account_id not in (None, 0)
            )
            if _is_scrap_ret:
                for _k in _inv_accounts_ret:
                    _inv_accounts_ret[_k] = int(scrap_purchase_gold_safe_account_id)

            _je_party_ret = (
                party_account
                or (Account.query.get(acc_id) if acc_id else None)
            )
            if _je_party_ret and _inv_accounts_ret:
                _pr_acc = purchase_returns_acc_id or get_account_id_by_number('513')
                _pr_weight_override = (
                    _je_adapter._resolve_weight_account_id(_pr_acc)
                    if _pr_acc else None
                )
                _je_adapter.weight_entries_for_party(
                    journal_entry_id=journal_entry.id,
                    gold_by_karat=gold_by_karat,
                    inventory_accounts=_inv_accounts_ret,
                    party_account_obj=_je_party_ret,
                    direction='sale',   # مرتجع شراء = ذهب يخرج من المخزون للعميل
                    customer_id=new_invoice.customer_id,
                    scrap_purchase_gold_safe_account_id=(
                        int(scrap_purchase_gold_safe_account_id)
                        if _is_scrap_ret else None
                    ),
                    party_weight_account_override=_pr_weight_override,
                )
        
        elif invoice_type == 'شراء':
            # 5. شراء (مورد)
            # السيناريو الجديد: المخزون يُثبت بالوزن والقيمة، المورد دائن بالذهب،
            # ويتم تسجيل التقييم النقدي على حساب جسر مستقل.

            # محاولة الحصول على حساب الجسر من الطلب أو إعدادات الربط
            bridge_acc_id = (
                data.get('bridge_account_id')
                or get_account_id_for_mapping('شراء', 'supplier_bridge')
            )

            if not bridge_acc_id:
                bridge_acc_id = (
                    get_account_id_for_mapping('شراء', 'suppliers')
                    or (party_account.id if party_account and not party_account.tracks_weight else None)
                    or (cash_account.id if cash_account else None)
                )

            if bridge_acc_id:
                operation_key = 'شراء'
                fallback_operation = None
                dual_entry_params = set(create_dual_journal_entry.__code__.co_varnames)

                def _mapping(account_type):
                    value = get_account_id_for_mapping(operation_key, account_type)
                    if value is None and fallback_operation:
                        value = get_account_id_for_mapping(fallback_operation, account_type)
                    return value

                def _normalize_karat(value):
                    try:
                        return str(int(round(float(value))))
                    except (TypeError, ValueError):
                        return None

                # حسابات أساسية
                vat_receivable_acc_id = _mapping('vat_receivable')

                # Root-fix: VAT receivable must exist if any VAT is present.
                # Correct account is 1400 (ضريبة مدفوعة على المشتريات).
                if not vat_receivable_acc_id:
                    try:
                        vat_acc = Account.query.filter_by(account_number='1400').first()
                        if not vat_acc:
                            vat_acc = Account.query.filter_by(account_number='1500').first()
                        vat_receivable_acc_id = vat_acc.id if vat_acc else None
                    except Exception:
                        vat_receivable_acc_id = vat_receivable_acc_id
                wage_mode = _get_manufacturing_wage_mode()
                wage_expense_acc_id = None
                wage_inventory_acc_id = None
                if wage_mode == 'inventory':
                    wage_inventory_acc_id = (
                        data.get('wage_inventory_account_id')
                        or _get_manufacturing_wage_inventory_account_id()
                        or _mapping('manufacturing_wage_inventory')
                        or _mapping('manufacturing_wage')
                    )
                if wage_mode != 'inventory' or not wage_inventory_acc_id:
                    wage_expense_acc_id = (
                        data.get('wage_expense_account_id')
                        or _mapping('manufacturing_wage')
                        or _mapping('manufacturing_wage_inventory')
                    )
                if wage_inventory_acc_id:
                    _ensure_weight_tracking_account(wage_inventory_acc_id)
                if wage_expense_acc_id:
                    _ensure_weight_tracking_account(wage_expense_acc_id)

                # بناء قاموس حسابات المخزون: دعم التوحيد (حساب واحد لكل العيارات)
                inventory_accounts = {}
                unified_inventory_acc_id = _resolve_inventory_account_id_for_invoice(invoice_type, gold_type)
                if unified_inventory_acc_id:
                    inventory_accounts = {k: unified_inventory_acc_id for k in ['18', '21', '22', '24']}
                else:
                    for karat in ['18', '21', '22', '24']:
                        acc_id = _mapping(f'inventory_{karat}k')
                        if acc_id:
                            inventory_accounts[karat] = acc_id

                # تحديد حساب المورد (يجب أن يتتبع الوزن دائماً)
                # Root-fix: supplier invoices must post to the supplier's own subledger account
                # (financial + memo), not to an aggregated control account.
                supplier_fin_account_id = None
                supplier_fin_account_obj = None
                supplier_memo_account_id = None

                if new_invoice.supplier_id and party_account:
                    supplier_fin_account_id = party_account.id
                    supplier_fin_account_obj = party_account
                    supplier_memo_account_id = party_account.memo_account_id

                # If somehow we still don't have supplier accounts, fall back to mapping (control account)
                # but keep weight postings on memo only when available.
                if not supplier_fin_account_id:
                    fallback_id = _mapping('suppliers') or _mapping('suppliers_weight')
                    if fallback_id:
                        supplier_fin_account_id = fallback_id
                        supplier_fin_account_obj = Account.query.get(fallback_id)
                        supplier_memo_account_id = getattr(supplier_fin_account_obj, 'memo_account_id', None)

                if supplier_memo_account_id:
                    _ensure_weight_tracking_account(supplier_memo_account_id)

                # تجميع أوزان المورد (يمكن تمريرها من الواجهة، وإلا نستخدم أوزان الأصناف)
                supplier_gold_lines = data.get('supplier_gold_lines') or data.get('supplier_gold_weights')
                supplier_gold_by_karat = {}

                if isinstance(supplier_gold_lines, list):
                    for line in supplier_gold_lines:
                        karat_key = _normalize_karat(line.get('karat'))
                        weight = _to_float(line.get('weight', 0), 0.0)
                        if not karat_key or weight <= 0:
                            continue
                        supplier_gold_by_karat[karat_key] = supplier_gold_by_karat.get(karat_key, 0.0) + weight
                elif isinstance(supplier_gold_lines, dict):
                    for karat, weight in supplier_gold_lines.items():
                        weight_val = _to_float(weight, 0.0)
                        if weight_val <= 0:
                            continue
                        karat_key = _normalize_karat(karat)
                        if not karat_key:
                            continue
                        supplier_gold_by_karat[karat_key] = supplier_gold_by_karat.get(karat_key, 0.0) + weight_val

                if not supplier_gold_by_karat:
                    # استخدام الأوزان الفعلية من karat_lines
                    supplier_gold_by_karat = {k: v for k, v in gold_by_karat.items() if v > 0}

                # حفظ إجمالي الذهب (عيار رئيسي) في الفاتورة للرجوع إليه لاحقاً
                supplier_gold_main = sum(
                    convert_to_main_karat(weight, int(round(float(karat))))
                    for karat, weight in supplier_gold_by_karat.items()
                )
                new_invoice.payment_gold_weight = round(supplier_gold_main, 3)
                new_invoice.payment_gold_karat = get_main_karat()

                # قراءة القيم النقدية من الطلب أو حسابها
                gold_tax_total = _to_float(data.get('gold_tax_total', 0), 0.0)
                wage_tax_total = _to_float(data.get('wage_tax_total', 0), 0.0)
                total_vat_source = (
                    data.get('vat_receivable_cash')
                    or data.get('total_tax')
                    or (gold_tax_total + wage_tax_total)
                    or new_invoice.total_tax
                    or 0
                )
                total_vat = _to_float(total_vat_source, 0.0)
                wage_cash = _to_float(
                    data.get('manufacturing_wage_cash')
                    or data.get('wage_cash')
                    or data.get('total_wage')
                    or data.get('wage_subtotal')
                    or 0
                , 0.0)

                valuation_cash_total = data.get('valuation_cash_total')
                if valuation_cash_total is None and isinstance(data.get('valuation'), dict):
                    valuation_cash_total = data['valuation'].get('cash_total')

                valuation_cash_total = _to_float(valuation_cash_total, None) if valuation_cash_total is not None else None
                if valuation_cash_total is None:
                    valuation_cash_total = _to_float(data.get('gold_subtotal', 0), None)
                if valuation_cash_total is None:
                    valuation_cash_total = new_invoice.total - wage_cash - total_vat
                valuation_cash_total = max(round(valuation_cash_total, 2), 0)

                # توزيع الوزن الخاص بالتقييم (يمكن أن يختلف عن الوزن الفعلي إن وجد)
                valuation_weights = {}
                raw_valuation_weights = None
                if isinstance(data.get('valuation_gold_weights'), dict):
                    raw_valuation_weights = data.get('valuation_gold_weights')
                elif isinstance(data.get('valuation'), dict) and isinstance(data['valuation'].get('weight_by_karat'), dict):
                    raw_valuation_weights = data['valuation'].get('weight_by_karat')

                if raw_valuation_weights:
                    for karat, weight in raw_valuation_weights.items():
                        weight_val = _to_float(weight, 0.0)
                        if weight_val <= 0:
                            continue
                        karat_key = _normalize_karat(karat)
                        if not karat_key:
                            continue
                        valuation_weights[karat_key] = weight_val
                else:
                    valuation_weights = {k: v for k, v in gold_by_karat.items() if v > 0}

                # إجمالي الوزن المستخدم للتوزيع النقدي
                total_weight_for_allocation = sum(
                    weight for karat, weight in valuation_weights.items()
                    if weight > 0 and str(karat) in inventory_accounts
                )

                cash_debit_booked = 0.0

                # 🆕 محاولة استخراج التوزيع النقدي الفعلي من بيانات الفاتورة
                # هذا يدعم: خصومات، تفاوت سعر حسب العيار، أسعار مخصصة
                explicit_cash_by_karat = {}
                
                # 1. التحقق من وجود توزيع نقدي صريح في البيانات
                if isinstance(data.get('cash_allocation_by_karat'), dict):
                    explicit_cash_by_karat = data['cash_allocation_by_karat']
                
                # 2. حساب التوزيع من سطور الفاتورة إن وجدت
                elif data.get('items') and isinstance(data['items'], list):
                    for item_data in data['items']:
                        item_karat = _normalize_karat(item_data.get('karat'))
                        if not item_karat:
                            continue
                        
                        # الحصول على القيمة النقدية الفعلية للصنف
                        item_cash_value = _to_float(
                            item_data.get('net') or 
                            item_data.get('net_price') or
                            item_data.get('selling_price', 0), 
                            0.0
                        )
                        
                        # طرح الضريبة والخصم للحصول على قيمة الذهب فقط
                        item_tax = _to_float(item_data.get('tax_amount', 0), 0.0)
                        item_discount = _to_float(item_data.get('discount_amount', 0), 0.0)
                        item_wage = _to_float(item_data.get('wage', 0), 0.0)
                        
                        # القيمة النقدية للذهب = السعر - الضريبة - الخصم - الأجور
                        gold_cash = item_cash_value - item_tax - item_discount
                        
                        if gold_cash > 0:
                            explicit_cash_by_karat[item_karat] = (
                                explicit_cash_by_karat.get(item_karat, 0.0) + gold_cash
                            )

                # --- 1) إثبات المخزون (نقد + وزن لكل عيار) ---
                # 🆕 تخزين الأوزان الفعلية للقيود الوزنية (من karat_lines فقط، بدون المصنعية)
                actual_gold_weights_for_memo = {}
                if karat_lines_data and isinstance(karat_lines_data, list):
                    for line_data in karat_lines_data:
                        k = _normalize_karat(line_data.get('karat'))
                        w = _to_float(line_data.get('weight_grams', 0), 0.0)
                        if k and w > 0:
                            actual_gold_weights_for_memo[k] = actual_gold_weights_for_memo.get(k, 0.0) + w
                
                # Weight source used for posting physical gold purchase lines.
                # Some clients do not send `karat_lines` for supplier purchases; in that case,
                # fall back to supplier_gold_by_karat derived from items/request.
                physical_gold_weights_for_posting = dict(actual_gold_weights_for_memo or {})
                if not physical_gold_weights_for_posting:
                    physical_gold_weights_for_posting = dict(supplier_gold_by_karat or {})

                # Track which karats had an inventory weight debit posted.
                posted_weight_debits = set()
                
                if valuation_cash_total > 0 or total_weight_for_allocation > 0:
                    positive_karats = [k for k in valuation_weights if k in inventory_accounts and valuation_weights[k] > 0]

                    # If we have explicit cash allocation by karat, normalize it to ALWAYS sum
                    # to valuation_cash_total to avoid JE cash imbalance.
                    cash_shares_by_karat = None
                    if explicit_cash_by_karat and positive_karats:
                        try:
                            cash_shares_by_karat = {}
                            for k in positive_karats:
                                if k in explicit_cash_by_karat:
                                    cash_shares_by_karat[k] = round(_to_float(explicit_cash_by_karat.get(k), 0.0), 2)

                            explicit_sum = round(sum(cash_shares_by_karat.values()), 2)
                            diff = round(valuation_cash_total - explicit_sum, 2)
                            if abs(diff) > 0.01:
                                # Absorb the remainder into the largest-weight karat to keep totals consistent.
                                adjust_karat = max(
                                    positive_karats,
                                    key=lambda kk: float(valuation_weights.get(kk, 0.0) or 0.0),
                                )
                                cash_shares_by_karat[adjust_karat] = round(
                                    float(cash_shares_by_karat.get(adjust_karat, 0.0) or 0.0) + diff,
                                    2,
                                )
                                # If adjustment leads to a negative share, fall back to proportional allocation.
                                if cash_shares_by_karat[adjust_karat] < 0:
                                    cash_shares_by_karat = None
                        except Exception:
                            cash_shares_by_karat = None

                    remaining_cash = valuation_cash_total

                    for index, karat in enumerate(positive_karats):
                        weight_value = valuation_weights[karat]
                        inv_account_id = inventory_accounts.get(karat)
                        if not inv_account_id:
                            continue

                        # 🆕 استخدام التوزيع النقدي الصريح إن وجد (بعد التطبيع)، وإلا التوزيع النسبي
                        if cash_shares_by_karat is not None:
                            cash_share = round(_to_float(cash_shares_by_karat.get(karat, 0.0), 0.0), 2)
                        elif total_weight_for_allocation > 0 and index < len(positive_karats) - 1:
                            # التوزيع النسبي التقليدي (fallback)
                            cash_share = round(valuation_cash_total * (weight_value / total_weight_for_allocation), 2)
                            remaining_cash = round(remaining_cash - cash_share, 2)
                        else:
                            # آخر عيار يأخذ الباقي لتجنب فروقات التقريب
                            cash_share = max(round(remaining_cash, 2), 0)
                            remaining_cash = 0

                        # المخزون: وزن فقط (بدون قيد نقدي)
                        # قيمة الذهب مرجعية (للضريبة/التقارير) وليست التزاماً على المورد
                        # الوزن يُثبَّت في حساب المذكرة أدناه
                        
                        # 🆕 القيد الوزني: استخدام الوزن الفعلي من karat_lines (بدون المصنعية)
                        actual_weight_for_karat = physical_gold_weights_for_posting.get(karat, 0.0)
                        if actual_weight_for_karat > 0:
                            # حاول استخدام حساب مذكرة مرتبط بحساب المخزون المالي
                            # وإذا لم يوجد، استخدم fallback ثم الحساب المالي نفسه لمنع عدم توازن الوزن.
                            weight_inventory_memo_acc_id = None
                            try:
                                inv_acc_obj = Account.query.get(inv_account_id)
                                if inv_acc_obj and inv_acc_obj.memo_account_id:
                                    weight_inventory_memo_acc_id = inv_acc_obj.memo_account_id
                            except Exception:
                                weight_inventory_memo_acc_id = None

                            # fallback على الحساب المذكرة الافتراضي 7521
                            if not weight_inventory_memo_acc_id:
                                weight_inventory_memo_acc_id = get_account_id_by_number('7521')

                            # لا نسقط على الحساب المالي — الأوزان يجب أن تذهب لحسابات مذكرة فقط.
                            if weight_inventory_memo_acc_id:
                                create_dual_journal_entry(
                                    journal_entry_id=journal_entry.id,
                                    account_id=weight_inventory_memo_acc_id,
                                    **_weight_kwargs_for_karat(karat, round(actual_weight_for_karat, 3), 'debit'),
                                    exclude_from_ledger=True,  # سطر مذكرة مخزون ليس التزاماً على المورد
                                    description=f"شراء وزني (مورد) - عيار {karat}"
                                )
                                posted_weight_debits.add(str(karat))

                        cash_debit_booked = round(cash_debit_booked + max(cash_share, 0), 2)

                    # في حال لم يُسجَّل أي سطر (لعدم وجود أوزان)، ننشئ سطر نقدي واحد للمخزون
                    if not positive_karats and valuation_cash_total > 0 and inventory_accounts:
                        fallback_account_id = next(iter(inventory_accounts.values()))
                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=fallback_account_id,
                            cash_debit=valuation_cash_total,
                            apply_golden_rule=False,
                            exclude_from_ledger=True,
                            description="إثبات مخزون شراء (مورد) (بدون توزيع عيارات)"
                        )
                        cash_debit_booked = round(cash_debit_booked + valuation_cash_total, 2)

                # --- Safety net: ensure physical weight is always balanced ---
                # If a karat weight exists but inventory mapping was missing, we still need a debit
                # to balance the supplier payable weight credit.
                for karat_key, weight_val in (physical_gold_weights_for_posting or {}).items():
                    karat_str = str(karat_key)
                    if weight_val <= 0 or karat_str in posted_weight_debits:
                        continue

                    # Prefer the memo account linked to the mapped inventory account (if any).
                    inv_for_karat = inventory_accounts.get(karat_str)
                    weight_target_acc_id = None

                    if inv_for_karat:
                        try:
                            inv_acc_obj = Account.query.get(inv_for_karat)
                            if inv_acc_obj and inv_acc_obj.memo_account_id:
                                weight_target_acc_id = inv_acc_obj.memo_account_id
                        except Exception:
                            weight_target_acc_id = None

                    # Fallback to a generic inventory memo account if present.
                    if not weight_target_acc_id:
                        weight_target_acc_id = get_account_id_by_number('7521')

                    # لا نسقط على الحساب المالي — الأوزان يجب أن تذهب لحسابات مذكرة فقط.
                    if not weight_target_acc_id:
                        print(f"⚠️ No memo account for weight safety-net posting (karat {karat_str}). Skipping.")

                    if weight_target_acc_id:
                        weight_kwargs = _weight_kwargs_for_karat(karat_str, round(float(weight_val), 3), 'debit')
                        if not weight_kwargs:
                            continue

                        created_line = create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=weight_target_acc_id,
                            **weight_kwargs,
                            exclude_from_ledger=True,
                            description=f"شراء وزني من مورد - عيار {karat_str} (fallback)"
                        )

                        # Extra safety: if for any reason the helper didn't set the expected weight field,
                        # insert a direct weight-only line to avoid blocking invoice save.
                        expected_field = f"debit_{karat_str}k"
                        expected_value = float(weight_val)
                        if not created_line or (getattr(created_line, expected_field, 0.0) or 0.0) <= 0.0:
                            try:
                                from models import JournalEntryLine
                                manual_line = JournalEntryLine(
                                    journal_entry_id=journal_entry.id,
                                    account_id=weight_target_acc_id,
                                    cash_debit=0.0,
                                    cash_credit=0.0,
                                    description=f"شراء وزني من مورد - عيار {karat_str} (manual fallback)",
                                )
                                setattr(manual_line, expected_field, round(expected_value, 3))
                                db.session.add(manual_line)
                                db.session.flush()
                            except Exception as manual_exc:
                                print(f"⚠️ Weight safety-net manual insert failed: {manual_exc}")
                        posted_weight_debits.add(karat_str)

                # --- 2) أجور المصنعية → مخزون أجور المصنعية ---
                # 🆕 النظام الجديد: فصل المصنعية في حساب مستقل
                wage_inventory_account_id = (
                    _get_manufacturing_wage_inventory_account_id()
                    or get_account_id_by_number('1320')
                    or get_account_id_by_number('1350')
                )
                
                if wage_cash > 0:
                    if not wage_inventory_account_id:
                        return jsonify({
                            'error': 'حساب مخزون أجور المصنعية غير موجود. يرجى إنشاؤه أولاً أو ضبط mapping (manufacturing_wage_inventory).'
                        }), 400
                    
                    # إضافة المصنعية لحساب مخزون المصنعية
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=wage_inventory_account_id,
                        cash_debit=round(wage_cash, 2),
                        apply_golden_rule=False,
                        exclude_from_ledger=True,  # المصنعية تُعرض على المورد كالتزام منفصل (نوسمها على سطر المورد فقط)
                        description="إضافة أجور مصنعية للمخزون - شراء (مورد)"
                    )
                    cash_debit_booked = round(cash_debit_booked + wage_cash, 2)

                # --- 3) ضريبة القيمة المضافة ---
                # ملاحظة: ضريبة الذهب تُضاف لقيمة المخزون، وضريبة الأجور تُسجل منفصلة
                # لذا نسجل فقط ضريبة الأجور كقيد مستقل
                if (wage_tax_total > 0 or gold_tax_total > 0) and not vat_receivable_acc_id:
                    return jsonify({
                        'error': 'vat_receivable_account_missing',
                        'message': 'حساب ضريبة القيمة المضافة (مدفوعة) غير موجود. الرجاء إنشاء الحساب رقم 1500 أو ضبط mapping (شراء -> vat_receivable).',
                    }), 400

                if wage_tax_total > 0 and vat_receivable_acc_id:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=vat_receivable_acc_id,
                        cash_debit=round(wage_tax_total, 2),
                        apply_golden_rule=False,
                        exclude_from_ledger=True,
                        description="ضريبة على أجور المصنعية - مشتريات من مورد"
                    )
                    cash_debit_booked = round(cash_debit_booked + wage_tax_total, 2)
                
                # إذا كانت هناك ضريبة على الذهب، تُضاف للمخزون (مدرجة ضمن valuation_cash_total)
                if gold_tax_total > 0 and vat_receivable_acc_id:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=vat_receivable_acc_id,
                        cash_debit=round(gold_tax_total, 2),
                        apply_golden_rule=False,
                        exclude_from_ledger=True,
                        description="ضريبة على قيمة الذهب - مشتريات من مورد"
                    )
                    cash_debit_booked = round(cash_debit_booked + gold_tax_total, 2)

                # --- 4) فصل الالتزامات ---
                # A) ضريبة الذهب فقط (إذا كانت الضريبة على كامل قيمة الذهب - نادر)
                # قيمة الذهب نفسها (valuation_cash_total) ليست التزاماً نقدياً على المورد:
                # الذهب يتم عبر مقايضة (يتسلم خاماً ويسلم مشغولاً) وليس بيعاً نقدياً.
                valuation_bridge_cash = round(gold_tax_total, 2)

                # Supplier-specific wage settlement type:
                # - cash: wage is a SAR payable on supplier
                # - gold: wage is converted to gold weight (main karat) payable on supplier,
                #         while the cash-equivalent wage cost is balanced via the bridge
                supplier_obj = None
                try:
                    supplier_obj = Supplier.query.get(new_invoice.supplier_id) if new_invoice.supplier_id else None
                except Exception:
                    supplier_obj = None

                supplier_wage_type = (
                    (supplier_obj.default_wage_type if supplier_obj else None) or 'cash'
                )
                if not isinstance(supplier_wage_type, str):
                    supplier_wage_type = str(supplier_wage_type)
                supplier_wage_type = supplier_wage_type.strip().lower()
                if supplier_wage_type not in ('cash', 'gold'):
                    supplier_wage_type = 'cash'

                wage_gold_weight_main = 0.0
                wage_cash_liability = wage_cash
                wage_gold_weight_field = None
                if supplier_wage_type == 'gold' and wage_cash > 0:
                    try:
                        price_snapshot = get_current_gold_price()
                        price_main = _to_float(price_snapshot.get('price_per_gram_main_karat'), 0.0)
                        if price_main > 0:
                            main_karat = get_main_karat() or 21
                            candidate_field = f"weight_{int(round(float(main_karat)))}k_credit"
                            if candidate_field in dual_entry_params:
                                wage_gold_weight_main = round(wage_cash / price_main, 3)
                                wage_gold_weight_field = candidate_field
                                wage_cash_liability = 0.0
                                valuation_bridge_cash = round(valuation_bridge_cash + wage_cash, 2)
                            else:
                                supplier_wage_type = 'cash'
                    except Exception as exc:
                        print(f"⚠️ Failed to convert wage cash to gold weight: {exc}")
                        wage_gold_weight_main = 0.0
                        wage_cash_liability = wage_cash
                if valuation_bridge_cash > 0:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=supplier_fin_account_id,
                        cash_credit=valuation_bridge_cash,
                        apply_golden_rule=False,
                        description="التزام المورد - ضريبة قيمة الذهب"
                    )

                # B) الأجور + ضريبة الأجور:
                # - Wage VAT is always cash.
                # - Wage itself may be cash payable or gold payable depending on supplier.
                wage_payable_cash = round(wage_cash_liability + wage_tax_total, 2)

                # --- 5) المورد دائن بالذهب (حسب العيارات) ---
                # Weight postings are always on supplier memo account.

                if supplier_gold_by_karat and not supplier_memo_account_id:
                    return jsonify({
                        'error': 'supplier_memo_account_missing',
                        'message': 'حساب المورد الوزني (memo) غير موجود أو غير مربوط. الرجاء إعادة ربط حساب المورد بوجود حساب مذكرة وزني.',
                    }), 400

                if supplier_memo_account_id and supplier_gold_by_karat:
                    supplier_weight_kwargs = {
                        f'weight_{karat}k_credit': round(weight, 3)
                        for karat, weight in physical_gold_weights_for_posting.items()  # ← أوزان الذهب الفعلية
                        if weight > 0 and f'weight_{karat}k_credit' in dual_entry_params
                    }

                    # إن لم تُطابق أسماء الوسائط (عيار غير مدعوم)، نحاول تحويله إلى العيار الرئيسي
                    unsupported_karats = [
                        karat for karat in physical_gold_weights_for_posting  # ← أوزان الذهب الفعلية
                        if f'weight_{karat}k_credit' not in dual_entry_params
                    ]

                    additional_21k = 0.0
                    for karat in unsupported_karats:
                        weight = physical_gold_weights_for_posting.get(karat, 0)  # ← أوزان الذهب الفعلية
                        additional_21k += convert_to_main_karat(weight, int(round(float(karat))))

                    if additional_21k > 0:
                        supplier_weight_kwargs['weight_21k_credit'] = round(
                            supplier_weight_kwargs.get('weight_21k_credit', 0.0) + additional_21k,
                            3
                        )

                    if supplier_weight_kwargs:
                        # سطر التزام المورد (ذهب)
                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=supplier_memo_account_id,
                            apply_golden_rule=False,
                            description="التزام المورد - ذهب (وزن)",
                            **supplier_weight_kwargs,
                        )

                # Supplier wage as gold (main karat)
                if supplier_memo_account_id and wage_gold_weight_main > 0 and wage_gold_weight_field:
                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=supplier_memo_account_id,
                            apply_golden_rule=False,
                            description="التزام المورد - أجور مصنعية (ذهب)",
                        **{wage_gold_weight_field: round(wage_gold_weight_main, 3)},
                        )
                        try:
                            new_invoice.payment_gold_weight = round(
                                _to_float(new_invoice.payment_gold_weight, 0.0) + wage_gold_weight_main,
                                3,
                            )
                        except Exception:
                            pass

                # سطر التزام المورد (نقد) للأجور + ضريبة الأجور
                if supplier_fin_account_id and wage_payable_cash > 0:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=supplier_fin_account_id,
                        cash_credit=wage_payable_cash,
                        apply_golden_rule=False,
                        description="التزام المورد - أجور مصنعية + ضريبة الأجور",
                    )

                # --- Gold settlements: reflect in journal lines for statements ---
                # `gold_settlements[]` previously created only SafeBoxTransaction rows, which do not show
                # up in supplier statements (statements are derived from JournalEntryLine).
                # We also book weight-only lines:
                # - Supplier memo (weight) side: payable decreases when gold goes out (debit)
                # - Gold safe (weight) side: safe gold decreases when gold goes out (credit)
                try:
                    settlement_direction = _direction_for_invoice_type(new_invoice.invoice_type)
                except Exception:
                    settlement_direction = 'out'

                supplier_side = 'debit' if settlement_direction == 'out' else 'credit'
                safe_side = 'credit' if settlement_direction == 'out' else 'debit'

                if supplier_memo_account_id and resolved_gold_settlements_lines and (not gold_settlement_voucher_created):
                    aggregated = {}
                    for row in resolved_gold_settlements_lines:
                        try:
                            safe_id = int(row.get('safe_box_id'))
                            karat_int = int(row.get('karat'))
                            weight_val = float(row.get('weight') or 0.0)
                        except Exception:
                            continue
                        if weight_val <= 0 or karat_int not in (18, 21, 22, 24):
                            continue
                        aggregated[(safe_id, karat_int)] = aggregated.get((safe_id, karat_int), 0.0) + weight_val

                    for (safe_id, karat_int), weight_val in aggregated.items():
                        if weight_val <= 0:
                            continue

                        safe_box = None
                        safe_name = None
                        safe_weight_account_id = None

                        try:
                            safe_box = SafeBox.query.get(int(safe_id))
                            safe_name = getattr(safe_box, 'name', None)
                        except Exception:
                            safe_box = None

                        try:
                            safe_account_id = int(getattr(safe_box, 'account_id', None) or 0)
                        except Exception:
                            safe_account_id = 0

                        if safe_account_id:
                            try:
                                safe_acc = Account.query.get(safe_account_id)
                            except Exception:
                                safe_acc = None
                            if safe_acc and getattr(safe_acc, 'memo_account_id', None):
                                safe_weight_account_id = safe_acc.memo_account_id
                            else:
                                safe_weight_account_id = safe_account_id

                        if not safe_weight_account_id:
                            continue

                        _ensure_weight_tracking_account(safe_weight_account_id)

                        supplier_kwargs = _weight_kwargs_for_karat(karat_int, round(float(weight_val), 3), supplier_side)
                        safe_kwargs = _weight_kwargs_for_karat(karat_int, round(float(weight_val), 3), safe_side)
                        if not supplier_kwargs or not safe_kwargs:
                            continue

                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=supplier_memo_account_id,
                            apply_golden_rule=False,
                            description=f"سداد ذهب للمورد{' - ' + safe_name if safe_name else ''} - عيار {karat_int}",
                            **supplier_kwargs,
                        )

                        create_dual_journal_entry(
                            journal_entry_id=journal_entry.id,
                            account_id=safe_weight_account_id,
                            apply_golden_rule=False,
                            exclude_from_ledger=True,
                            description=f"سداد ذهب للمورد{' - ' + supplier.name if supplier else ''} - عيار {karat_int}",
                            **safe_kwargs,
                        )
                
                # 🆕 التحقق من توازن حساب الجسر بعد الفاتورة
                db.session.flush()  # تطبيق التغييرات قبل التحقق
                bridge_validation = validate_bridge_account_balance(bridge_acc_id, tolerance=0.01)
                
                if not bridge_validation['is_balanced']:
                    # تسجيل تحذير في السجل
                    print(f"⚠️ BRIDGE ACCOUNT IMBALANCE DETECTED:")
                    print(f"   Invoice ID: {new_invoice.id}")
                    print(f"   Invoice Type: {invoice_type}")
                    print(f"   Bridge Balance: {bridge_validation['bridge_balance']} SAR")
                    print(f"   Warning: {bridge_validation['warning']}")
                    
                    # يمكن إضافة تنبيه للمستخدم أو إرسال إشعار للمدير
                    # لكن لا نوقف العملية لأنها قد تكون بسبب فواصل عشرية

            else:
                return jsonify({
                    'error': 'لم يتم تهيئة حساب الجسر لمشتريات الموردين، يرجى ضبط mapping "supplier_bridge" أو حساب المورد النقدي.'
                }), 400
        
        elif invoice_type == 'مرتجع شراء (مورد)':
            # 6. مرتجع شراء (مورد) (عكس الشراء)
            # الهيكل الصحيح: يعكس فاتورة الشراء بالضبط
            #   مدين: حساب الجسر (قيمة الذهب النقدية - exclude_from_ledger كما في الشراء)
            #   مدين: حساب المورد المالي (أجور مصنعية فقط إن وجدت)
            #   مدين: حساب مذكرة المورد الوزني (وزن الذهب)
            #   دائن: حساب المخزون النقدي (exclude_from_ledger)
            #   دائن: حساب مذكرة المخزون الوزني (exclude_from_ledger)

            # 🔥 استخدام الربط المحاسبي (نفس إعدادات "شراء")
            cash_acc_id = get_account_id_for_mapping('شراء', 'cash')
            suppliers_acc_id = get_account_id_for_mapping('شراء', 'suppliers')

            # حساب الجسر (نفس منطق الشراء الأصلي)
            bridge_acc_id = (
                data.get('bridge_account_id')
                or get_account_id_for_mapping('شراء', 'supplier_bridge')
                or get_account_id_for_mapping('شراء', 'suppliers')
                or cash_acc_id
            )

            # Prefer posting to the supplier's own subledger account (root-fix).
            supplier_fin_account_id = None
            try:
                if getattr(new_invoice, 'supplier_id', None) and party_account:
                    supplier_fin_account_id = int(party_account.id)
            except Exception:
                supplier_fin_account_id = None

            # حسابات المخزون
            inventory_acc_id = None
            try:
                unified_inventory_acc_id = _resolve_inventory_account_id_for_invoice('شراء', gold_type)
            except Exception:
                unified_inventory_acc_id = None
            if unified_inventory_acc_id:
                inventory_acc_id = unified_inventory_acc_id
            else:
                for karat in ['18', '21', '22', '24']:
                    inv_acc_id = get_account_id_for_mapping('شراء', f'inventory_{karat}k')
                    if inv_acc_id:
                        inventory_acc_id = inv_acc_id
                        break

            if not inventory_acc_id:
                db.session.rollback()
                return jsonify({
                    'error': 'account_mapping_missing',
                    'message': 'لم يتم العثور على حساب مخزون لمرتجع شراء (مورد). الرجاء ضبط Accounting Mapping (inventory_XXk) أو تفعيل المخزون الموحد.',
                    'missing': [{'mapping': 'inventory_*', 'operation_type': 'شراء'}],
                }), 400

            acc_id = supplier_fin_account_id or (party_account.id if party_account else None) or suppliers_acc_id or cash_acc_id

            # حساب مذكرة المورد الوزني
            supplier_weight_acc_id = memo_party_account.id if memo_party_account else acc_id

            # حساب مذكرة المخزون الوزني
            weight_inventory_acc_id = None
            try:
                inv_acc_obj = Account.query.get(inventory_acc_id)
                if inv_acc_obj and inv_acc_obj.memo_account_id:
                    weight_inventory_acc_id = inv_acc_obj.memo_account_id
            except Exception:
                weight_inventory_acc_id = None
            if not weight_inventory_acc_id:
                weight_inventory_acc_id = get_account_id_by_number('7521')
            if not weight_inventory_acc_id:
                weight_inventory_acc_id = inventory_acc_id  # last resort fallback

            has_gold_weight = any(float(v or 0) > 0 for v in gold_by_karat.values())

            # استخراج جزء الأجور المصنعية (إن وجد) - يذهب لحساب المورد المالي مباشرة
            wage_cash = _to_float(
                data.get('manufacturing_wage_cash') or data.get('wage_cash')
                or data.get('total_wage') or data.get('wage_subtotal') or 0,
                0.0
            )
            # في المرتجع يجب عكس ضريبة الأجور من حساب المورد تماماً كما أُضيفت في الشراء
            _ret_wage_tax = _to_float(
                data.get('wage_tax_total') or data.get('wage_tax') or 0,
                0.0
            )
            _ret_gold_tax = _to_float(
                data.get('gold_tax_total') or data.get('gold_tax') or 0,
                0.0
            )
            # مبلغ الذهب = قيمة الذهب + ضريبة الذهب فقط (لا تشمل أجور ولا ضريبة الأجور)
            gold_subtt = _to_float(data.get('gold_subtotal') or 0, 0.0)
            if gold_subtt > 0:
                gold_value_cash = max(round(gold_subtt + _ret_gold_tax, 2), 0.0)
            else:
                gold_value_cash = max(round(total_cash - wage_cash - _ret_wage_tax, 2), 0.0)

            # Line 1: مدين حساب الجسر (قيمة الذهب النقدية - كما في الشراء)
            # يُعكس الدائن على الجسر من فاتورة الشراء الأصلية
            if gold_value_cash > 0 and bridge_acc_id:
                create_dual_journal_entry(
                    journal_entry_id=journal_entry.id,
                    account_id=bridge_acc_id,
                    cash_debit=gold_value_cash,
                    apply_golden_rule=False,
                    exclude_from_ledger=True,
                    description="مرتجع شراء (مورد) - عكس جسر التقييم"
                )
            elif total_cash > 0 and not bridge_acc_id:
                # Fallback: لا يوجد حساب جسر → استخدم حساب المورد المالي
                create_dual_journal_entry(
                    journal_entry_id=journal_entry.id,
                    account_id=acc_id,
                    cash_debit=total_cash,
                    description="مرتجع شراء (مورد)"
                )

            # Line 1b: مدين حساب المورد المالي (أجور مصنعية + ضريبة الأجور - عكس كامل لسطر الشراء)
            _wage_with_tax = round(wage_cash + _ret_wage_tax, 2)
            if _wage_with_tax > 0:
                create_dual_journal_entry(
                    journal_entry_id=journal_entry.id,
                    account_id=acc_id,
                    cash_debit=_wage_with_tax,
                    description="مرتجع شراء (مورد) - رد أجور مصنعية وضريبتها"
                )

            # Line 2: مدين حساب مذكرة المورد الوزني (تخفيض التزام الذهب - وزن فقط)
            if has_gold_weight:
                vendor_return_weight_debit = _weight_kwargs_from_map(gold_by_karat, 'debit')
                if vendor_return_weight_debit:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=supplier_weight_acc_id,
                        apply_golden_rule=False,
                        **vendor_return_weight_debit,
                        description="مرتجع شراء (مورد) - تخفيض وزن ذهب المورد"
                    )

            # Line 3: دائن حساب المخزون النقدي (تقليص القيمة النقدية)
            create_dual_journal_entry(
                journal_entry_id=journal_entry.id,
                account_id=inventory_acc_id,
                cash_credit=total_cash,
                exclude_from_ledger=True,
                description="خصم من المخزون النقدي (مرتجع شراء مورد)"
            )

            # Line 4: دائن حساب مذكرة المخزون الوزني
            if has_gold_weight and weight_inventory_acc_id:
                vendor_return_weight_credit = _weight_kwargs_from_map(gold_by_karat, 'credit')
                if vendor_return_weight_credit:
                    create_dual_journal_entry(
                        journal_entry_id=journal_entry.id,
                        account_id=weight_inventory_acc_id,
                        apply_golden_rule=False,
                        **vendor_return_weight_credit,
                        exclude_from_ledger=True,
                        description="خصم وزني من المخزون (مرتجع شراء مورد)"
                    )

        # --- 6. Verify Dual Balance Before Commit ---
        db.session.flush()  # Ensure all entries are in DB before verification
        print(f"🔍 Verifying dual balance for journal entry #{journal_entry.id}...")
        balance_check = verify_dual_balance(journal_entry.id)
        print(f"Balance check result: {balance_check}")
        if not balance_check['balanced']:
            # محاولة موازنة فروقات الوزن الصغيرة تلقائياً (مثل فروقات التقريب)
            try:
                from models import JournalEntryLine

                weight_balances = balance_check.get('weight_balances') or {}
                imbalanced = [
                    (k, v) for k, v in weight_balances.items()
                    if abs(v) > 0.001
                ]

                # محاولة موازنة فروقات الوزن الصغيرة تلقائياً (مثل فروقات التقريب)
                # ملاحظة: حالات المقايضة/الأوزان متعددة العيارات قد تُنتج فروقات صغيرة عبر أكثر من عيار.
                AUTO_WEIGHT_TOLERANCE = 1.5  # grams
                try:
                    # For barter-linked scrap purchases (offset), valuation may differ materially
                    # from the current gold price used for cash↔weight conversion. In those cases
                    # we prefer not to fail the save; instead we allow a wider auto-rebalance.
                    inv_type_key = str(getattr(new_invoice, 'invoice_type', '') or '').strip()
                    gold_type_key = str(getattr(new_invoice, 'gold_type', '') or '').strip().lower()
                    settlement_key = str(getattr(new_invoice, 'settlement_method', '') or '').strip().lower()
                    has_barter_link = getattr(new_invoice, 'barter_sale_invoice_id', None) not in (None, '', False)
                    is_offset_like = settlement_key in ('offset', 'barter', 'trade', 'swap') or has_barter_link
                    if inv_type_key == 'شراء من عميل' and gold_type_key == 'scrap' and is_offset_like:
                        AUTO_WEIGHT_TOLERANCE = 10.0
                except Exception:
                    pass
                if (
                    abs(balance_check.get('cash_balance', 0.0)) <= 0.01
                    and imbalanced
                    and all(abs(diff) <= AUTO_WEIGHT_TOLERANCE for _, diff in imbalanced)
                ):
                    lines = (
                        db.session.query(JournalEntryLine)
                        .filter_by(journal_entry_id=journal_entry.id)
                        .order_by(JournalEntryLine.id.desc())
                        .all()
                    )

                    for karat_label, diff in imbalanced:  # diff = debit - credit
                        try:
                            karat_int = int(str(karat_label).replace('k', '').strip())
                        except Exception:
                            karat_int = 21

                        debit_field = f'debit_{karat_int}k'
                        credit_field = f'credit_{karat_int}k'

                        target_line = None
                        if diff > 0:
                            # debit > credit → نزيد credit
                            for line in lines:
                                if (getattr(line, credit_field, 0) or 0) > 0:
                                    target_line = line
                                    break
                        else:
                            # credit > debit → نزيد debit
                            for line in lines:
                                if (getattr(line, debit_field, 0) or 0) > 0:
                                    target_line = line
                                    break

                        if not target_line and lines:
                            target_line = lines[0]

                        if target_line:
                            if diff > 0:
                                setattr(
                                    target_line,
                                    credit_field,
                                    round((getattr(target_line, credit_field, 0) or 0) + diff, 3),
                                )
                            else:
                                setattr(
                                    target_line,
                                    debit_field,
                                    round((getattr(target_line, debit_field, 0) or 0) + abs(diff), 3),
                                )
                            db.session.add(target_line)

                    db.session.flush()

                    # إعادة التحقق بعد التصحيح
                    balance_check = verify_dual_balance(journal_entry.id)
                    print(f"Balance check after auto-weight-balance: {balance_check}")

                if not balance_check['balanced']:
                    # Capture a small preview of journal lines before rollback to make debugging easier.
                    try:
                        from models import JournalEntryLine

                        je_lines = (
                            db.session.query(JournalEntryLine)
                            .filter_by(journal_entry_id=journal_entry.id)
                            .order_by(JournalEntryLine.id.asc())
                            .all()
                        )
                        preview = []
                        for ln in je_lines[:30]:
                            acc = ln.account or db.session.query(Account).get(ln.account_id)
                            preview.append({
                                'line_id': ln.id,
                                'account_id': ln.account_id,
                                'account_number': getattr(acc, 'account_number', None),
                                'account_name': getattr(acc, 'name', None),
                                'description': getattr(ln, 'description', None),
                                'cash_debit': round(float(getattr(ln, 'cash_debit', 0.0) or 0.0), 2),
                                'cash_credit': round(float(getattr(ln, 'cash_credit', 0.0) or 0.0), 2),
                                'debit_18k': round(float(getattr(ln, 'debit_18k', 0.0) or 0.0), 3),
                                'credit_18k': round(float(getattr(ln, 'credit_18k', 0.0) or 0.0), 3),
                                'debit_21k': round(float(getattr(ln, 'debit_21k', 0.0) or 0.0), 3),
                                'credit_21k': round(float(getattr(ln, 'credit_21k', 0.0) or 0.0), 3),
                                'debit_22k': round(float(getattr(ln, 'debit_22k', 0.0) or 0.0), 3),
                                'credit_22k': round(float(getattr(ln, 'credit_22k', 0.0) or 0.0), 3),
                                'debit_24k': round(float(getattr(ln, 'debit_24k', 0.0) or 0.0), 3),
                                'credit_24k': round(float(getattr(ln, 'credit_24k', 0.0) or 0.0), 3),
                            })
                        balance_check = dict(balance_check or {})
                        balance_check['journal_lines_preview'] = preview
                        balance_check['posting_context'] = {
                            'invoice_type': str(data.get('type') or ''),
                            'total_cash': round(float(total_cash or 0.0), 2) if 'total_cash' in locals() else None,
                            'net_amount': round(float(net_amount or 0.0), 2) if 'net_amount' in locals() else None,
                            'commission_amount': round(float(commission_amount or 0.0), 2) if 'commission_amount' in locals() else None,
                            'commission_vat_total': round(float(commission_vat_total or 0.0), 2) if 'commission_vat_total' in locals() else None,
                            'payment_method_id': payment_method_id,
                            'payment_commission_timing': pm_commission_timing if 'pm_commission_timing' in locals() else None,
                            'safe_box_id': safe_box_id,
                            'payments_count': len(payments) if isinstance(payments, list) else None,
                        }
                    except Exception as preview_exc:
                        try:
                            balance_check = dict(balance_check or {})
                            balance_check['journal_lines_preview_error'] = str(preview_exc)
                        except Exception:
                            pass

                    db.session.rollback()
                    extra_parts = []
                    try:
                        ctx = (balance_check or {}).get('posting_context') if isinstance(balance_check, dict) else None
                        if isinstance(ctx, dict):
                            ctx_bits = []
                            for k in (
                                'payment_commission_timing',
                                'total_cash',
                                'net_amount',
                                'commission_amount',
                                'commission_vat_total',
                                'payment_method_id',
                                'safe_box_id',
                                'payments_count',
                            ):
                                if k in ctx and ctx.get(k) is not None:
                                    ctx_bits.append(f"{k}={ctx.get(k)}")
                            if ctx_bits:
                                extra_parts.append('ctx: ' + ', '.join(ctx_bits))

                        preview = (balance_check or {}).get('journal_lines_preview') if isinstance(balance_check, dict) else None
                        if isinstance(preview, list) and preview:
                            cash_lines = []
                            for ln in preview:
                                try:
                                    d = float((ln or {}).get('cash_debit') or 0.0)
                                    c = float((ln or {}).get('cash_credit') or 0.0)
                                except Exception:
                                    d, c = 0.0, 0.0
                                if abs(d) < 0.005 and abs(c) < 0.005:
                                    continue
                                acc_no = (ln or {}).get('account_number')
                                acc_nm = (ln or {}).get('account_name')
                                label = f"{acc_no} {acc_nm}".strip()
                                cash_lines.append(f"{label}: {d:.2f}/{c:.2f}")
                                if len(cash_lines) >= 8:
                                    break
                            if cash_lines:
                                extra_parts.append('lines: ' + ' | '.join(cash_lines))
                    except Exception:
                        pass

                    extra = (' | ' + ' | '.join(extra_parts)) if extra_parts else ''
                    error_msg = f"Journal entry is not balanced: {', '.join(balance_check['errors'])}{extra}"
                    print(f"❌ Balance Error: {error_msg}")
                    return jsonify({'error': error_msg, 'balance_details': balance_check}), 400
            except Exception as auto_exc:
                db.session.rollback()
                error_msg = f"Journal entry is not balanced: {', '.join(balance_check.get('errors') or [])}"
                print(f"❌ Balance Error (auto-balance failed): {auto_exc} :: {error_msg}")
                return jsonify({'error': error_msg, 'balance_details': balance_check}), 400

        # --- 7. Mark as Posted and Commit ---
        now = datetime.now()

        # Determine if invoice should be unposted:
        # 1) Approval required (below_cost / large_discount)
        # 2) Auto-post disabled in server settings (and not force_post)
        _auto_post_disabled = (not _posting_auto_post and not force_post)

        if approval_required or _auto_post_disabled:
            if approval_required:
                print("✅ Balance verified! Approval required; skipping posting/safebox effects...")
            else:
                print("✅ Balance verified! Auto-post disabled; saving as unposted...")

            new_invoice.is_posted = False
            # Keep posted_by as creator name, but do not set posted_at.
            if not new_invoice.posted_by:
                new_invoice.posted_by = posted_by_username or 'system'

            # Keep entry unposted; it will be posted when invoice is approved.
            journal_entry.is_posted = False
            if hasattr(journal_entry, 'posted_at'):
                journal_entry.posted_at = None
            if hasattr(journal_entry, 'posted_by'):
                journal_entry.posted_by = None

            # 🧾 Sync payment status for unposted invoices too (UI correctness)
            try:
                total_amount = float(new_invoice.total or 0.0)
                paid_amount = float(new_invoice.amount_paid or 0.0)
                barter_total_status = float(getattr(new_invoice, 'barter_total', 0.0) or 0.0)
                total_settled = paid_amount + barter_total_status
                eps = 0.01
                if total_amount <= eps:
                    # Edge-case: zero-total invoices; consider any settlement as paid.
                    new_invoice.status = 'paid' if total_settled > eps else 'unpaid'
                elif total_settled <= eps:
                    new_invoice.status = 'unpaid'
                elif total_settled >= total_amount - eps:
                    new_invoice.status = 'paid'
                else:
                    new_invoice.status = 'partially_paid'
            except Exception:
                pass

            # Persistent manager alert (only for approval-gated invoices, not auto-post-disabled).
            if approval_required:
              try:
                from models import SystemAlert

                reason_labels = {
                    'large_discount': 'خصم كبير',
                    'below_cost': 'بيع تحت التكلفة',
                    'above_live_price': 'شراء أعلى من السعر المباشر',
                }
                reasons_human = [reason_labels.get(r, r) for r in (approval_reasons or ([approval_reason] if approval_reason else []))]
                reasons_human = [r for r in reasons_human if r]

                message_parts = []
                if 'large_discount' in (approval_reasons or []):
                    try:
                        message_parts.append(
                            f"خصم كبير: {round(float(discount_pct or 0.0), 2)}% "
                            f"(الحد {float(large_discount_pct_threshold or 0.0)}%)"
                        )
                    except Exception:
                        message_parts.append("خصم كبير")

                if 'below_cost' in (approval_reasons or []):
                    try:
                        sale_ex_vat = float((below_cost_details or {}).get('effective_sale_cash_ex_vat', 0.0) or 0.0)
                        cost_cash = float((below_cost_details or {}).get('cost_cash', 0.0) or 0.0)
                        profit_est = float((below_cost_details or {}).get('profit_cash_estimate', 0.0) or 0.0)
                        message_parts.append(
                            f"بيع تحت التكلفة: صافي {round(sale_ex_vat, 2)} مقابل تكلفة {round(cost_cash, 2)} "
                            f"(فرق {round(profit_est, 2)})"
                        )
                    except Exception:
                        message_parts.append("بيع تحت التكلفة")

                if 'above_live_price' in (approval_reasons or []):
                    try:
                        items = (purchase_above_live_price_details or {}).get('items') or []
                        first_item = items[0] if items else {}
                        message_parts.append(
                            f"شراء أعلى من السعر المباشر: {first_item.get('name', 'صنف')} "
                            f"بسعر/جرام {first_item.get('paid_per_gram', 0)} مقابل مباشر {first_item.get('live_per_gram', 0)}"
                        )
                    except Exception:
                        message_parts.append("شراء أعلى من السعر المباشر")

                if not message_parts and reasons_human:
                    message_parts.append(" / ".join(reasons_human))

                alert_details = {
                    'invoice_type': str(invoice_type).strip(),
                    'approval_reason': approval_reason,
                    'approval_reasons': approval_reasons or ([] if not approval_reason else [approval_reason]),
                }

                # Include discount info when applicable
                if 'large_discount' in (approval_reasons or []):
                    alert_details.update({
                        'discount_total_cash': round(float(total_discount_cash), 2),
                        'gross_total_cash': round(float(total_gross_cash), 2),
                        'discount_pct': round(float(discount_pct or 0.0), 2),
                        'threshold_pct': float(large_discount_pct_threshold or 0.0),
                    })

                # Include below-cost info when applicable
                if 'below_cost' in (approval_reasons or []):
                    try:
                        alert_details['below_cost'] = below_cost_details
                    except Exception:
                        pass
                if 'above_live_price' in (approval_reasons or []):
                    try:
                        alert_details['above_live_price'] = purchase_above_live_price_details
                    except Exception:
                        pass

                alert = SystemAlert(
                    alert_type='invoice_approval',
                    severity='critical',
                    title='فاتورة تحتاج اعتماد قبل الترحيل',
                    message=" | ".join([p for p in message_parts if p]) or 'تحتاج اعتماد قبل الترحيل',
                    entity_type='Invoice',
                    entity_id=new_invoice.id,
                    entity_number=getattr(new_invoice, 'invoice_number', None),
                    details=json.dumps(alert_details, ensure_ascii=False),
                    created_by=posted_by_username or 'system',
                )
                db.session.add(alert)
              except Exception:
                pass

            # Audit: approval required (per-reason)
            if approval_required:
              try:
                from models import AuditLog

                audit_base = {
                    'invoice_type': str(invoice_type).strip(),
                    'approval_required': True,
                    'approval_reason': approval_reason,
                    'approval_reasons': approval_reasons or ([] if not approval_reason else [approval_reason]),
                }

                if 'large_discount' in (approval_reasons or []):
                    AuditLog.log_action(
                        user_name=new_invoice.posted_by or posted_by_username or 'system',
                        action='large_discount',
                        entity_type='Invoice',
                        entity_id=new_invoice.id,
                        entity_number=getattr(new_invoice, 'invoice_number', None),
                        details=json.dumps({
                            **audit_base,
                            'discount_total_cash': round(float(total_discount_cash), 2),
                            'gross_total_cash': round(float(total_gross_cash), 2),
                            'discount_pct': round(float(discount_pct or 0.0), 2),
                            'threshold_pct': float(large_discount_pct_threshold or 0.0),
                        }, ensure_ascii=False),
                        ip_address=request.remote_addr,
                        user_agent=request.headers.get('User-Agent'),
                        success=True,
                    )

                if 'below_cost' in (approval_reasons or []):
                    AuditLog.log_action(
                        user_name=new_invoice.posted_by or posted_by_username or 'system',
                        action='below_cost',
                        entity_type='Invoice',
                        entity_id=new_invoice.id,
                        entity_number=getattr(new_invoice, 'invoice_number', None),
                        details=json.dumps({
                            **audit_base,
                            'below_cost': below_cost_details,
                        }, ensure_ascii=False),
                        ip_address=request.remote_addr,
                        user_agent=request.headers.get('User-Agent'),
                        success=True,
                    )

                if 'above_live_price' in (approval_reasons or []):
                    AuditLog.log_action(
                        user_name=new_invoice.posted_by or posted_by_username or 'system',
                        action='above_live_price',
                        entity_type='Invoice',
                        entity_id=new_invoice.id,
                        entity_number=getattr(new_invoice, 'invoice_number', None),
                        details=json.dumps({
                            **audit_base,
                            'above_live_price': purchase_above_live_price_details,
                        }, ensure_ascii=False),
                        ip_address=request.remote_addr,
                        user_agent=request.headers.get('User-Agent'),
                        success=True,
                    )
              except Exception:
                pass

            db.session.commit()
            resp = new_invoice.to_dict()
            resp['approval_required'] = bool(approval_required)
            resp['auto_post_disabled'] = bool(_auto_post_disabled and not approval_required)
            resp['approval_reason'] = approval_reason
            resp['approval_reasons'] = approval_reasons or ([] if not approval_reason else [approval_reason])
            if 'below_cost' in (approval_reasons or []):
                resp['below_cost'] = below_cost_details
            if 'above_live_price' in (approval_reasons or []):
                resp['above_live_price'] = purchase_above_live_price_details
            resp['discount_pct'] = round(float(discount_pct or 0.0), 2) if discount_pct is not None else None
            resp['threshold_pct'] = float(large_discount_pct_threshold or 0.0)
            _pos_claims_confirmed = True
            if _pos_claims:
                try:
                    from services.commerce_availability import _confirm_pos_claims_best_effort
                    _confirm_pos_claims_best_effort(_pos_claims)
                except Exception:
                    pass
            return jsonify(resp), 201

        print(f"✅ Balance verified! Marking invoice and journal entry as posted...")
        new_invoice.is_posted = True
        if not new_invoice.posted_at:
            new_invoice.posted_at = now
        if not new_invoice.posted_by:
            new_invoice.posted_by = posted_by_username or 'system'

        # 🧾 Sync payment status from amount_paid vs total
        try:
            total_amount = float(new_invoice.total or 0.0)
            paid_amount = float(new_invoice.amount_paid or 0.0)
            barter_total_status = float(getattr(new_invoice, 'barter_total', 0.0) or 0.0)
            total_settled = paid_amount + barter_total_status
            eps = 0.01
            if total_settled <= eps:
                new_invoice.status = 'unpaid'
            elif total_settled >= total_amount - eps:
                new_invoice.status = 'paid'
            else:
                new_invoice.status = 'partially_paid'
        except Exception:
            pass

        journal_entry.is_posted = True
        journal_entry.is_draft = False
        if hasattr(journal_entry, 'posted_at') and not getattr(journal_entry, 'posted_at', None):
            journal_entry.posted_at = now
        if hasattr(journal_entry, 'posted_by') and not getattr(journal_entry, 'posted_by', None):
            journal_entry.posted_by = new_invoice.posted_by

        # ── Auto-post voucher JEs linked to this invoice ──
        try:
            _linked_vouchers = Voucher.query.filter_by(
                reference_type='invoice', reference_id=new_invoice.id
            ).all()
            _voucher_ids = [v.id for v in _linked_vouchers]
            if _voucher_ids:
                _voucher_jes = JournalEntry.query.filter(
                    JournalEntry.reference_type == 'voucher',
                    JournalEntry.reference_id.in_(_voucher_ids),
                ).all()
                for _vje in _voucher_jes:
                    if not _vje.is_posted:
                        _vje.is_posted = True
                        _vje.is_draft = False
                        if not getattr(_vje, 'posted_at', None):
                            _vje.posted_at = now
                        if not getattr(_vje, 'posted_by', None):
                            _vje.posted_by = new_invoice.posted_by
        except Exception as exc:
            print(f"⚠️ Auto-post voucher JEs skipped: {exc}")

        print(f"✅ Committing transaction...")

        # 📦 Category-weight tracking (by location / gold SafeBox)
        # We do this only for posted invoices to avoid counting approval-required drafts.
        try:
            record_category_weight_movements_for_invoice_payload(
                invoice_id=new_invoice.id,
                items_payload=(data.get('items') if isinstance(data, dict) else None),
            )
        except Exception as exc:
            # Do not block posting for tracking failures.
            print(f"⚠️ Category weight tracking skipped: {exc}")

        # 📒 Inventory ledger posting (Event Log — append-only)
        try:
            from services.inventory_posting_service import InventoryPostingService
            InventoryPostingService.post(new_invoice)
        except Exception as exc:
            print(f"⚠️ Inventory ledger posting skipped: {exc}")

        # Audit: large discount (sales)
        try:
            if str(invoice_type).strip() == 'بيع' and total_gross_cash > 0 and total_discount_cash > 0:
                discount_pct2 = (total_discount_cash / total_gross_cash) * 100.0
                if discount_pct2 >= float(large_discount_pct_threshold or 0.0):
                    from models import AuditLog

                    AuditLog.log_action(
                        user_name=new_invoice.posted_by or posted_by_username or 'system',
                        action='large_discount',
                        entity_type='Invoice',
                        entity_id=new_invoice.id,
                        entity_number=getattr(new_invoice, 'invoice_number', None),
                        details=json.dumps({
                            'invoice_type': str(invoice_type).strip(),
                            'discount_total_cash': round(float(total_discount_cash), 2),
                            'gross_total_cash': round(float(total_gross_cash), 2),
                            'discount_pct': round(float(discount_pct2), 2),
                            'threshold_pct': float(large_discount_pct_threshold),
                        }, ensure_ascii=False),
                        ip_address=request.remote_addr,
                        user_agent=request.headers.get('User-Agent'),
                        success=True,
                    )
        except Exception:
            pass

        # ── Sync safe-box transactions for invoice JE lines ──
        try:
            db.session.flush()
            _inv_jes = JournalEntry.query.filter_by(
                reference_type='invoice', reference_id=new_invoice.id
            ).all()
            for _ij in _inv_jes:
                _ensure_safe_box_transactions_for_invoice_je(
                    invoice_id=new_invoice.id,
                    journal_entry_id=_ij.id,
                    created_by=posted_by_username or 'system',
                )
        except Exception as exc:
            print(f"⚠️ safe-box SBT sync skipped: {exc}")

        # Recompute stored Account.balance_* for all accounts touched by this
        # invoice's posted JE lines (POS inline-post path).
        try:
            _pos_affected_ids = set()
            for _pje in JournalEntry.query.filter_by(
                reference_type='invoice', reference_id=new_invoice.id
            ).all():
                _pos_affected_ids.update(
                    l.account_id for l in (_pje.lines or []) if l.account_id
                )
            if _pos_affected_ids:
                _recalculate_account_balances_for_accounts(_pos_affected_ids)
        except Exception as _rc_exc:
            print(f"⚠️ recalculate balances after inline post skipped: {_rc_exc}")

        db.session.commit()
        try:
            created_payment_method_ids = {
                int(payment.payment_method_id)
                for payment in (getattr(new_invoice, 'payments', None) or [])
                if getattr(payment, 'payment_method_id', None) not in (None, '', False)
            }
        except Exception:
            created_payment_method_ids = set()
        _try_process_due_auto_clearing_settlements(payment_method_ids=created_payment_method_ids)
        _pos_claims_confirmed = True
        if _pos_claims:
            try:
                from services.commerce_availability import _confirm_pos_claims_best_effort
                _confirm_pos_claims_best_effort(_pos_claims)
            except Exception:
                pass
        return jsonify(new_invoice.to_dict()), 201

    except (ValueError, IntegrityError) as e:
        db.session.rollback()
        # Log the error for debugging
        print(f"Error adding invoice: {str(e)}")
        import traceback
        traceback.print_exc()
        _err_detail = str(e)
        # Expose first 600 chars of detail to help diagnose production-only failures
        return jsonify({'error': 'Failed to create invoice', 'detail': _err_detail, 'detail_short': _err_detail[:600]}), 400
    except Exception as e:
        db.session.rollback()
        print(f"An unexpected error occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        _err_detail2 = str(e)
        return jsonify({
            'error': 'An unexpected server error occurred.',
            'detail': _err_detail2,
            'detail_short': _err_detail2[:600],
            'error_type': type(e).__name__,
        }), 500
    finally:
        # Release any unclaimed pos-claims so the item is freed immediately
        # rather than waiting for TTL expiry.  Best-effort: never blocks return.
        if _pos_claims and not _pos_claims_confirmed:
            try:
                from services.commerce_availability import _release_pos_claims_best_effort
                _release_pos_claims_best_effort(_pos_claims)
            except Exception:
                pass

@invoices_bp.route('/devtools/import/sales-invoices', methods=['POST'])
@require_admin
def devtools_import_sales_invoices_from_excel():
    """Import sales invoices from an uploaded Excel file.

    Admin-only endpoint.

    Request: multipart/form-data
      - file: .xlsx
      - apply: '0'|'1' (default: 0 dry-run)
      - sheet: optional sheet name
    - as_categories: '0'|'1' (default: 0)

    Notes:
      - Dedupe is always enabled.
      - Uses the same invoice creation route logic (`add_invoice`).
      - For imports only, an admin can override invoice employee_id/posted_by
        via a guarded flag injected server-side.
    """

    import os
    import tempfile
    from datetime import datetime

    try:
        from auth_decorators import generate_token
    except Exception:
        generate_token = None

    # Local import to avoid heavy imports at module load time.
    from devtools.import_sales_invoices import (
        _read_xlsx_rows,
        _normalize_header,
        _parse_row,
        _group_invoices,
        _load_employee_map,
        _infer_payment_method_ids,
        _build_invoice_payload,
        _ensure_default_customer,
        _find_existing_invoice_ids,
        _cached_categories,
    )

    def _extract_error_message(obj) -> str:
        try:
            if isinstance(obj, dict):
                for k in ('message', 'detail', 'error'):
                    v = obj.get(k)
                    if v not in (None, '', False):
                        return str(v)
                # Common patterns: {'errors': [...]} or {'validation_errors': {...}}
                for k in ('errors', 'validation_errors'):
                    v = obj.get(k)
                    if v not in (None, '', False):
                        return str(v)
        except Exception:
            pass
        return ''

    def _safe_len(x) -> int:
        try:
            return len(x) if x is not None else 0
        except Exception:
            return 0

    def _to_bool(raw, default: bool = False) -> bool:
        if raw in (None, ''):
            return default
        s = str(raw).strip().lower()
        return s in ('1', 'true', 'yes', 'y', 'on', 'apply')

    uploaded = request.files.get('file')
    if not uploaded:
        return jsonify({'success': False, 'error': 'missing_file', 'message': 'Missing uploaded file'}), 400

    filename = (getattr(uploaded, 'filename', None) or 'sales.xlsx').strip() or 'sales.xlsx'
    if not filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'invalid_file', 'message': 'Only .xlsx files are supported'}), 400

    apply_mode = _to_bool(request.form.get('apply'), default=False)
    sheet_name = request.form.get('sheet') or None
    as_categories = _to_bool(request.form.get('as_categories'), default=False)

    if as_categories:
        try:
            _cached_categories.cache_clear()
        except Exception:
            pass

    # Persist to a temp file because the importer expects a path.
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(uploaded.read())
            tmp_path = tmp.name

        # Resolve the actual sheet name that will be read.
        actual_sheet_name = sheet_name
        try:
            from openpyxl import load_workbook as _lw
            _wb = _lw(tmp_path, read_only=True, data_only=True)
            if sheet_name and sheet_name in _wb.sheetnames:
                actual_sheet_name = sheet_name
            else:
                actual_sheet_name = _wb.sheetnames[0] if _wb.sheetnames else sheet_name
            _wb.close()
        except Exception:
            pass

        raw_rows = _read_xlsx_rows(tmp_path, sheet_name=sheet_name)
        if not raw_rows:
            return jsonify({'success': False, 'error': 'empty_file', 'message': 'No rows found in Excel sheet'}), 400

        fieldnames = [_normalize_header(h) for h in (raw_rows[0].keys() if raw_rows else [])]

        parsed = []
        errors = []
        last_date = None
        for i, rr in enumerate(raw_rows, start=1):
            try:
                pr = _parse_row(rr, fieldnames)
                if pr:
                    parsed.append(pr)
                    last_date = pr.date
            except Exception as exc:
                # Carry-forward date from previous row when Excel export omits it.
                if 'Missing date' in str(exc) and last_date is not None:
                    try:
                        rr2 = dict(rr)
                        rr2['التاريخ'] = last_date.strftime('%Y/%m/%d')
                        pr = _parse_row(rr2, fieldnames)
                        if pr:
                            parsed.append(pr)
                        continue
                    except Exception as exc2:
                        errors.append(f'row {i}: {exc2}')
                        continue

                errors.append(f'row {i}: {exc}')

        grouped = _group_invoices(parsed)
        group_keys = sorted(grouped.keys(), key=lambda x: (str(x)))

        employee_map = _load_employee_map()
        pm_ids = _infer_payment_method_ids()
        default_customer_id = _ensure_default_customer(None, 'عميل نقدي')

        # Diagnostic: resolve safe_box_ids for each payment method so we can
        # confirm what the importer will send to add_invoice.
        from devtools.import_sales_invoices import _resolve_safe_box_id_for_payment_method
        pm_safe_diag = {}
        for _pm_key, _pm_id in pm_ids.items():
            try:
                _sb = _resolve_safe_box_id_for_payment_method(_pm_id)
                pm_safe_diag[_pm_key] = {'pm_id': _pm_id, 'resolved_safe_box_id': _sb}
            except Exception:
                pm_safe_diag[_pm_key] = {'pm_id': _pm_id, 'resolved_safe_box_id': None}

        summary = {
            'success': True,
            '_code_version': 'safebox_fix_v3',
            'apply': bool(apply_mode),
            'as_categories': bool(as_categories),
            'filename': filename,
            'sheet_name': actual_sheet_name,
            'parsed_rows': len(parsed),
            'invoice_groups_total': len(group_keys),
            'row_parse_errors_count': len(errors),
            'row_parse_errors': errors[:50],
            'warnings_count': 0,
            'warnings': [],
            'groups_would_create': 0,
            'groups_skipped_existing': 0,
            'groups_skipped_missing_employee': 0,
            'created_invoices': 0,
            'created_invoice_ids': [],
            '_pm_ids': pm_ids,
            '_pm_safe_diag': pm_safe_diag,
        }

        # Create a short-lived token for internal calls.
        current_user = getattr(g, 'current_user', None)
        token = None
        if apply_mode:
            if generate_token and current_user is not None:
                try:
                    token = generate_token(current_user, expires_in_minutes=5)
                except Exception:
                    token = None

        for gk in group_keys:
            lines = grouped[gk]
            payload, warns = _build_invoice_payload(
                gk,
                lines,
                employee_map=employee_map,
                pm_ids=pm_ids,
                assume_cash_remainder=True,
                as_categories=bool(as_categories),
            )

            if payload.get('invoice_type') == 'بيع' and default_customer_id:
                payload['customer_id'] = int(default_customer_id)

            if warns:
                summary['warnings_count'] += len(warns)
                if len(summary['warnings']) < 50:
                    summary['warnings'].extend(warns[: max(0, 50 - len(summary['warnings']))])

            # Skip if employee_id missing (prevents mis-attribution)
            if not payload.get('employee_id'):
                summary['groups_skipped_missing_employee'] += 1
                continue

            # Dedupe is always on
            existing_ids = _find_existing_invoice_ids(payload)
            if existing_ids:
                summary['groups_skipped_existing'] += 1
                continue

            summary['groups_would_create'] += 1

            if not apply_mode:
                continue

            # Guarded override flag (admin only) to preserve employee attribution.
            payload['allow_employee_override'] = True
            payload['force_post'] = True  # Historical imports bypass approval gates

            headers = {}
            if token:
                headers['Authorization'] = f'Bearer {token}'

            from app import app as _app  # type: ignore
            with _app.test_request_context('/api/invoices', method='POST', json=payload, headers=headers):
                rv = add_invoice()

            status = None
            resp = None
            if isinstance(rv, tuple) and len(rv) >= 2:
                resp, status = rv[0], rv[1]
            else:
                resp, status = rv, 200

            try:
                data = resp.get_json(silent=True) if hasattr(resp, 'get_json') else None
            except Exception:
                data = None

            if int(status) >= 400:
                # Provide actionable context to the caller/UI.
                reason = _extract_error_message(data) or 'unknown_error'
                try:
                    if isinstance(data, dict) and str(data.get('error') or '') == 'missing_safe_box_for_payment_method':
                        pm_id = data.get('payment_method_id')
                        pm_name = data.get('payment_method_name')
                        extra = None
                        if pm_name not in (None, '', False):
                            extra = str(pm_name)
                            if pm_id not in (None, '', False):
                                extra = f"{extra} (id={pm_id})"
                        elif pm_id not in (None, '', False):
                            extra = f"id={pm_id}"
                        if extra:
                            reason = f"{reason} - {extra}"
                except Exception:
                    pass
                group_preview = []
                try:
                    for r in (lines or [])[:5]:
                        group_preview.append({
                            'date': getattr(r, 'date', None).isoformat() if getattr(r, 'date', None) else None,
                            'employee_code': getattr(r, 'employee_code', None),
                            'employee_name': getattr(r, 'employee_name', None),
                            'branch_name': getattr(r, 'branch_name', None),
                            'item_name': getattr(r, 'item_name', None),
                            'karat': getattr(r, 'karat', None),
                            'total_weight': getattr(r, 'total_weight', None),
                            'line_net': getattr(r, 'line_net', None),
                            'line_total': getattr(r, 'line_total', None),
                            'cash_amount': getattr(r, 'cash_amount', None),
                            'card_amount': getattr(r, 'card_amount', None),
                            'card_type': getattr(r, 'card_type', None),
                        })
                except Exception:
                    group_preview = []

                summary['success'] = False
                summary['error'] = 'invoice_create_failed'
                summary['message'] = f'Failed at group {gk} (status={status}): {reason}'
                summary['failed_group_key'] = str(gk)
                summary['failed_response'] = data
                summary['failed_reason'] = reason
                summary['failed_group_preview'] = group_preview
                try:
                    summary['failed_payload_preview'] = {
                        'invoice_type': payload.get('invoice_type'),
                        'date': payload.get('date'),
                        'total': payload.get('total'),
                        'total_tax': payload.get('total_tax'),
                        'amount_paid': payload.get('amount_paid'),
                        'employee_id': payload.get('employee_id'),
                        'customer_id': payload.get('customer_id'),
                        'items_count': _safe_len(payload.get('items') or []),
                        'payments_count': _safe_len(payload.get('payments') or []),
                        'is_posted': payload.get('is_posted'),
                        'force_post': payload.get('force_post'),
                    }
                except Exception:
                    pass
                return jsonify(summary), 400

            summary['created_invoices'] += 1
            try:
                if isinstance(data, dict) and 'invoice_id' in data:
                    if len(summary['created_invoice_ids']) < 200:
                        summary['created_invoice_ids'].append(int(data['invoice_id']))
            except Exception:
                pass

        summary['completed_at'] = datetime.now().isoformat() + 'Z'
        return jsonify(summary), (200 if summary.get('success') else 400)

    except Exception as exc:
        return jsonify({'success': False, 'error': 'import_failed', 'message': str(exc)}), 500
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

# 🆕 Endpoints للمرتجعات
@invoices_bp.route('/invoices/<int:invoice_id>/returns', methods=['GET'])
def get_invoice_returns(invoice_id):
    """
    الحصول على جميع المرتجعات المرتبطة بفاتورة معينة
    """
    invoice = Invoice.query.get_or_404(invoice_id)
    
    # الحصول على جميع المرتجعات
    returns = Invoice.query.filter_by(original_invoice_id=invoice_id).all()
    
    return jsonify({
        'original_invoice': {
            'id': invoice.id,
            'invoice_type_id': invoice.invoice_type_id,
            'invoice_type': invoice.invoice_type,
            'date': invoice.date.isoformat(),
            'total': invoice.total,
            'status': invoice.status
        },
        'returns': [r.to_dict() for r in returns],
        'total_returns': len(returns)
    })

@invoices_bp.route('/invoices/<int:invoice_id>/can-return', methods=['GET'])
def check_can_return(invoice_id):
    """
    التحقق من إمكانية إرجاع فاتورة
    """
    invoice = Invoice.query.get_or_404(invoice_id)
    
    # الفواتير التي يمكن إرجاعها
    returnable_types = ['بيع', 'شراء من عميل', 'شراء']

    invoice_type_value = (invoice.invoice_type or '').strip()
    if 'مورد' in invoice_type_value and 'شراء' in invoice_type_value and 'مرتجع' not in invoice_type_value:
        invoice_type_value = 'شراء'

    can_return = invoice_type_value in returnable_types
    
    # التحقق من المرتجعات السابقة
    existing_returns = Invoice.query.filter_by(original_invoice_id=invoice_id).all()
    total_returned = sum(r.total for r in existing_returns)
    
    return jsonify({
        'can_return': can_return,
        'invoice_type': invoice_type_value,
        'original_total': invoice.total,
        'total_returned': total_returned,
        'remaining_amount': invoice.total - total_returned,
        'existing_returns_count': len(existing_returns),
        'message': 'يمكن إرجاع هذه الفاتورة' if can_return else 'لا يمكن إرجاع هذا النوع من الفواتير'
    })

@invoices_bp.route('/invoices/returnable', methods=['GET'])
def get_returnable_invoices():
    """
    الحصول على جميع الفواتير القابلة للإرجاع.
    - إذا كان المستخدم لا يملك مطلق الصلاحية (admin أو invoices.edit_others)،
      يتم إرجاع الفواتير التي أنشأها هو فقط (posted_by == username).
    - يدعم البحث النصي (search) والفلترة بالتاريخ (date_from / date_to).
    """
    returnable_types = ['بيع', 'شراء من عميل', 'شراء']

    invoice_type_filter = request.args.get('invoice_type')
    customer_id = request.args.get('customer_id', type=int)
    supplier_id = request.args.get('supplier_id', type=int)
    search = (request.args.get('search') or '').strip()
    date_from_str = (request.args.get('date_from') or '').strip()
    date_to_str = (request.args.get('date_to') or '').strip()

    query = Invoice.query.filter(Invoice.invoice_type.in_(returnable_types))

    # --- User-based scoping ---
    _current_user = getattr(g, 'current_user', None)
    _is_admin = bool(getattr(_current_user, 'is_admin', False))
    _has_view_others = _is_admin or (
        _current_user is not None
        and callable(getattr(_current_user, 'has_permission', None))
        and _current_user.has_permission('invoices.edit_others')
    )
    if not _has_view_others and _current_user is not None:
        _username = getattr(_current_user, 'username', None)
        if _username:
            query = query.filter(
                func.lower(func.coalesce(Invoice.posted_by, '')) == _username.strip().lower()
            )
    # ---

    if invoice_type_filter:
        query = query.filter(Invoice.invoice_type == invoice_type_filter)

    if customer_id:
        query = query.filter(Invoice.customer_id == customer_id)

    if supplier_id:
        query = query.filter(Invoice.supplier_id == supplier_id)

    if search:
        like = f'%{search}%'
        search_clauses = []
        query = query.outerjoin(Customer, Invoice.customer_id == Customer.id)
        query = query.outerjoin(Supplier, Invoice.supplier_id == Supplier.id)
        search_clauses.extend([
            Customer.name.ilike(like),
            Supplier.name.ilike(like),
        ])
        try:
            search_clauses.append(Invoice.id == int(search))
        except (ValueError, TypeError):
            pass
        # البحث بالمبلغ أو الوزن
        try:
            numeric_val = float(search.replace(',', ''))
            search_clauses.append(func.abs(func.coalesce(Invoice.total, 0) - numeric_val) < 0.01)
            search_clauses.append(func.abs(func.coalesce(Invoice.total_weight, 0) - numeric_val) < 0.001)
        except (ValueError, TypeError):
            pass
        query = query.filter(or_(*search_clauses))

    if date_from_str:
        try:
            _date_from = datetime.fromisoformat(date_from_str.replace('Z', '+00:00'))
            query = query.filter(Invoice.date >= _date_from)
        except Exception:
            pass

    if date_to_str:
        try:
            _date_to = datetime.fromisoformat(date_to_str.replace('Z', '+00:00'))
            query = query.filter(Invoice.date <= _date_to)
        except Exception:
            pass

    invoices = query.order_by(Invoice.date.desc()).all()

    result = []
    for inv in invoices:
        existing_returns = Invoice.query.filter_by(original_invoice_id=inv.id).all()
        total_returned = sum(r.total for r in existing_returns)

        result.append({
            'id': inv.id,
            'invoice_type_id': inv.invoice_type_id,
            'invoice_type': inv.invoice_type,
            'date': inv.date.isoformat(),
            'total': inv.total,
            'total_returned': total_returned,
            'remaining_amount': inv.total - total_returned,
            'can_return': (inv.total - total_returned) > 0,
            'customer_name': inv.customer.name if inv.customer else None,
            'supplier_name': inv.supplier.name if inv.supplier else None,
            'items_count': len(inv.items),
        })

    return jsonify({
        'invoices': result,
        'total_count': len(result),
    })
