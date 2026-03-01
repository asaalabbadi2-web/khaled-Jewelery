#!/usr/bin/env python3
"""Import sales invoices (بيع) from an Excel-exported file (.xlsx / .csv / .tsv).

This tool is designed for the Yasar Gold POS where invoices are weight-based.
It expects the same Arabic columns you provided (exported from Excel).

Key behaviors:
- Groups rows by the *first column* (invoice group number) to form one invoice.
- Tags each invoice with `employee_id` by matching `رقم الموظف` to Employee.employee_code.
- Builds invoice `items[]` using line totals and VAT split:
    - price (selling_price) = "الإجمالي" (line total)
    - tax_amount = "الإجمالي" - "الصافي" (if both exist)
    - net is stored by backend as price - tax
- Supports cash + card payments (مدى/فيزا/ماستركارد) via the backend multi-payments feature.

Safety:
- Default is dry-run (no writes). Use --apply to actually create invoices.

Usage examples:
    # Dry-run (.xlsx recommended)
    python backend/devtools/import_sales_invoices.py --input SalesDB.xlsx

    # Apply (create invoices)
    python backend/devtools/import_sales_invoices.py --input SalesDB.xlsx --apply

    # Exported TSV/CSV is also supported
    python backend/devtools/import_sales_invoices.py --input sales.tsv --apply

Notes:
- This script calls the backend's `add_invoice()` route internally (no HTTP),
  so it will create invoice items, payments, and accounting side-effects the
  same way as the app.
- If invoice creation requires authentication in Settings, this script will
  stop with a clear error (because per-employee tagging would be blocked).
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from datetime import date as date_type
from typing import Any, Dict, Iterable, List, Optional, Tuple


# Ensure `backend/` is importable when running from repo root.
BACKEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)


def _strip(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def _normalize_header(h: str) -> str:
    h = (h or "").strip()
    h = h.replace("\ufeff", "")
    h = re.sub(r"\s+", " ", h)
    return h


def _detect_delimiter(sample_line: str) -> str:
    # Prefer tab if present (Excel TSV).
    if "\t" in sample_line:
        return "\t"
    # Otherwise fall back to comma.
    return ","


def _parse_number(raw: Any, default: float = 0.0) -> float:
    if raw in (None, "", False):
        return float(default)

    s = str(raw).strip()
    if not s:
        return float(default)

    # Remove common thousand separators.
    s = s.replace(",", "").replace("،", "").replace(" ", "")

    # Normalize Arabic-Indic digits if present.
    arabic_digits = {
        "٠": "0",
        "١": "1",
        "٢": "2",
        "٣": "3",
        "٤": "4",
        "٥": "5",
        "٦": "6",
        "٧": "7",
        "٨": "8",
        "٩": "9",
    }
    for k, v in arabic_digits.items():
        s = s.replace(k, v)

    try:
        return float(s)
    except Exception:
        return float(default)


def _parse_int(raw: Any, default: int = 0) -> int:
    try:
        v = _parse_number(raw, default=float(default))
        if v <= 0:
            return int(default)
        return int(round(v))
    except Exception:
        return int(default)


def _parse_date(raw: Any) -> datetime:
    # Excel readers may return datetime/date objects directly.
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date_type):
        return datetime(raw.year, raw.month, raw.day)

    s = _strip(raw)
    if not s:
        raise ValueError("Missing date")

    # Common Excel exports: YYYY/MM/DD or YYYY-MM-DD
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass

    # Last resort: try ISO
    try:
        return datetime.fromisoformat(s)
    except Exception as exc:
        # Also try normalizing '/' to '-'
        try:
            return datetime.fromisoformat(s.replace('/', '-'))
        except Exception:
            raise ValueError(f"Unrecognized date format: {s}") from exc


@dataclass(frozen=True)
class ParsedRow:
    group_key: str
    date: datetime
    employee_code: str
    employee_name: str
    branch_name: str
    item_name: str
    karat: int
    wage_per_gram: float
    quantity: int
    total_weight: float
    line_cost: float
    line_net: float
    line_total: float
    cash_amount: float
    card_amount: float
    card_type: str


def _normalize_employee_code(raw_code: Any) -> str:
    code = _strip(raw_code)
    if not code:
        return ""
    # If the Excel code is numeric like 102, normalize to the app's common format E-000102.
    if code.isdigit():
        return f"E-{code.zfill(6)}"
    return code


def _resolve_employee_id(employee_code_raw: Any, employee_name: str, employee_map: Dict[str, int]) -> Optional[int]:
    code = _strip(employee_code_raw)
    if code:
        # Try exact
        if code in employee_map:
            return employee_map[code]
        # Try normalized (e.g. 102 -> E-000102)
        norm = _normalize_employee_code(code)
        if norm and norm in employee_map:
            return employee_map[norm]

    # Fallback by name (best-effort)
    if employee_name:
        from models import Employee  # type: ignore

        e = Employee.query.filter(Employee.is_active == True).filter(Employee.name == employee_name).first()
        if e is not None:
            return int(e.id)

    return None


def _read_rows(path: str, delimiter: Optional[str] = None) -> List[Dict[str, str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        first = f.readline()
        if not first:
            return []
        delim = delimiter or _detect_delimiter(first)
        f.seek(0)
        reader = csv.DictReader(f, delimiter=delim)
        if not reader.fieldnames:
            return []
        reader.fieldnames = [_normalize_header(h) for h in reader.fieldnames]
        out: List[Dict[str, str]] = []
        for row in reader:
            normalized = {(_normalize_header(k)): (v if v is not None else "") for k, v in row.items()}
            out.append(normalized)
        return out


def _read_xlsx_rows(path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required to read .xlsx files. Install it then retry.") from exc

    wb = load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [_normalize_header(h) for h in header_row]
    out: List[Dict[str, Any]] = []
    for r in rows_iter:
        if r is None:
            continue
        row_dict = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
        # Drop fully-empty rows
        if not any(v not in (None, "") for v in row_dict.values()):
            continue
        out.append(row_dict)
    return out


def _extract_group_key(row: Dict[str, str], fieldnames: List[str]) -> str:
    # The first column in your sample is numeric (0/1/2/3...) and acts as invoice group.
    # When exported, its header might be "0" or empty.
    if fieldnames:
        first_key = fieldnames[0]
        candidate = _strip(row.get(first_key))
        if candidate:
            return candidate

    # Fallbacks if user renamed columns.
    for k in ("رقم الفاتورة", "invoice", "invoice_no", "invoice_id", "group", "المجموعة"):
        candidate = _strip(row.get(k))
        if candidate:
            return candidate

    raise ValueError("Missing invoice group key (first column)")


def _parse_row(row: Dict[str, str], fieldnames: List[str]) -> Optional[ParsedRow]:
    # Skip fully-empty lines
    if not any(_strip(v) for v in row.values()):
        return None

    group_key = _extract_group_key(row, fieldnames)

    date = _parse_date(row.get("التاريخ"))
    employee_name = _strip(row.get("اسم الموظف"))
    employee_code = _strip(row.get("رقم الموظف"))
    branch_name = _strip(row.get("الفرع"))

    item_name = _strip(row.get("الصنف"))
    karat = _parse_int(row.get("العيار"), default=0)
    wage_per_gram = _parse_number(row.get("أجرة الجرام"), default=0.0)
    quantity = _parse_int(row.get("العدد"), default=1) or 1
    total_weight = _parse_number(row.get("الوزن"), default=0.0)

    line_cost = _parse_number(row.get("التكلفة"), default=0.0)
    line_net = _parse_number(row.get("الصافي"), default=0.0)
    line_total = _parse_number(row.get("الإجمالي"), default=0.0)

    cash_amount = _parse_number(row.get("النقد"), default=0.0)
    card_amount = _parse_number(row.get("الشبكة"), default=0.0)
    card_type = _strip(row.get("نوع الشبكة"))

    # Skip rows that don't look like invoice lines.
    # (Some exports include separators or blank item rows.)
    if not item_name and line_total <= 0 and total_weight <= 0:
        return None

    return ParsedRow(
        group_key=str(group_key).strip(),
        date=date,
        employee_code=employee_code,
        employee_name=employee_name,
        branch_name=branch_name,
        item_name=item_name or "صنف",
        karat=karat,
        wage_per_gram=wage_per_gram,
        quantity=quantity,
        total_weight=total_weight,
        line_cost=line_cost,
        line_net=line_net,
        line_total=line_total,
        cash_amount=cash_amount,
        card_amount=card_amount,
        card_type=card_type,
    )


def _group_invoices(rows: List[ParsedRow]) -> Dict[str, List[ParsedRow]]:
    grouped: Dict[str, List[ParsedRow]] = defaultdict(list)
    for r in rows:
        grouped[r.group_key].append(r)
    return dict(grouped)


def _infer_payment_method_ids() -> Dict[str, int]:
    """Resolve payment method IDs from the DB.

    We match by payment_type first (preferred), otherwise by Arabic name keywords.

    Returns keys:
      - cash
      - mada
      - visa
      - mastercard
      - card_default
    """

    from models import PaymentMethod  # type: ignore

    methods = PaymentMethod.query.filter_by(is_active=True).all()

    def pick(pred) -> Optional[int]:
        for m in methods:
            try:
                if pred(m):
                    return int(m.id)
            except Exception:
                continue
        return None

    def name_contains(m: Any, needle: str) -> bool:
        try:
            return needle in str(getattr(m, "name", "") or "")
        except Exception:
            return False

    def type_is(m: Any, t: str) -> bool:
        try:
            return str(getattr(m, "payment_type", "") or "").strip().lower() == t
        except Exception:
            return False

    cash_id = pick(lambda m: type_is(m, "cash") or name_contains(m, "نقد"))
    mada_id = pick(lambda m: type_is(m, "mada") or name_contains(m, "مدى"))
    visa_id = pick(lambda m: name_contains(m, "فيزا") or name_contains(m, "VISA") or name_contains(m, "Visa"))
    mc_id = pick(lambda m: name_contains(m, "ماستر") or name_contains(m, "Master") or name_contains(m, "ماستركارد"))

    card_default = pick(lambda m: type_is(m, "mada") or type_is(m, "visa") or type_is(m, "mastercard") or name_contains(m, "بطاقة"))

    out: Dict[str, int] = {}
    if cash_id:
        out["cash"] = cash_id
    if mada_id:
        out["mada"] = mada_id
    if visa_id:
        out["visa"] = visa_id
    if mc_id:
        out["mastercard"] = mc_id
    if card_default:
        out["card_default"] = card_default

    return out


def _load_employee_map() -> Dict[str, int]:
    from models import Employee  # type: ignore

    employees = Employee.query.filter_by(is_active=True).all()
    mapping: Dict[str, int] = {}
    for e in employees:
        code = _strip(getattr(e, "employee_code", ""))
        if code:
            mapping[code] = int(e.id)
    return mapping


def _check_auth_required() -> bool:
    # Mirror backend/routes.py logic:
    # - default from config REQUIRE_AUTH_FOR_INVOICE_CREATE
    # - override from Settings.require_auth_for_invoice_create if row exists
    from config import REQUIRE_AUTH_FOR_INVOICE_CREATE  # type: ignore
    from models import Settings  # type: ignore

    auth_required = bool(REQUIRE_AUTH_FOR_INVOICE_CREATE)
    try:
        settings = Settings.query.first()
        if settings is not None:
            auth_required = bool(getattr(settings, "require_auth_for_invoice_create", False))
    except Exception:
        pass

    return auth_required


def _build_invoice_payload(group_key: str, lines: List[ParsedRow], employee_map: Dict[str, int], pm_ids: Dict[str, int], assume_cash_remainder: bool) -> Tuple[Dict[str, Any], List[str]]:
    warnings: List[str] = []

    # Basic consistency checks
    date = lines[0].date
    emp_code = lines[0].employee_code
    emp_name = lines[0].employee_name

    for ln in lines[1:]:
        if ln.employee_code and emp_code and ln.employee_code != emp_code:
            warnings.append(f"group {group_key}: mixed employee_code ({emp_code} vs {ln.employee_code})")
        if ln.date.date() != date.date():
            warnings.append(f"group {group_key}: mixed dates ({date.date()} vs {ln.date.date()})")

    employee_id = _resolve_employee_id(emp_code, emp_name, employee_map)
    if not employee_id:
        warnings.append(f"group {group_key}: employee_code not found: {emp_code} ({emp_name})")

    items: List[Dict[str, Any]] = []

    # Heuristic: some Excel reports repeat invoice-level totals on every item row.
    line_totals = [round(float(ln.line_total or 0.0), 2) for ln in lines if float(ln.line_total or 0.0) > 0]
    line_nets = [round(float(ln.line_net or 0.0), 2) for ln in lines if float(ln.line_net or 0.0) > 0]
    totals_invoice_level = len(line_totals) > 1 and len(set(line_totals)) == 1
    nets_invoice_level = len(line_nets) > 1 and len(set(line_nets)) == 1

    invoice_total_value = line_totals[0] if line_totals else 0.0
    invoice_net_value = line_nets[0] if line_nets else 0.0

    inv_total = invoice_total_value if totals_invoice_level else 0.0
    inv_tax = (
        max(0.0, round(invoice_total_value - invoice_net_value, 2))
        if (totals_invoice_level and nets_invoice_level and invoice_total_value > 0 and invoice_net_value > 0)
        else 0.0
    )
    inv_cost = 0.0

    # Allocation weights (for invoice-level totals): proportional by total_weight.
    weight_sum = sum(max(0.0, float(ln.total_weight or 0.0)) for ln in lines)
    if totals_invoice_level and weight_sum <= 0:
        warnings.append(f"group {group_key}: invoice totals appear invoice-level, but weights are all zero; allocating equally")

    allocated_line_totals: List[float] = []
    allocated_line_taxes: List[float] = []

    for ln in lines:
        qty = ln.quantity if ln.quantity > 0 else 1
        total_w = max(0.0, float(ln.total_weight or 0.0))
        weight_per_item = round(total_w / qty, 6) if (total_w > 0 and qty > 0) else 0.0

        raw_line_total = float(ln.line_total or 0.0)
        raw_line_net = float(ln.line_net or 0.0)

        if totals_invoice_level:
            share = (total_w / weight_sum) if weight_sum > 0 else (1.0 / max(1, len(lines)))
            line_total = round(inv_total * share, 2)
        else:
            line_total = raw_line_total

        if totals_invoice_level and inv_tax > 0:
            share = (total_w / weight_sum) if weight_sum > 0 else (1.0 / max(1, len(lines)))
            line_tax = round(inv_tax * share, 2)
        else:
            # Derive tax from raw net/total if available.
            line_tax = 0.0
            if raw_line_total > 0 and raw_line_net > 0:
                line_tax = max(0.0, round(raw_line_total - raw_line_net, 2))
            elif raw_line_total > 0 and raw_line_net <= 0:
                assumed_net = round(raw_line_total / 1.15, 2)
                line_tax = max(0.0, round(raw_line_total - assumed_net, 2))

        if not totals_invoice_level:
            inv_total += line_total
            inv_tax += line_tax

        inv_cost += max(0.0, float(ln.line_cost or 0.0))

        allocated_line_totals.append(round(line_total, 2))
        allocated_line_taxes.append(round(line_tax, 2))

        price_per_item = round((line_total / qty), 2) if (qty > 0 and line_total > 0) else 0.0
        tax_per_item = round((line_tax / qty), 2) if (qty > 0 and line_tax > 0) else 0.0
        wage_per_item = round(float(ln.wage_per_gram or 0.0) * weight_per_item, 2) if weight_per_item > 0 else 0.0

        items.append(
            {
                "name": ln.item_name,
                "karat": ln.karat if ln.karat > 0 else 21,
                "weight": weight_per_item,
                "quantity": qty,
                # Backend multiplies price*quantity, so this must be per-item.
                "selling_price": price_per_item,
                # Backend stores tax per item (and totals via *quantity).
                "tax_amount": tax_per_item,
                "discount_amount": 0.0,
                "wage": wage_per_item,
                "manufacturing_wage_per_gram": float(ln.wage_per_gram or 0.0),
            }
        )

    inv_total = round(inv_total, 2)
    inv_tax = round(inv_tax, 2)
    inv_cost = round(inv_cost, 2)

    # Fix rounding drift for invoice-level allocation so item totals match invoice total.
    if totals_invoice_level and items:
        computed = round(sum(float(it.get("selling_price") or 0.0) * float(it.get("quantity") or 1) for it in items), 2)
        drift = round(inv_total - computed, 2)
        if abs(drift) >= 0.01:
            last = items[-1]
            q = float(last.get("quantity") or 1)
            if q <= 0:
                q = 1
            last["selling_price"] = round(float(last.get("selling_price") or 0.0) + (drift / q), 2)

    # Payments: many exports repeat invoice-level payment fields on every line.
    # Using max() is safer than sum() in that case.
    cash_total = round(max((max(0.0, float(ln.cash_amount or 0.0)) for ln in lines), default=0.0), 2)
    card_by_type: Dict[str, float] = {}
    for ln in lines:
        amt = round(max(0.0, float(ln.card_amount or 0.0)), 2)
        if amt <= 0:
            continue
        ct = ln.card_type.strip() if ln.card_type else ""
        ct = ct or "card"
        # Keep the maximum amount seen for this type (avoids duplicate repeats).
        card_by_type[ct] = max(float(card_by_type.get(ct) or 0.0), float(amt))

    payments: List[Dict[str, Any]] = []

    if cash_total > 0:
        cash_pm = pm_ids.get("cash")
        if not cash_pm:
            warnings.append(f"group {group_key}: cash payment present but no cash payment method found")
        else:
            payments.append({"payment_method_id": cash_pm, "amount": cash_total})

    for card_type, amt in sorted(card_by_type.items(), key=lambda x: x[0]):
        amt = round(amt, 2)
        if amt <= 0:
            continue

        key = "card_default"
        if "مدى" in card_type:
            key = "mada"
        elif "فيزا" in card_type or "visa" in card_type.lower():
            key = "visa"
        elif "ماستر" in card_type or "master" in card_type.lower():
            key = "mastercard"

        pm_id = pm_ids.get(key) or pm_ids.get("card_default")
        if not pm_id:
            warnings.append(f"group {group_key}: card payment ({card_type}) but no card payment method found")
            continue

        payments.append({"payment_method_id": pm_id, "amount": amt})

    # If there are payments but rounding differences exist, adjust cash (or add cash remainder).
    payments_sum = round(sum(float(p.get("amount") or 0.0) for p in payments), 2)

    # Never allow payments to exceed invoice total (can happen due to rounding in source export).
    if payments and payments_sum > inv_total:
        over = round(payments_sum - inv_total, 2)
        if over > 0:
            cash_pm = pm_ids.get("cash")
            adjusted = False
            if cash_pm:
                for p in payments:
                    if int(p.get("payment_method_id") or 0) == int(cash_pm):
                        p["amount"] = round(max(0.0, float(p.get("amount") or 0.0) - over), 2)
                        adjusted = True
                        break
            if not adjusted:
                # Fallback: reduce the last payment.
                payments[-1]["amount"] = round(max(0.0, float(payments[-1].get("amount") or 0.0) - over), 2)

            payments[:] = [p for p in payments if round(float(p.get("amount") or 0.0), 2) > 0]
            payments_sum = round(sum(float(p.get("amount") or 0.0) for p in payments), 2)

    diff = round(inv_total - payments_sum, 2)

    if payments and abs(diff) >= 0.01:
        if assume_cash_remainder:
            cash_pm = pm_ids.get("cash")
            if cash_pm:
                if diff > 0:
                    # Add missing remainder to cash.
                    payments.append({"payment_method_id": cash_pm, "amount": diff})
                    payments_sum = round(payments_sum + diff, 2)
                else:
                    # Reduce the first cash payment if possible.
                    for p in payments:
                        if int(p.get("payment_method_id") or 0) == int(cash_pm):
                            new_amt = round(float(p.get("amount") or 0.0) + diff, 2)
                            if new_amt >= 0:
                                p["amount"] = new_amt
                                payments_sum = round(payments_sum + diff, 2)
                            break
        else:
            warnings.append(f"group {group_key}: payments sum {payments_sum} != invoice total {inv_total}")

    payload: Dict[str, Any] = {
        "invoice_type": "بيع",
        "gold_type": "new",
        "date": date.replace(hour=0, minute=0, second=0, microsecond=0).isoformat(),
        "total": inv_total,
        "total_tax": inv_tax,
        "total_cost": inv_cost if inv_cost > 0 else 0.0,
        "amount_paid": inv_total if (not payments or abs(inv_total - payments_sum) <= 0.01) else payments_sum,
        "employee_id": employee_id,
        "posted_by": emp_name or None,
        "items": items,
    }

    if payments:
        payload["payments"] = payments

    return payload, warnings


def _ensure_default_customer(default_customer_id: Optional[int], default_customer_name: str) -> Optional[int]:
    if default_customer_id not in (None, '', 0, '0', False):
        try:
            return int(default_customer_id)
        except Exception:
            return None

    name = (default_customer_name or '').strip() or 'عميل نقدي'
    from models import db, Customer  # type: ignore
    from code_generator import generate_customer_code  # type: ignore

    existing = Customer.query.filter(Customer.active == True).filter(Customer.name == name).first()
    if existing is not None:
        return int(existing.id)

    c = Customer(customer_code=generate_customer_code(), name=name, active=True)
    db.session.add(c)
    db.session.flush()
    return int(c.id)


def _find_existing_invoice_ids(
    payload: Dict[str, Any],
    total_tolerance: float = 0.01,
    weight_tolerance: float = 0.001,
) -> List[int]:
    """Best-effort dedupe check.

        We look for an existing posted/unposted sales invoice with the same:
      - invoice_type
      - date (day-level ISO normalized by this importer)
      - employee_id (if present)
      - customer_id (if present)
            - total (within tolerance)

        Additionally (when payload items are available), we also compare:
            - total_weight (derived from payload items vs DB invoice.total_weight)
            - items signature (name/karat/weight/qty/price/tax)

        This extra weight check prevents false-positive dedupe collisions when two
        different invoices share the same day/total/employee/customer.

    This is intentionally conservative to avoid creating duplicates when resuming.
    """

    try:
        from models import Invoice  # type: ignore
    except Exception:
        return []

    try:
        inv_type = _strip(payload.get('invoice_type') or '').strip() or 'بيع'
        inv_date_raw = payload.get('date')
        inv_date = datetime.fromisoformat(str(inv_date_raw))
        total = float(payload.get('total') or 0.0)
    except Exception:
        return []

    payload_weight: Optional[float] = None
    payload_item_sig: Optional[List[Tuple[Any, ...]]] = None
    try:
        items = payload.get('items') or []
        if isinstance(items, list) and items:
            w_sum = 0.0
            sig: List[Tuple[Any, ...]] = []
            for it in items:
                if not isinstance(it, dict):
                    continue
                w = float(it.get('weight') or 0.0)
                q = float(it.get('quantity') or 1.0)
                if q <= 0:
                    q = 1.0
                w_sum += max(0.0, w) * q

                name = str(it.get('name') or '').strip()
                karat = float(it.get('karat') or 0.0)
                price = float(it.get('selling_price') or 0.0)
                tax = float(it.get('tax_amount') or 0.0)
                sig.append(
                    (
                        name,
                        round(karat, 3),
                        round(float(w or 0.0), 6),
                        int(round(q)),
                        round(price, 2),
                        round(tax, 2),
                    )
                )
            if w_sum > 0:
                payload_weight = float(w_sum)
            if sig:
                payload_item_sig = sorted(sig)
    except Exception:
        payload_weight = None
        payload_item_sig = None

    q = Invoice.query.filter(Invoice.invoice_type == inv_type).filter(Invoice.date == inv_date)

    # Filter by employee/customer when available.
    try:
        if payload.get('employee_id') not in (None, '', False):
            q = q.filter(Invoice.employee_id == int(payload.get('employee_id')))
    except Exception:
        pass

    try:
        if payload.get('customer_id') not in (None, '', False):
            q = q.filter(Invoice.customer_id == int(payload.get('customer_id')))
    except Exception:
        pass

    try:
        cands = q.all()
    except Exception:
        return []

    matches: List[int] = []
    for inv in cands:
        try:
            if abs(float(getattr(inv, 'total', 0.0) or 0.0) - total) > float(total_tolerance):
                continue

            if payload_weight is not None:
                inv_w = float(getattr(inv, 'total_weight', 0.0) or 0.0)
                if abs(inv_w - float(payload_weight)) > float(weight_tolerance):
                    continue

            if payload_item_sig is not None:
                inv_sig: List[Tuple[Any, ...]] = []
                try:
                    inv_items = getattr(inv, 'items', None) or []
                    for ii in inv_items:
                        inv_sig.append(
                            (
                                str(getattr(ii, 'name', '') or '').strip(),
                                round(float(getattr(ii, 'karat', 0.0) or 0.0), 3),
                                round(float(getattr(ii, 'weight', 0.0) or 0.0), 6),
                                int(getattr(ii, 'quantity', 1) or 1),
                                round(float(getattr(ii, 'price', 0.0) or 0.0), 2),
                                round(float(getattr(ii, 'tax', 0.0) or 0.0), 2),
                            )
                        )
                except Exception:
                    inv_sig = []

                if sorted(inv_sig) != payload_item_sig:
                    continue

            matches.append(int(inv.id))
        except Exception:
            continue

    return matches


def main() -> int:
    parser = argparse.ArgumentParser(description="Import sales invoices from Excel-exported CSV/TSV")
    parser.add_argument("--input", required=True, help="Path to .xlsx, .csv, or .tsv file")
    parser.add_argument("--delimiter", default=None, help="Optional delimiter override (e.g. '\\t' or ',')")
    parser.add_argument("--sheet", default=None, help="Excel sheet name (only for .xlsx)")
    parser.add_argument("--apply", action="store_true", help="Actually create invoices (default: dry-run)")
    parser.add_argument("--skip-groups", type=int, default=0, help="Skip first N invoice groups (resume support)")
    parser.add_argument("--min-group", type=int, default=0, help="Only process groups with numeric key >= this")
    parser.add_argument("--max-group", type=int, default=0, help="Only process groups with numeric key <= this")
    parser.add_argument("--limit", type=int, default=0, help="Only process first N invoice groups")
    parser.add_argument(
        "--no-assume-cash-remainder",
        action="store_true",
        help="Don't auto-fix payment differences by adjusting cash (not recommended)",
    )
    parser.add_argument(
        "--create-missing-employees",
        action="store_true",
        help="Create missing employees based on employee code/name in the file (dry-run won't commit)",
    )
    parser.add_argument(
        "--default-customer-id",
        default=None,
        help="Customer ID to attach to sales invoices (needed to create payment vouchers)",
    )
    parser.add_argument(
        "--default-customer-name",
        default="عميل نقدي",
        help="Customer name to auto-create/find and attach to sales invoices when customer_id is missing",
    )

    parser.add_argument(
        "--no-dedupe",
        action="store_true",
        help="Disable dedupe check (NOT recommended). By default the importer skips invoices that already exist (same date/total/employee/customer).",
    )

    args = parser.parse_args()

    from app import app  # type: ignore
    import routes as api_routes  # type: ignore

    from models import db, Employee  # type: ignore

    with app.app_context():
        if _check_auth_required():
            print(
                "ERROR: Settings.require_auth_for_invoice_create is enabled.\n"
                "This importer needs unauthenticated invoice creation so it can set employee_id from the payload.\n"
                "Disable it temporarily (Settings -> require_auth_for_invoice_create = false), then retry.",
                file=sys.stderr,
            )
            return 2

        if str(args.input).lower().endswith(".xlsx"):
            raw_rows = _read_xlsx_rows(args.input, sheet_name=args.sheet)
        else:
            raw_rows = _read_rows(args.input, delimiter=args.delimiter)
        if not raw_rows:
            print("No rows found.", file=sys.stderr)
            return 2

        fieldnames = [_normalize_header(h) for h in (raw_rows[0].keys() if raw_rows else [])]

        parsed: List[ParsedRow] = []
        errors: List[str] = []
        last_date: Optional[datetime] = None
        for i, rr in enumerate(raw_rows, start=1):
            try:
                pr = _parse_row(rr, fieldnames)
                if pr:
                    parsed.append(pr)
                    last_date = pr.date
            except Exception as exc:
                # Some Excel exports leave date empty on subsequent lines; carry-forward.
                if "Missing date" in str(exc) and last_date is not None:
                    try:
                        rr = dict(rr)
                        rr["التاريخ"] = last_date.strftime("%Y/%m/%d")
                        pr = _parse_row(rr, fieldnames)
                        if pr:
                            parsed.append(pr)
                        continue
                    except Exception as exc2:
                        errors.append(f"row {i}: {exc2}")
                        continue

                errors.append(f"row {i}: {exc}")

        grouped = _group_invoices(parsed)
        group_keys = sorted(grouped.keys(), key=lambda x: (_parse_int(x, 0), x))

        if args.min_group and args.min_group > 0:
            group_keys = [k for k in group_keys if _parse_int(k, 0) >= int(args.min_group)]
        if args.max_group and args.max_group > 0:
            group_keys = [k for k in group_keys if _parse_int(k, 0) <= int(args.max_group)]

        if args.skip_groups and args.skip_groups > 0:
            group_keys = group_keys[args.skip_groups :]

        if args.limit and args.limit > 0:
            group_keys = group_keys[: args.limit]

        employee_map = _load_employee_map()
        pm_ids = _infer_payment_method_ids()

        # Payment vouchers for sales invoices require a party (customer/supplier).
        default_customer_id = _ensure_default_customer(args.default_customer_id, args.default_customer_name)

        created_employees = 0
        if args.create_missing_employees:
            # Pre-create employees for codes present in the file to enable employee_id tagging.
            # In dry-run, we won't commit; in apply mode, they'll be persisted (either via invoice commits or final commit).
            wanted: Dict[str, str] = {}
            for r in parsed:
                raw = _strip(r.employee_code)
                if not raw:
                    continue
                norm = _normalize_employee_code(raw)
                code_candidate = norm or raw
                if code_candidate and code_candidate not in wanted:
                    wanted[code_candidate] = _strip(r.employee_name) or code_candidate

            missing = [(code, name) for code, name in wanted.items() if code not in employee_map]
            for code, name in missing:
                if Employee.query.filter_by(employee_code=code).first():
                    continue
                e = Employee(employee_code=code, name=name, is_active=True, created_by="import_sales")
                db.session.add(e)
                created_employees += 1

            if created_employees:
                db.session.flush()
                employee_map = _load_employee_map()

        print(f"Parsed rows: {len(parsed)}")
        print(f"Invoice groups: {len(group_keys)}")
        if errors:
            print(f"Row parse errors: {len(errors)} (showing first 10)")
            for e in errors[:10]:
                print(f"  - {e}")

        created = 0
        skipped = 0
        total_warnings = 0

        for gk in group_keys:
            lines = grouped[gk]
            payload, warns = _build_invoice_payload(
                gk,
                lines,
                employee_map=employee_map,
                pm_ids=pm_ids,
                assume_cash_remainder=not bool(args.no_assume_cash_remainder),
            )

            if payload.get('invoice_type') == 'بيع' and default_customer_id:
                payload['customer_id'] = int(default_customer_id)

            if warns:
                total_warnings += len(warns)
                print(f"WARN group {gk}: {warns[0]}")

            if not args.apply:
                continue

            # Skip if employee_id missing (strong default; prevents mis-attribution)
            if not payload.get("employee_id"):
                skipped += 1
                continue

            # Best-effort dedupe: avoid creating duplicates when resuming.
            if not bool(args.no_dedupe):
                existing_ids = _find_existing_invoice_ids(payload)
                if existing_ids:
                    print(f"SKIP group {gk}: already imported (invoice_id={existing_ids[0]})")
                    continue

            # Call the backend route internally.
            with app.test_request_context("/api/invoices", method="POST", json=payload):
                rv = api_routes.add_invoice()

            # Unpack (response, status)
            status = None
            resp = None
            if isinstance(rv, tuple) and len(rv) >= 2:
                resp, status = rv[0], rv[1]
            else:
                resp, status = rv, 200

            try:
                data = resp.get_json(silent=True) if hasattr(resp, "get_json") else None
            except Exception:
                data = None

            if int(status) >= 400:
                print(f"ERROR group {gk}: status={status} response={data}")
                return 3

            created += 1

        if args.apply:
            print(f"Created invoices: {created}")
            print(f"Skipped (missing employee): {skipped}")
        else:
            print("Dry-run only (no invoices created). Use --apply to import.")

        if created_employees:
            if args.apply:
                try:
                    db.session.commit()
                except Exception:
                    pass
                print(f"Created employees: {created_employees}")
            else:
                db.session.rollback()
                print(f"Dry-run: would create employees: {created_employees}")

        if total_warnings:
            print(f"Total warnings: {total_warnings}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
